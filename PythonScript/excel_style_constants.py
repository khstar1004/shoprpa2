from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# --- Styling Constants ---
HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green fill
HEADER_FONT = Font(bold=True, size=11, name='맑은 고딕')
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define alignments based on column type
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center", wrap_text=False)  # Numbers right-aligned

DEFAULT_FONT = Font(name='맑은 고딕', size=10)

THIN_BORDER_SIDE = Side(style='thin')
DEFAULT_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

LINK_FONT = Font(color="0000FF", underline="single", name='맑은 고딕', size=10)
INVALID_LINK_FONT = Font(color="FF0000", name='맑은 고딕', size=10)  # Red for invalid links

NEGATIVE_PRICE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill for negative diff < -1

# --- Result File Specific Constants ---
RESULT_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
RESULT_HEADER_HEIGHT = 30
RESULT_DATA_ROW_HEIGHT = 400  # Increased from 380 for better image display
RESULT_ROW_HEIGHT = 400  # Same as RESULT_DATA_ROW_HEIGHT for consistency
RESULT_IMAGE_CELL_WIDTH = 85   # Increased from 80
RESULT_IMAGE_WIDTH = 160       # Increased from 80
RESULT_IMAGE_HEIGHT = 160      # Increased from 80

# --- Upload File Specific Constants ---
UPLOAD_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray background
UPLOAD_HEADER_HEIGHT = 34.5
UPLOAD_DATA_ROW_HEIGHT = 16.9
UPLOAD_COLUMN_DEFAULT_WIDTH = 7

# Column width settings for different column types
COLUMN_WIDTH_SETTINGS = {
    'image': 85,          # Image columns - increased from 21.44
    'name': 45,          # Product name
    'link': 35,          # URLs and links
    'price': 14,         # Price columns
    'percent': 10,       # Percentage columns
    'quantity': 10,      # Quantity columns
    'code': 12,          # Code columns
    'category': 20,      # Category columns
    'text_short': 7,     # Short text columns (구분, 담당자 등)
    'default': 15        # Default width
}

# Special width settings for upload file
UPLOAD_COLUMN_WIDTHS = {
    '상품명': 35,
    '상품코드': 12,
    '카테고리(중분류)': 15,
    '해오름(이미지링크)': 40,
    '고려기프트(이미지링크)': 40,
    '네이버쇼핑(이미지링크)': 40,
    '본사링크': 30,
    '고려 링크': 30,
    '네이버 링크': 30
}

# Image processing constants
IMAGE_SETTINGS = {
    'MAX_SIZE': (2000, 2000),  # Maximum supported image size
    'STANDARD_SIZE': (160, 160),  # Standard display size - increased from 80x80
    'QUALITY': 85,  # JPEG compression quality
    'SUPPORTED_FORMATS': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'],
    'NOBG_SUFFIX': '_nobg.png'  # Suffix for transparent background images
}

# Cell style combinations for different file types
RESULT_FILE_STYLES = {
    'header': {
        'fill': RESULT_HEADER_FILL,
        'font': HEADER_FONT,
        'alignment': HEADER_ALIGNMENT,
        'border': DEFAULT_BORDER
    },
    'data': {
        'font': DEFAULT_FONT,
        'border': DEFAULT_BORDER
    }
}

UPLOAD_FILE_STYLES = {
    'header': {
        'fill': UPLOAD_HEADER_FILL,
        'font': Font(bold=True, size=10, name='맑은 고딕'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': DEFAULT_BORDER
    },
    'data': {
        'font': DEFAULT_FONT,
        'border': DEFAULT_BORDER,
        'alignment': Alignment(vertical='center', wrap_text=True)
    }
}

# Image cell specific styling
IMAGE_CELL_HEIGHT = 420  # Increased from 360 for larger images
IMAGE_CELL_WIDTH = 60   # Increased from 44 for wider image cells

# Error Messages Constants
ERROR_MESSAGES = {
    'no_match': '가격 범위내에 없거나 텍스트 유사율을 가진 상품이 없음',
    'no_price_match': '가격이 범위내에 없거나 검색된 상품이 없음',
    'low_similarity': '일정 정확도 이상의 텍스트 유사율을 가진 상품이 없음',
    'no_results': '검색 결과 0',
    'no_image': '이미지를 찾을 수 없음',
    'file_not_found': '-이미지 없음-',
    'invalid_image': '유효하지 않은 이미지 형식',
    'processing_error': '-처리 오류-',
    'too_small': '이미지 크기가 너무 작음 (저해상도)',
    'format_error': '지원하지 않는 이미지 형식',
    'download_failed': '이미지 다운로드 실패',
    'excel_limit': '이미지 크기가 Excel 제한을 초과함'
} 