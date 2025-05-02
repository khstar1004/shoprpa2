#!/usr/bin/env python3
"""
고려기프트 이미지 URL 처리 문제 해결 스크립트
"""

import os
import pandas as pd
import logging
import re
import json
import shutil
from pathlib import Path
import sys
import traceback
import requests
from urllib.parse import urlparse
import hashlib
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Set, Tuple, Any
import configparser

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('kogift_image_fix.log')
    ]
)
logger = logging.getLogger('kogift_image_fix')

def get_config():
    """Load configuration from config.ini file"""
    config = configparser.ConfigParser()
    config_path = Path(__file__).resolve().parent.parent / 'config.ini'
    
    try:
        config.read(config_path, encoding='utf-8')
        logger.info(f"Successfully loaded configuration from {config_path}")
    except Exception as e:
        logger.error(f"Error loading config from {config_path}: {e}")
        # Create minimal default config
        config['Paths'] = {
            'image_main_dir': 'C:\\RPA\\Image\\Main',
            'output_dir': 'C:\\RPA\\Output'
        }
    
    return config

def download_image(url, save_path):
    """Download an image from URL to the specified path"""
    try:
        response = requests.get(url, stream=True, timeout=10)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                for chunk in response.iter_content(1024):
                    f.write(chunk)
            logging.info(f"Downloaded image from {url} to {save_path}")
            return True
        else:
            logging.error(f"Failed to download image. Status code: {response.status_code}")
            return False
    except Exception as e:
        logging.error(f"Error downloading image: {e}")
        return False

def scan_kogift_images(base_dirs=None) -> Dict[str, str]:
    """
    Scan all potential Kogift image directories and build a comprehensive mapping
    of image filenames to their local paths.
    
    Args:
        base_dirs: Optional list of base directories to scan. If None, default directories will be used.
        
    Returns:
        Dictionary mapping filenames to full local paths for all Kogift images found
    """
    # Get config for default directories
    config = get_config()
    base_img_dir = config.get('Paths', 'image_main_dir', fallback='C:\\RPA\\Image\\Main')
    
    # Default Kogift image directory candidates
    if base_dirs is None:
        base_dirs = [
            Path(base_img_dir),
            Path(base_img_dir) / 'Kogift',
            Path(base_img_dir) / 'kogift',
            Path(base_img_dir).parent / 'Kogift',
            Path(base_img_dir).parent / 'kogift',
            Path(base_img_dir).parent / 'Target' / 'Kogift',
            Path(base_img_dir).parent / 'Target' / 'kogift',
            Path('C:\\RPA\\Image\\Main\\Kogift'),
            Path('C:\\RPA\\Image\\Main\\kogift'),
            Path('C:\\RPA\\Image\\Kogift'),
            Path('C:\\RPA\\Image\\kogift')
        ]
    
    # Dictionary to store all found Kogift images
    kogift_images = {}
    total_images = 0
    
    logger.info("Starting scan of Kogift image directories...")
    
    # Scan all directories
    for base_dir in base_dirs:
        if not base_dir.exists():
            logger.debug(f"Directory does not exist: {base_dir}")
            continue
        
        logger.info(f"Scanning directory: {base_dir}")
        
        # Scan recursively for image files
        try:
            for root, _, files in os.walk(base_dir):
                for file in files:
                    if file.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        full_path = os.path.join(root, file)
                        file_size = os.path.getsize(full_path)
                        
                        # Skip empty or very small files
                        if file_size < 1000:  # Less than 1KB
                            continue
                        
                        # Store image by various key combinations for flexible matching
                        # 1. Store by full base filename (case preserved)
                        base_name = os.path.basename(file)
                        kogift_images[base_name] = full_path
                        
                        # 2. Store by lowercase filename
                        kogift_images[base_name.lower()] = full_path
                        
                        # 3. For filenames with kogift_ prefix
                        if base_name.lower().startswith('kogift_'):
                            # 3a. Store without the prefix
                            no_prefix = base_name[7:]  # Remove 'kogift_'
                            kogift_images[no_prefix] = full_path
                            kogift_images[no_prefix.lower()] = full_path
                            
                            # 3b. Try to extract a hash part if present
                            # Pattern: kogift_PRODUCTNAME_HASH.ext
                            hash_match = re.search(r'kogift_.*?_([a-f0-9]{8,})\.', base_name.lower())
                            if hash_match:
                                hash_val = hash_match.group(1)
                                # Store hash patterns
                                kogift_images[f"kogift_{hash_val}.jpg"] = full_path
                                kogift_images[f"kogift_{hash_val}.png"] = full_path
                                kogift_images[hash_val] = full_path
                                
                            # 3c. Check if the file has _nobg suffix (new)
                            if '_nobg' in base_name.lower():
                                # Extract the part before _nobg
                                name_without_nobg = re.sub(r'_nobg\.[^.]+$', '', base_name)
                                # Store alternative versions for mapping
                                kogift_images[f"{name_without_nobg}.jpg"] = full_path
                                kogift_images[f"{name_without_nobg}.png"] = full_path
                                
                                # Also store without the kogift_ prefix
                                if name_without_nobg.lower().startswith('kogift_'):
                                    base_without_prefix = name_without_nobg[7:] # Remove 'kogift_'
                                    kogift_images[f"{base_without_prefix}.jpg"] = full_path
                                    kogift_images[f"{base_without_prefix}.png"] = full_path
                            # 3d. Also map regular image names to their _nobg counterparts (new)
                            else:
                                # Create the _nobg variant names to check if they exist
                                name_without_ext, ext = os.path.splitext(base_name)
                                nobg_variant = f"{name_without_ext}_nobg.png"
                                nobg_path = os.path.join(root, nobg_variant)
                                # If the _nobg file exists, map the regular name to it as well
                                if os.path.exists(nobg_path):
                                    kogift_images[f"{base_name}"] = nobg_path
                                    kogift_images[f"{base_name.lower()}"] = nobg_path
                        else:
                            # 4. For files without kogift_ prefix, add it as an alternate key
                            with_prefix = f"kogift_{base_name}"
                            kogift_images[with_prefix] = full_path
                            kogift_images[with_prefix.lower()] = full_path
                        
                        # 5. Special handling for shop_ prefix (common in Kogift URLs)
                        if base_name.lower().startswith('shop_'):
                            # 5a. Store without shop_ prefix
                            no_shop = base_name[5:]  # Remove 'shop_'
                            kogift_images[no_shop] = full_path
                            kogift_images[no_shop.lower()] = full_path
                            
                            # 5b. Add kogift_ prefix but without shop_
                            kogift_without_shop = f"kogift_{no_shop}"
                            kogift_images[kogift_without_shop] = full_path
                            kogift_images[kogift_without_shop.lower()] = full_path
                        elif 'shop_' in base_name.lower():
                            # 5c. If shop_ appears elsewhere in the name
                            alt_version = base_name.lower().replace('shop_', '')
                            kogift_images[alt_version] = full_path
                
                total_images += len(files)
        except Exception as e:
            logger.error(f"Error scanning directory {base_dir}: {e}")
    
    logger.info(f"Scan complete. Found {total_images} total image files, created {len(kogift_images)} lookup entries")
    
    return kogift_images

def find_kogift_image_for_url(url: str, kogift_images: Dict[str, str]) -> Optional[str]:
    """
    Find the corresponding local Kogift image file for a given URL.
    
    Args:
        url: The URL to find a matching image for
        kogift_images: Dictionary of filename to local path mappings
        
    Returns:
        Local file path if found, None otherwise
    """
    if not url or not url.startswith(('http://', 'https://')):
        return None
    
    # 1. Try direct filename matching
    filename = os.path.basename(url)
    if filename in kogift_images:
        logger.debug(f"Direct filename match found for: {filename}")
        return kogift_images[filename]
    
    # Check lowercase version
    if filename.lower() in kogift_images:
        logger.debug(f"Lowercase filename match found for: {filename.lower()}")
        return kogift_images[filename.lower()]
    
    # 2. Try with kogift_ prefix
    prefixed = f"kogift_{filename}"
    if prefixed in kogift_images:
        logger.debug(f"Prefixed filename match found for: {prefixed}")
        return kogift_images[prefixed]
    
    if prefixed.lower() in kogift_images:
        logger.debug(f"Lowercase prefixed filename match found for: {prefixed.lower()}")
        return kogift_images[prefixed.lower()]
    
    # 3. Try URL patterns from Kogift
    # 3a. mall/shop_ pattern
    if 'mall/shop_' in url:
        product_part = url.split('mall/shop_')[1].split('?')[0]
        if product_part in kogift_images:
            logger.debug(f"Product match via mall/shop_ pattern: {product_part}")
            return kogift_images[product_part]
        
        # Try with kogift_ prefix
        prefixed_product = f"kogift_{product_part}"
        if prefixed_product in kogift_images:
            logger.debug(f"Prefixed product match via mall/shop_ pattern: {prefixed_product}")
            return kogift_images[prefixed_product]
    
    # 4. Try hash-based matching
    url_hash = hashlib.md5(url.encode()).hexdigest()[:10]
    hash_patterns = [
        f"kogift_{url_hash}.jpg",
        f"kogift_{url_hash}.png", 
        f"kogift_{url_hash}_nobg.png",
        url_hash
    ]
    
    for pattern in hash_patterns:
        if pattern in kogift_images:
            logger.debug(f"Hash pattern match found: {pattern}")
            return kogift_images[pattern]
    
    # 5. Try fuzzy matching as a last resort
    best_match = None
    highest_similarity = 0.4  # Threshold for accepting a match
    
    url_base = os.path.basename(url).lower()
    
    # Only consider main part of URL (before query params)
    if '?' in url_base:
        url_base = url_base.split('?')[0]
    
    # Skip very short URLs to avoid false matches
    if len(url_base) < 5:
        return None
    
    for img_name in kogift_images:
        img_name_lower = img_name.lower()
        
        # Skip very short filenames
        if len(img_name_lower) < 5:
            continue
        
        # Skip obvious non-matches
        if not any(part in img_name_lower for part in ['kogift', 'shop_']):
            continue
        
        # Calculate similarity score
        # Common substring approach
        i, j, max_len = 0, 0, 0
        len1, len2 = len(url_base), len(img_name_lower)
        
        for i in range(len1):
            for j in range(len2):
                k = 0
                while (i + k < len1 and j + k < len2 and 
                       url_base[i + k] == img_name_lower[j + k]):
                    k += 1
                max_len = max(max_len, k)
                
        similarity = max_len / max(len1, len2)
        
        if similarity > highest_similarity:
            highest_similarity = similarity
            best_match = kogift_images[img_name]
            
    if best_match:
        logger.debug(f"Fuzzy match found with similarity {highest_similarity:.2f}: {os.path.basename(best_match)}")
        return best_match
    
    return None

def fix_excel_kogift_images(excel_file: str, output_file: Optional[str] = None) -> str:
    """
    Fix Kogift images in an Excel file by ensuring all image URLs are properly linked to local files.
    
    Args:
        excel_file: Path to the Excel file to fix
        output_file: Optional path for the output file. If None, will generate one
    
    Returns:
        Path to the fixed Excel file
    """
    if not os.path.exists(excel_file):
        logger.error(f"Excel file not found: {excel_file}")
        return None
    
    # If no output file provided, create one
    if output_file is None:
        output_dir = os.path.dirname(excel_file)
        file_name = os.path.basename(excel_file)
        name_part, ext = os.path.splitext(file_name)
        output_file = os.path.join(output_dir, f"{name_part}_fixed{ext}")
    
    logger.info(f"Starting to fix Kogift images in: {excel_file}")
    logger.info(f"Output will be saved to: {output_file}")
    
    # Build kogift image database
    kogift_images = scan_kogift_images()
    
    # Load the Excel file with pandas to locate Kogift columns
    try:
        df = pd.read_excel(excel_file)
        logger.info(f"Successfully loaded Excel file with {len(df)} rows and {len(df.columns)} columns")
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return None
    
    # Find Kogift image columns
    kogift_columns = [col for col in df.columns if '고려' in col and ('이미지' in col or 'image' in col)]
    
    if not kogift_columns:
        logger.warning("No Kogift image columns found in the Excel file.")
        return None
    
    logger.info(f"Found Kogift image columns: {kogift_columns}")
    
    # Extract and process URLs from Kogift columns
    url_to_local_map = {}
    url_patterns = re.compile(r'https?://[^\s"\'<>]+')
    
    # Extract URLs and map to local files
    for col in kogift_columns:
        for idx, cell_value in enumerate(df[col]):
            if not isinstance(cell_value, str) or not cell_value:
                continue
                
            # Extract URLs from the cell
            urls = url_patterns.findall(cell_value)
            
            for url in urls:
                if url in url_to_local_map:
                    continue  # Already processed this URL
                    
                # Try to find local file for URL
                local_file = find_kogift_image_for_url(url, kogift_images)
                
                if local_file:
                    url_to_local_map[url] = local_file
                    logger.debug(f"Mapped URL: {url} -> {local_file}")
    
    logger.info(f"Found {len(url_to_local_map)} URL to local file mappings")
    
    # Now load the Excel file with openpyxl to modify it
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        # Get column indices for Kogift image columns
        kogift_col_indices = {}
        for i, col_name in enumerate(df.columns, 1):
            if col_name in kogift_columns:
                kogift_col_indices[col_name] = i
        
        # Track counts for reporting
        url_replaced = 0
        images_added = 0
        already_images = 0
        errors = 0
        
        # Process each cell in the Kogift columns
        for row_idx in range(2, worksheet.max_row + 1):  # Skip header row
            for col_name, col_idx in kogift_col_indices.items():
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Skip empty cells
                if not cell.value:
                    continue
                
                # Get cell value as string
                cell_text = str(cell.value)
                
                # Skip cells that are already images (no text)
                if cell_text == "":
                    already_images += 1
                    continue
                
                # Extract URLs from cell text
                urls = url_patterns.findall(cell_text)
                
                # If no URLs found, skip cell
                if not urls:
                    continue
                
                # Process first valid URL in the cell
                for url in urls:
                    if url in url_to_local_map:
                        local_file = url_to_local_map[url]
                        
                        # Verify file still exists
                        if not os.path.exists(local_file):
                            logger.warning(f"Mapped file no longer exists: {local_file}")
                            continue
                        
                        try:
                            # Replace cell value with empty string (for image)
                            cell.value = ""
                            
                            # Add hyperlink to the original URL
                            cell.hyperlink = url
                            
                            # Add image to the cell
                            img = Image(local_file)
                            
                            # Set image size
                            img.width = 200  # pixels
                            img.height = 200  # pixels
                            
                            # Position image
                            img.anchor = f"{get_column_letter(col_idx)}{row_idx}"
                            
                            # Add image to worksheet
                            worksheet.add_image(img)
                            
                            images_added += 1
                            logger.debug(f"Added image at row {row_idx}, column {col_idx} (Sheet position {get_column_letter(col_idx)}{row_idx})")
                            break
                        except Exception as e:
                            logger.error(f"Error adding image at row {row_idx}, column {col_idx}: {e}")
                            logger.debug(traceback.format_exc())
                            errors += 1
                            # Restore original cell value
                            cell.value = cell_text
        
        # Adjust row heights for rows with images
        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in kogift_col_indices.values():
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value == "":  # Empty cell may contain an image
                    # Set height to accommodate the image
                    worksheet.row_dimensions[row_idx].height = 200
                    break
        
        # Save the modified workbook
        workbook.save(output_file)
        logger.info(f"Successfully saved fixed Excel file: {output_file}")
        logger.info(f"Summary: {images_added} images added, {already_images} already had images, {errors} errors")
        
        return output_file
    
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        logger.debug(traceback.format_exc())
        return None

def main():
    """Main function to run the Kogift image fix tool."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Fix Kogift images in Excel files')
    parser.add_argument('--input', '-i', required=True, help='Input Excel file')
    parser.add_argument('--output', '-o', help='Output Excel file (optional)')
    
    args = parser.parse_args()
    
    result = fix_excel_kogift_images(args.input, args.output)
    
    if result:
        print(f"Successfully fixed Kogift images. Output saved to: {result}")
        return 0
    else:
        print("Failed to fix Kogift images. Check the log for details.")
        return 1

if __name__ == "__main__":
    sys.exit(main()) 