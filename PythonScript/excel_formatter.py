import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_upload_excel(file_path):
    """
    Apply formatting to the upload Excel file according to requirements:
    1) "제목"부분(1열)은 2줄로 보여지게 할것 (바탕색 : 그레이), 칼럼(제목) 부분 높이는 34.5
    2) "제목"부분을 제외한 리스트 부분의 서식은 "셀에맞춤"으로 설정할것
    3) 칸의 너비는 그림처럼 보여지게 조정할것 (열너비 : 7, 행 높이 : 17, )
    4) 표에 테두리를 넣을것
    """
    try:
        if not os.path.exists(file_path):
            logging.error(f"Upload Excel file not found: {file_path}")
            return False

        logging.info(f"Applying formatting to upload Excel file: {file_path}")
        wb = load_workbook(file_path)
        sheet = wb.active

        # 1) Format header row (gray background, 2 lines, height 34.5)
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Get max row and column
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Format headers (first row)
        sheet.row_dimensions[1].height = 34.5  # Header height
        
        # Dictionary to store header text lengths
        header_lengths = {}
        
        # First pass: collect header text lengths
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                header_lengths[col] = len(str(cell.value))
            else:
                header_lengths[col] = 0
        
        # Second pass: format headers and set column widths
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = gray_fill
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = thin_border
            
            # Calculate column width - base width is 7, but adjust for longer headers
            length = header_lengths[col]
            # Calculate width to ensure headers show in about 2 lines
            # For Korean text, each character needs more width
            if length > 0:
                # Check if header contains Korean characters
                has_korean = any('\uAC00' <= char <= '\uD7A3' for char in str(cell.value))
                
                if has_korean:
                    # Korean characters need more width
                    if length <= 4:
                        width = 7  # Default width for short headers
                    elif length <= 8:
                        width = 9  # Slightly wider for medium headers
                    else:
                        # For longer headers, calculate width to fit approximately 2 lines
                        width = min(max(7, length * 0.9), 16)
                else:
                    # For non-Korean text
                    if length <= 7:
                        width = 7  # Default width for short headers
                    elif length <= 14:
                        width = 9  # Slightly wider for medium headers
                    else:
                        # For longer headers, calculate width to fit approximately 2 lines
                        width = min(max(7, length * 0.6), 14)
            else:
                width = 7  # Default width
                
            # Set the column width
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        # 2) Set "fit to cell" for data rows
        for row in range(2, max_row + 1):
            # Set row height to 17
            sheet.row_dimensions[row].height = 17
            
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = thin_border
        
        # Save the workbook
        wb.save(file_path)
        logging.info(f"Successfully formatted upload Excel file: {file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error formatting upload Excel file: {e}", exc_info=True)
        return False

def format_result_excel(file_path):
    """
    Apply formatting to the result Excel file according to requirements:
    1) Adjust column widths based on content type
    2) Make headers wrap text (2 lines)
    3) Right-align specified numeric columns
    """
    try:
        if not os.path.exists(file_path):
            logging.error(f"Result Excel file not found: {file_path}")
            return False

        logging.info(f"Applying formatting to result Excel file: {file_path}")
        wb = load_workbook(file_path)
        sheet = wb.active
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Get max row and column
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Column names that need to be right-aligned (matches the result file's actual column names)
        right_align_columns = ['기본수량(1)', '판매단가(V포함)', '기본수량(2)', '판매가(V포함)(2)', '가격차이(2)', '가격차이(2)(%)', 
                              '기본수량(3)', '판매단가(V포함)(3)', '가격차이(3)', '가격차이(3)(%)']
        right_align_column_indices = []
        
        # Image column names that need wider width
        image_columns = ['본사 이미지', '고려기프트 이미지', '네이버 이미지']
        image_column_indices = []
        
        # Find column indices for special formatting
        for col in range(1, max_col + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value in right_align_columns:
                right_align_column_indices.append(col)
                logging.debug(f"Found right-align column at index {col}: {header_value}")
            if header_value in image_columns:
                image_column_indices.append(col)
                logging.debug(f"Found image column at index {col}: {header_value}")
        
        # Format headers (first row)
        sheet.row_dimensions[1].height = 34.5  # Set header row height
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = thin_border
            cell.fill = gray_fill  # Add gray background to header row
        
        # Format data rows
        for row in range(2, max_row + 1):
            # Set row height to 120 (for all rows to accommodate images)
            sheet.row_dimensions[row].height = 120
            
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Apply border to all cells
                cell.border = thin_border
                
                # Apply right alignment to numeric columns
                if col in right_align_column_indices:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Adjust column widths
        for col in range(1, max_col + 1):
            if col in image_column_indices:
                # Image columns: width 21.44
                sheet.column_dimensions[get_column_letter(col)].width = 21.44
                logging.debug(f"Set column {col} width to 21.44 for image")
            else:
                # Standard columns: width 7
                sheet.column_dimensions[get_column_letter(col)].width = 7
                
                # Get header text to adjust width if needed
                header_cell = sheet.cell(row=1, column=col)
                header_text = str(header_cell.value) if header_cell.value else ""
                
                # Adjust width for longer headers (similar to upload_excel logic)
                if header_text:
                    length = len(header_text)
                    has_korean = any('\uAC00' <= char <= '\uD7A3' for char in header_text)
                    
                    if has_korean:
                        if length <= 4:
                            width = 7
                        elif length <= 8:
                            width = 9
                        else:
                            width = min(max(7, length * 0.9), 16)
                    else:
                        if length <= 7:
                            width = 7
                        elif length <= 14:
                            width = 9
                        else:
                            width = min(max(7, length * 0.6), 14)
                            
                    sheet.column_dimensions[get_column_letter(col)].width = width
                    logging.debug(f"Adjusted column {col} width to {width} based on header length")
        
        # Save the workbook
        wb.save(file_path)
        logging.info(f"Successfully formatted result Excel file: {file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error formatting result Excel file: {e}", exc_info=True)
        return False

def apply_excel_formatting(result_path=None, upload_path=None):
    """
    Apply formatting to both result and upload Excel files.
    Returns tuple: (success_count, total_files)
    """
    success_count = 0
    total_files = 0
    
    if result_path:
        total_files += 1
        if format_result_excel(result_path):
            success_count += 1
            logging.info(f"Successfully applied formatting to result file: {result_path}")
        else:
            logging.error(f"Failed to apply formatting to result file: {result_path}")
            
    if upload_path:
        total_files += 1
        if format_upload_excel(upload_path):
            success_count += 1
            logging.info(f"Successfully applied formatting to upload file: {upload_path}")
        else:
            logging.error(f"Failed to apply formatting to upload file: {upload_path}")
            
    return success_count, total_files 