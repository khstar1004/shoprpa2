import os
import logging
import shutil
import requests
import pandas as pd
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import tempfile
import hashlib
from pathlib import Path

# 로거 설정
logger = logging.getLogger(__name__)

def create_excel_with_placed_images(df, output_excel_path):
    """
    DataFrame의 이미지 URL을 사용하여 이미지를 다운로드하고 Excel 파일에 배치하는 함수
    
    Args:
        df (pandas.DataFrame): 처리할 원본 데이터 ('본사 이미지', '고려기프트 이미지', '네이버 이미지' 열 포함)
        output_excel_path (str): 결과 Excel 파일 저장 경로
    
    Returns:
        None: 결과는 지정된 경로에 Excel 파일로 저장됨
    """
    logger.info(f"이미지 배치 Excel 생성 시작 - 출력 경로: {output_excel_path}")
    
    # 1. DataFrame 복사본 생성
    df_copy = df.copy()
    
    # 2. URL 컬럼 삭제 (접미사가 _URL인 컬럼들)
    url_columns = [col for col in df_copy.columns if col.endswith('_URL')]
    if url_columns:
        logger.info(f"URL 컬럼 {len(url_columns)}개 삭제: {url_columns}")
        df_copy = df_copy.drop(columns=url_columns)
    
    # 3. 이미지 열 확인 및 처리
    image_columns = ['본사 이미지', '고려기프트 이미지', '네이버 이미지']
    # 존재하는 이미지 컬럼만 필터링
    image_columns = [col for col in image_columns if col in df_copy.columns]
    
    if not image_columns:
        logger.warning("이미지 컬럼이 DataFrame에 존재하지 않습니다. 일반 Excel로 저장합니다.")
        df_copy.to_excel(output_excel_path, index=False, engine='openpyxl')
        return
    
    # 4. 임시 폴더 생성
    temp_dir = os.path.join(tempfile.gettempdir(), 'temp_downloaded_images')
    os.makedirs(temp_dir, exist_ok=True)
    logger.info(f"임시 이미지 폴더 생성: {temp_dir}")
    
    try:
        # 5. 이미지 컬럼의 URL은 Excel에 저장 시 빈 문자열로 대체
        for col in image_columns:
            df_copy[col] = ''  # URL 텍스트를 빈 문자열로 설정
        
        # 6. 기본 Excel 파일 생성
        logger.info(f"기본 Excel 파일 생성: {output_excel_path}")
        df_copy.to_excel(output_excel_path, index=False, engine='openpyxl')
        
        # 7. Excel 파일 열기
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb.active
        
        # 8. 이미지 다운로드 및 배치
        for col_idx, col_name in enumerate(image_columns):
            if col_name not in df.columns:
                continue
                
            # Excel에서의 컬럼 위치 찾기
            excel_col_idx = None
            for i, cell in enumerate(ws[1], 1):  # 첫 번째 행의 헤더 확인
                if cell.value == col_name:
                    excel_col_idx = i
                    break
            
            if excel_col_idx is None:
                logger.warning(f"Excel에서 컬럼 '{col_name}'을 찾을 수 없습니다.")
                continue
                
            excel_col_letter = get_column_letter(excel_col_idx)
            
            # 열 너비 조정 (한 번만)
            ws.column_dimensions[excel_col_letter].width = 22  # 150픽셀에 맞게 조정
            
            # 각 행의 데이터 처리
            for row_idx, row in df.iterrows():
                url = row.get(col_name)
                excel_row = row_idx + 2  # Excel은 1부터 시작, 헤더 행 1개
                
                # URL 유효성 검사
                if not isinstance(url, str) or not url or url == '-' or not url.startswith(('http://', 'https://')):
                    continue  # 유효하지 않은 URL은 건너뜀
                
                try:
                    # 이미지 파일 이름 (URL의 해시 사용)
                    img_hash = hashlib.md5(url.encode()).hexdigest()[:10]
                    temp_img_path = os.path.join(temp_dir, f"img_{img_hash}_row{row_idx}_col{excel_col_idx}.png")
                    
                    # 이미지 다운로드
                    logger.debug(f"이미지 다운로드 시작: {url}")
                    response = requests.get(url, timeout=30)
                    response.raise_for_status()  # HTTP 오류 확인
                    
                    # 이미지 저장
                    with open(temp_img_path, 'wb') as img_file:
                        img_file.write(response.content)
                    
                    # 이미지 리사이즈
                    with Image.open(temp_img_path) as img:
                        # LANCZOS 리샘플링으로 150x150 리사이즈
                        resized_img = img.resize((150, 150), Image.Resampling.LANCZOS)
                        resized_img.save(temp_img_path)
                    
                    # Excel에 이미지 추가
                    img_obj = XLImage(temp_img_path)
                    img_obj.width = 150
                    img_obj.height = 150
                    
                    # 셀 주소 계산 (예: A2, B3)
                    cell_address = f"{excel_col_letter}{excel_row}"
                    
                    # 행 높이 조정
                    ws.row_dimensions[excel_row].height = 115  # 150픽셀 높이에 맞게 조정
                    
                    # 이미지 배치
                    ws.add_image(img_obj, cell_address)
                    
                    # 이미지 배치 셀의 텍스트 값 비우기
                    ws[cell_address].value = None
                    
                    logger.debug(f"이미지 배치 완료: 셀 {cell_address}")
                    
                except Exception as e:
                    logger.error(f"이미지 처리 오류 (행 {row_idx+1}, 열 '{col_name}'): {str(e)}")
        
        # 9. Excel 파일 저장
        logger.info(f"이미지 배치 완료, Excel 파일 저장: {output_excel_path}")
        wb.save(output_excel_path)
        
    except Exception as e:
        logger.error(f"Excel 파일 생성 중 오류 발생: {str(e)}")
        raise
    finally:
        # 10. 임시 폴더 정리
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                logger.info(f"임시 이미지 폴더 삭제 완료: {temp_dir}")
        except Exception as cleanup_err:
            logger.warning(f"임시 폴더 삭제 중 오류: {str(cleanup_err)}")
    
    logger.info("이미지 배치 Excel 생성 완료")
    return


# 이 함수는 네이버 이미지 URL로부터 로컬 이미지 파일을 찾는 도우미 함수입니다.
def find_local_image_from_url(url, config=None):
    """
    이미지 URL에 해당하는 로컬 이미지 파일을 찾는 함수
    
    Args:
        url (str): 이미지 URL
        config (ConfigParser, optional): 설정 객체
    
    Returns:
        str or None: 로컬 이미지 경로 또는 None
    """
    if not url or not isinstance(url, str):
        return None
        
    try:
        # 설정에서 이미지 기본 경로 가져오기 (config가 제공된 경우)
        image_main_dir = None
        if config:
            image_main_dir = config.get('Paths', 'image_main_dir', fallback='C:\\RPA\\Image\\Main')
        else:
            # 기본 경로 설정
            image_main_dir = 'C:\\RPA\\Image\\Main'
        
        # URL에서 파일 이름 추출
        filename = os.path.basename(url.split('?')[0])  # URL 파라미터 제거
        
        # 이미지 소스 판별 및 해당 폴더 설정
        img_subfolder = None
        if 'pstatic.net' in url.lower() or 'naver' in url.lower():
            img_subfolder = 'Naver'
        elif 'koreagift' in url.lower() or 'kogift' in url.lower():
            img_subfolder = 'Kogift'
        elif 'jclgift' in url.lower() or 'haereum' in url.lower():
            img_subfolder = 'Haereum'
            
        if img_subfolder:
            # 해당 폴더에서 이미지 파일 찾기
            search_dir = os.path.join(image_main_dir, img_subfolder)
            if os.path.exists(search_dir):
                # 파일명으로 직접 검색
                direct_path = os.path.join(search_dir, filename)
                if os.path.exists(direct_path):
                    return direct_path
                
                # 파일 이름 일부로 검색 (URL 해시나 ID 부분이 포함된 파일)
                name_without_ext = os.path.splitext(filename)[0]
                for file in os.listdir(search_dir):
                    if name_without_ext in file:
                        return os.path.join(search_dir, file)
        
        # 전체 이미지 디렉토리에서 검색
        for root, dirs, files in os.walk(image_main_dir):
            for file in files:
                if filename == file:
                    return os.path.join(root, file)
                
                # URL 해시나 ID 부분이 파일 이름에 포함된 경우
                name_without_ext = os.path.splitext(filename)[0]
                if name_without_ext and len(name_without_ext) > 5 and name_without_ext in file:
                    return os.path.join(root, file)
        
        return None
        
    except Exception as e:
        logger.error(f"로컬 이미지 검색 오류 (URL: {url}): {str(e)}")
        return None 