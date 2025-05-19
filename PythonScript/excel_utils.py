import os
import logging
import pandas as pd
import datetime
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image
import functools
from functools import wraps
import configparser
import time
import re
from pathlib import Path
import traceback
import numpy as np
import json
from copy import copy
from decimal import Decimal
from typing import Optional
import hashlib
import glob
import io


# Check Python/PIL version for proper resampling constant
try:
    # Python 3.10+ with newer Pillow
    RESAMPLING_FILTER = Image.Resampling.LANCZOS
except (AttributeError, ImportError):
    try:
        # Older Pillow versions
        RESAMPLING_FILTER = Image.LANCZOS
    except (AttributeError, ImportError):
        # Very old Pillow versions
        RESAMPLING_FILTER = Image.ANTIALIAS

# Initialize logger
logger = logging.getLogger(__name__)

# Define constants for image source directory names
HAEREUM_DIR_NAME = 'Haereum'
KOGIFT_DIR_NAME = 'Kogift' # Changed from kogift_pre / Kogift
NAVER_DIR_NAME = 'Naver'   # Ensure this is uppercase 'Naver'
OTHER_DIR_NAME = 'Other'

# Initialize config parser
CONFIG = configparser.ConfigParser()
config_ini_path = Path(__file__).resolve().parent.parent / 'config.ini'
try:
    CONFIG.read(config_ini_path, encoding='utf-8')
    logger.info(f"Successfully loaded configuration from {config_ini_path}")
    # Get main paths from config
    IMAGE_MAIN_DIR = Path(CONFIG.get('Paths', 'image_main_dir', fallback='C:\\RPA\\Image\\Main'))
except Exception as e:
    logger.error(f"Error loading config from {config_ini_path}: {e}, using default values")
    # Use default directory if config fails
    IMAGE_MAIN_DIR = Path('C:\\RPA\\Image\\Main')

# --- Constants ---
PROMO_KEYWORDS = ['판촉', '기프트', '답례품', '기념품', '인쇄', '각인', '제작', '호갱', '몽키', '홍보']

# Column Rename Mapping (Ensure keys cover variations, values match FINAL_COLUMN_ORDER)
COLUMN_RENAME_MAP = {
    # Standard column renames - map FROM old names TO new names
    '구분(승인관리:A/가격관리:P)': '구분',
    '공급사명': '업체명',
    '공급처코드': '업체코드',
    '상품코드': 'Code',
    '카테고리(중분류)': '중분류카테고리',
    '본사 기본수량': '기본수량(1)',
    '판매단가1(VAT포함)': '판매단가(V포함)',
    '본사링크': '본사상품링크',
    '고려 기본수량': '기본수량(2)',
    '판매단가2(VAT포함)': '판매가(V포함)(2)',
    '고려 가격차이': '가격차이(2)',
    '고려 가격차이(%)': '가격차이(2)(%)',
    '고려 링크': '고려기프트 상품링크',
    '네이버 기본수량': '기본수량(3)',
    '판매단가3 (VAT포함)': '판매단가(V포함)(3)',
    '네이버 가격차이': '가격차이(3)',
    '네이버가격차이(%)': '가격차이(3)(%)',
    '네이버 공급사명': '공급사명',
    '네이버 링크': '네이버 쇼핑 링크',
    '해오름(이미지링크)': '본사 이미지',
    '고려기프트(이미지링크)': '고려기프트 이미지',
    '네이버쇼핑(이미지링크)': '네이버 이미지',
    
    # Self-maps (columns that already have correct names)
    '구분': '구분',
    '담당자': '담당자',
    '업체명': '업체명',
    '업체코드': '업체코드',
    'Code': 'Code',
    '중분류카테고리': '중분류카테고리',
    '상품명': '상품명',
    '기본수량(1)': '기본수량(1)',
    '판매단가(V포함)': '판매단가(V포함)',
    '본사상품링크': '본사상품링크',
    '기본수량(2)': '기본수량(2)',
    '판매가(V포함)(2)': '판매가(V포함)(2)',
    '가격차이(2)': '가격차이(2)',
    '가격차이(2)(%)': '가격차이(2)(%)',
    '고려기프트 상품링크': '고려기프트 상품링크',
    '기본수량(3)': '기본수량(3)',
    '판매단가(V포함)(3)': '판매단가(V포함)(3)',
    '가격차이(3)': '가격차이(3)',
    '가격차이(3)(%)': '가격차이(3)(%)',
    '공급사명': '공급사명',
    '네이버 쇼핑 링크': '네이버 쇼핑 링크',
    '공급사 상품링크': '공급사 상품링크',
    '본사 이미지': '본사 이미지',
    '고려기프트 이미지': '고려기프트 이미지',
    '네이버 이미지': '네이버 이미지'
}

# Final Target Column Order (Based on "엑셀 골든" sample)
# THIS IS THE STRICT ORDER AND NAMING FOR THE OUTPUT FILE
FINAL_COLUMN_ORDER = [
    '구분', '담당자', '업체명', '업체코드', 'Code', '중분류카테고리', '상품명',
    '기본수량(1)', '판매단가(V포함)', '본사상품링크',
    '기본수량(2)', '판매가(V포함)(2)', '가격차이(2)', '가격차이(2)(%)', '고려기프트 상품링크',
    '기본수량(3)', '판매단가(V포함)(3)', '가격차이(3)', '가격차이(3)(%)', '공급사명', 
    '네이버 쇼핑 링크', '공급사 상품링크',
    '본사 이미지', '고려기프트 이미지', '네이버 이미지'
]

# Columns that must be present in the input file for processing
# Update this based on the new FINAL_COLUMN_ORDER if necessary,
# focusing on the absolutely essential input fields needed.
REQUIRED_INPUT_COLUMNS = [
    '구분', '담당자', '업체명', '업체코드', 'Code', '중분류카테고리',
    '상품명', '기본수량(1)', '판매단가(V포함)', '본사상품링크'
]

# --- Column Type Definitions for Formatting ---
# Update these lists based on the FINAL_COLUMN_ORDER names
PRICE_COLUMNS = [
    '판매단가(V포함)', '판매가(V포함)(2)', '판매단가(V포함)(3)',
    '가격차이(2)', '가격차이(3)'
]
QUANTITY_COLUMNS = ['기본수량(1)', '기본수량(2)', '기본수량(3)']
PERCENTAGE_COLUMNS = ['가격차이(2)(%)', '가격차이(3)(%)']
TEXT_COLUMNS = ['구분', '담당자', '업체명', '업체코드', 'Code', '중분류카테고리', '상품명', '공급사명']
LINK_COLUMNS_FOR_HYPERLINK = {
    # Map final column names used for links
    '본사상품링크': '본사상품링크',
    '고려기프트 상품링크': '고려기프트 상품링크',
    '네이버 쇼핑 링크': '네이버 쇼핑 링크',
    '공급사 상품링크': '공급사 상품링크'
}
# Define IMAGE_COLUMNS based on FINAL_COLUMN_ORDER
IMAGE_COLUMNS = ['본사 이미지', '고려기프트 이미지', '네이버 이미지']

# Upload file columns (based on '엑셀골든_upload' notepad)
UPLOAD_COLUMN_ORDER = [
    '구분(승인관리:A/가격관리:P)', '담당자', '공급사명', '공급처코드', '상품코드', '카테고리(중분류)', '상품명',
    '본사 기본수량', '판매단가1(VAT포함)', '본사링크',
    '고려 기본수량', '판매단가2(VAT포함)', '고려 가격차이', '고려 가격차이(%)', '고려 링크',
    '네이버 기본수량', '판매단가3 (VAT포함)', '네이버 가격차이', '네이버가격차이(%)', '네이버 공급사명', 
    '네이버 링크', '해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)'
]

# Mapping between FINAL_COLUMN_ORDER and UPLOAD_COLUMN_ORDER
COLUMN_MAPPING_FINAL_TO_UPLOAD = {
    '구분': '구분(승인관리:A/가격관리:P)',
    '담당자': '담당자',  
    '업체명': '공급사명',
    '업체코드': '공급처코드',
    'Code': '상품코드',
    '중분류카테고리': '카테고리(중분류)',
    '상품명': '상품명',
    '기본수량(1)': '본사 기본수량',
    '판매단가(V포함)': '판매단가1(VAT포함)',
    '본사상품링크': '본사링크',
    '기본수량(2)': '고려 기본수량',
    '판매가(V포함)(2)': '판매단가2(VAT포함)',
    '가격차이(2)': '고려 가격차이',
    '가격차이(2)(%)': '고려 가격차이(%)',
    '고려기프트 상품링크': '고려 링크',
    '기본수량(3)': '네이버 기본수량',
    '판매단가(V포함)(3)': '판매단가3 (VAT포함)',
    '가격차이(3)': '네이버 가격차이',
    '가격차이(3)(%)': '네이버가격차이(%)',
    '공급사명': '네이버 공급사명',
    '네이버 쇼핑 링크': '네이버 링크',
    '본사 이미지': '해오름(이미지링크)',
    '고려기프트 이미지': '고려기프트(이미지링크)',
    '네이버 이미지': '네이버쇼핑(이미지링크)'
}

# Error Messages Constants (Can be used for conditional formatting or checks)
ERROR_MESSAGES = {
    'no_match': '가격 범위내에 없거나 텍스트 유사율을 가진 상품이 없음',
    'no_price_match': '가격이 범위내에 없거나 검색된 상품이 없음',
    'low_similarity': '일정 정확도 이상의 텍스트 유사율을 가진 상품이 없음',
    'no_results': '검색 결과 0',
    'no_image': '이미지를 찾을 수 없음',
    'file_not_found': '-이미지 없음-',
    'invalid_image': '유효하지 않은 이미지 형식',
    'processing_error': '-처리 오류-',
    'too_small': '이미지 크기가 너무 작음 (저해상도)',
    'format_error': '지원하지 않는 이미지 형식',
    'download_failed': '이미지 다운로드 실패',
    'excel_limit': '이미지 크기가 Excel 제한을 초과함'
}
ERROR_MESSAGE_VALUES = list(ERROR_MESSAGES.values()) # Cache list for faster checking

# --- Styling Constants ---
HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light green fill
HEADER_FONT = Font(bold=True, size=11, name='맑은 고딕')
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define alignments based on column type
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center", wrap_text=False) # Numbers right-aligned

DEFAULT_FONT = Font(name='맑은 고딕', size=10)

THIN_BORDER_SIDE = Side(style='thin')
DEFAULT_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

LINK_FONT = Font(color="0000FF", underline="single", name='맑은 고딕', size=10)
INVALID_LINK_FONT = Font(color="FF0000", name='맑은 고딕', size=10) # Red for invalid links

NEGATIVE_PRICE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow fill for negative diff < -1

# --- Image Processing Constants ---
IMAGE_MAX_SIZE = (2000, 2000)  # Excel 2021 maximum supported image size (increased from 1200x1200)
IMAGE_STANDARD_SIZE = (600, 600)  # Standard display size in Excel (increased from 400x400)
IMAGE_QUALITY = 85  # JPEG compression quality
SUPPORTED_IMAGE_FORMATS = ['.jpg', '.jpeg', '.png', '.gif', '.bmp']  # Supported by Excel 2021

# Image cell specific styling
IMAGE_CELL_HEIGHT = 420  # Increased from 360 for larger images
IMAGE_CELL_WIDTH = 60   # Increased from 44 for wider image cells

# --- Utility Functions ---

def retry_on_failure(max_retries=3, delay=1):
    """Decorator for retrying functions on failure."""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if attempt == max_retries - 1:
                        logger.error(f"Function {func.__name__} failed after {max_retries} attempts: {str(e)}")
                        raise
                    logger.warning(f"Attempt {attempt + 1} failed for {func.__name__}: {str(e)}")
                    time.sleep(delay)
            return None
        return wrapper
    return decorator

@retry_on_failure()
def find_excel_file(directory: str, extension: str = '.xlsx') -> Optional[str]:
    """Find the first Excel file with the specified extension in the directory."""
    try:
        # Ignore temporary Excel files starting with ~$
        files = [f for f in os.listdir(directory) if f.lower().endswith(extension) and not f.startswith('~$')]
        return files[0] if files else None
    except Exception as e:
        logger.error(f"Error finding Excel file in '{directory}': {str(e)}")
        raise

# validate_excel_file is now handled by check_excel_file.py
# def validate_excel_file(...) -> removed

def convert_text_to_numbers(df: pd.DataFrame) -> pd.DataFrame:
    """(Deprecated/Simplified) Initial conversion, formatting is now primarily handled in _prepare_data_for_excel."""
    logger.debug("Skipping deprecated convert_text_to_numbers function. Formatting handled in _prepare_data.")
    return df

def preprocess_product_name(name: str) -> str:
    """Preprocess product name (basic cleaning)."""
    if not isinstance(name, str):
        return str(name)
    # Keep basic cleaning, more advanced logic might be in matching modules
    return re.sub(r'[\(\)\[\]{}]+', '', name).strip() # Example: Remove only brackets


# --- Core Excel Creation Logic ---

def _apply_column_widths(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Sets appropriate column widths based on column names/types."""
    # Define width hints (can be adjusted)
    width_hints = {
        'image': 21.44, # Width for image columns - UPDATED to 21.44 per requirement
        'name': 45,  # 상품명
        'link': 35,
        'price': 14,
        'percent': 10,
        'quantity': 10,
        'code': 12,
        'category': 20,
        'text_short': 7, # UPDATED to 7 per requirement
        'default': 15
    }
    logger.debug(f"Applying column widths. DataFrame columns: {df.columns.tolist()}")
    for idx, col_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        width = width_hints['default'] # Default width

        col_name_str = str(col_name) # Ensure col_name is string for checks

        # Determine width based on column name patterns
        if col_name_str in IMAGE_COLUMNS:
            width = width_hints['image']
        elif '상품명' in col_name_str:
            width = width_hints['name']
        elif col_name_str in LINK_COLUMNS_FOR_HYPERLINK or '링크' in col_name_str:
            width = width_hints['link']
        elif col_name_str in PRICE_COLUMNS:
            width = width_hints['price']
        elif col_name_str in PERCENTAGE_COLUMNS:
            width = width_hints['percent']
        elif col_name_str in QUANTITY_COLUMNS:
             width = width_hints['quantity']
        elif 'Code' in col_name_str or '코드' in col_name_str:
            width = width_hints['code']
        elif '카테고리' in col_name_str or '분류' in col_name_str: # Added '분류'
            width = width_hints['category']
        elif col_name_str in ['구분', '담당자', '업체명', '업체코드']: # UPDATED: Added more columns to use short width
            width = width_hints['text_short']
        # Add more specific rules if needed

        worksheet.column_dimensions[column_letter].width = width
        # logger.debug(f"Set width for column '{col_name_str}' ({column_letter}) to {width}") # Reduce log verbosity
    logger.debug("Finished applying column widths.")

def _apply_cell_styles_and_alignment(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Applies formatting (font, border, alignment) to header and data cells."""
    logger.debug("Applying cell styles and alignments.")
    # Header Styling
    for cell in worksheet[1]: # First row is header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        
        # UPDATED: Enable wrap text for headers to display in 2 lines
        header_alignment = copy(HEADER_ALIGNMENT)
        header_alignment.wrap_text = True
        cell.alignment = header_alignment
        
        cell.border = DEFAULT_BORDER

    # Data Cell Styling
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = DEFAULT_FONT
            cell.border = DEFAULT_BORDER

            # Apply alignment based on column type
            col_name_str = str(col_name)
            is_pct_col = col_name_str in ['가격차이(2)(%)', '가격차이(3)(%)'] # Explicit check for percentage columns

            # Check if the cell value is likely numeric (ignoring error messages)
            is_numeric_value = False
            cell_value_str = str(cell.value)
            if cell_value_str not in ERROR_MESSAGE_VALUES and cell_value_str != '-':
                 # Basic check if it looks like a number (might need refinement)
                 try:
                      float(cell_value_str.replace(',', '').replace('%',''))
                      is_numeric_value = True
                 except ValueError:
                      is_numeric_value = False

            # Apply right alignment to numbers and specifically formatted percentage strings
            # UPDATED: Added explicit right alignment for requested columns
            quantity_columns = ['기본수량(1)', '기본수량(2)', '판매가(V포함)(2)', '기본수량(3)']
            if col_name_str in quantity_columns:
                cell.alignment = RIGHT_ALIGNMENT
            elif is_pct_col or ((col_name_str in PRICE_COLUMNS or col_name_str in QUANTITY_COLUMNS) and is_numeric_value):
                cell.alignment = RIGHT_ALIGNMENT
            # Update checks for center alignment based on new names
            elif col_name_str in IMAGE_COLUMNS or '코드' in col_name_str or 'Code' in col_name_str or col_name_str == '구분': # Use new '구분', added 'Code'
                 cell.alignment = CENTER_ALIGNMENT
            else:
                cell.alignment = LEFT_ALIGNMENT # Default left align for text/links/errors
    logger.debug("Finished applying cell styles.")


def clean_naver_images_and_data(worksheet, df):
    """
    Clean Naver images and data for invalid entries:
    1. Remove any displayed images for rows without a valid Naver URL or local path
    2. Clear all Naver-related cells for those rows
    
    Args:
        worksheet: The Excel worksheet
        df: DataFrame with the source data
    """
    try:
        if worksheet is None:
            return
        
        # Check if we should be more lenient with Naver data validation
        # This flag can be set in various parts of the application
        lenient_validation = True  # Always set to True to make it more lenient
        
        # If lenient validation is enabled, skip cleaning to preserve Naver data
        if lenient_validation:
            logger.info("Using lenient Naver data validation - skipping strict cleaning")
            return
            
        # The rest of the function will be skipped due to the early return above
        # Find column indexes for Naver-related columns
        header_row = 1  # First row is header
        naver_related_headers = ['네이버 이미지', '네이버 쇼핑 링크', '공급사 상품링크', '공급사명',
                              '판매단가(V포함)(3)', '기본수량(3)', '가격차이(3)', '가격차이(3)(%)']
        
        # Find column indices for all relevant columns
        naver_related_col_indices = []
        naver_img_col_idx = None
        
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col_idx).value
            if cell_value in naver_related_headers:
                naver_related_col_indices.append(col_idx)
                if cell_value == '네이버 이미지':
                    naver_img_col_idx = col_idx
        
        if not naver_img_col_idx or not naver_related_col_indices:
            logger.warning("Naver image column or related columns not found in worksheet")
            return
        
        # Track images to be removed
        images_to_remove = []
        
        # 네이버 이미지 컬럼의 각 행을 순회하며 링크 유무 확인
        for row_idx in range(2, worksheet.max_row + 1):  # 1행은 헤더
            cell_value = worksheet.cell(row=row_idx, column=naver_img_col_idx).value
            
            # Check if the cell value is a valid URL or if it has a valid local path
            is_valid = False
            
            # If the value is in the DataFrame and is a dictionary with URL
            if len(df) >= row_idx - 1:  # Adjust for header row
                df_idx = row_idx - 2  # DataFrame index (0-based, and accounting for header)
                if df_idx < len(df) and '네이버 이미지' in df.columns:
                    naver_img_value = df.iloc[df_idx]['네이버 이미지']
                    if isinstance(naver_img_value, dict):
                        # Check if it has a URL
                        if 'url' in naver_img_value and naver_img_value['url'] and isinstance(naver_img_value['url'], str):
                            # More lenient URL check - only basic requirement that it starts with http
                            if naver_img_value['url'].startswith(('http://', 'https://')):
                                is_valid = True
                        
                        # Also check for local_path
                        if 'local_path' in naver_img_value and naver_img_value['local_path'] and isinstance(naver_img_value['local_path'], str):
                            if os.path.exists(naver_img_value['local_path']):
                                is_valid = True
            
            # Cell value check - less strict, just check for URL-like value
            if isinstance(cell_value, str) and cell_value.startswith(('http://', 'https://')):
                is_valid = True
            
            # 링크가 없거나 빈값이면 처리 대상
            if not is_valid and cell_value in (None, '', '-', '-이미지 없음-', '-처리 오류-'):
                # 1) 해당 셀 위치에 삽입된 이미지 삭제
                # openpyxl은 이미지 객체가 워크시트에 리스트로 존재함
                # 이미지의 anchor 속성으로 위치를 확인 가능
                cell_coordinate = f"{get_column_letter(naver_img_col_idx)}{row_idx}"
                
                for img in worksheet._images:
                    # img.anchor는 openpyxl.drawing.spreadsheet_drawing.Anchor 객체 또는 문자열 좌표
                    # 좌표 문자열인 경우가 많으므로 문자열 비교
                    if hasattr(img.anchor, 'from_'):
                        # Anchor 객체인 경우 좌표 추출
                        anchor_coord = f"{get_column_letter(img.anchor._from.col + 1)}{img.anchor._from.row + 1}"
                    else:
                        anchor_coord = str(img.anchor)
                    
                    if anchor_coord == cell_coordinate:
                        images_to_remove.append(img)
                
                # 2) 네이버 관련 컬럼들 초기화 ('-')
                for col_idx in naver_related_col_indices:
                    worksheet.cell(row=row_idx, column=col_idx).value = '-'
        
        # 이미지 삭제 (리스트에서 제거)
        for img in images_to_remove:
            try:
                worksheet._images.remove(img)
            except Exception as e:
                # 로그 남기거나 무시
                pass
    except Exception as e:
        logger.error(f"Error cleaning Naver images and data: {e}")
        # Continue with execution even if this function fails

def _process_image_columns(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Process image columns in the DataFrame and add images to the worksheet.
    
    This function handles complex nested data structures for images, which may be dictionaries
    with paths or URLs, or plain string paths. The function will try to find the image files
    and embed them in the Excel file if they exist.
    
    Args:
        worksheet: The worksheet to add images to
        df: DataFrame containing the data with image columns
    """
    import openpyxl
    from openpyxl.drawing.image import Image
    
    # Initialize tracking variables
    successful_embeddings = 0
    attempted_embeddings = 0
    kogift_successful = 0
    kogift_attempted = 0
    naver_successful = 0
    naver_attempted = 0
    
    # Only handle these image-specific columns
    global IMAGE_COLUMNS
    columns_to_process = [col for col in IMAGE_COLUMNS if col in df.columns]
    
    if not columns_to_process:
        logger.debug("No image columns found in DataFrame")
        return 0
        
    # Fallback image to use if an image file is not found
    default_img_path = os.environ.get('RPA_DEFAULT_IMAGE', None)
    default_exists = default_img_path and os.path.exists(default_img_path)
    if not default_exists and default_img_path:
        logger.warning(f"Default image not found at {default_img_path}")
        default_img_path = None
        
    # Track if the fallback image was used
    used_fallback = False
    
    # Note: This is a reference to PIL.Image to avoid confusion with openpyxl.drawing.image.Image
    from PIL import Image as PILImage
    
    # Function to safely load and resize an image for Excel
    def safe_load_image(path, max_height=150, max_width=150, retry_count=2):
        """Safely load and resize an image, with error handling and retry logic"""
        if path is None or not isinstance(path, str):
            logger.warning(f"Invalid image path: {path}")
            return None
            
        # Handle file:// protocol
        if path.startswith('file:///'):
            path = path.replace('file:///', '')
            # Normalize slashes for Windows
            path = path.replace('/', os.sep)
        
        # Skip URLs - we need local files
        if path.startswith(('http://', 'https://')):
            logger.debug(f"Skipping URL path (need local file): {path}")
            # Check if we have a matching local path for this URL in the image directory
            # Try to find a matching file in the appropriate image directory
            if 'phinf.pstatic.net' in path:  # Naver image
                local_path = find_naver_image_from_url(path)
                if local_path:
                    logger.info(f"Found local Naver image for URL: {local_path}")
                    path = local_path
                else:
                    return None
            else:
                return None
        
        # If path still exists, try to open and resize the image
        retry = 0
        while retry <= retry_count:
            try:
                if not os.path.exists(path):
                    logger.warning(f"Image path does not exist: {path}")
                    return None
                    
                if os.path.getsize(path) <= 0:
                    logger.warning(f"Image file is empty: {path}")
                    return None
                
                img = PILImage.open(path)
                
                # Handle RGBA images (convert to RGB for Excel compatibility)
                if img.mode == 'RGBA':
                    # Create white background
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    # Paste using alpha as mask
                    background.paste(img, mask=img.split()[3])  # 3 is the alpha channel
                    img = background
                    
                # Calculate new dimensions while maintaining aspect ratio
                width, height = img.size
                if width > max_width or height > max_height:
                    ratio = min(max_width / width, max_height / height)
                    new_width = int(width * ratio)
                    new_height = int(height * ratio)
                    img = img.resize((new_width, new_height), Image.LANCZOS)
                    
                # Convert PIL image to BytesIO buffer
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format=img.format or 'JPEG')
                img_byte_arr.seek(0)
                
                return img_byte_arr
                
            except Exception as e:
                retry += 1
                error_msg = f"Error loading image {path} (attempt {retry}/{retry_count}): {e}"
                if retry <= retry_count:
                    logger.warning(error_msg)
                    time.sleep(0.5)  # Short delay before retry
                else:
                    logger.error(error_msg)
                    return None
    
    # Helper function to find Naver image file using hash patterns
    def find_naver_image_from_url(url):
        """Find local Naver image file from URL using hash patterns and intelligent matching"""
        if not url or not isinstance(url, str):
            return None
            
        # Skip non-image URLs like shopping links - only process phinf.pstatic.net URLs
        if not 'phinf.pstatic.net' in url:
            logger.debug(f"Not a Naver image URL (skipping): {url[:50]}...")
            return None
            
        # Naver image directory
        naver_dir = os.path.join(IMAGE_MAIN_DIR, 'Naver')
        if not os.path.exists(naver_dir):
            logger.warning(f"Naver image directory not found: {naver_dir}")
            return None
            
        # Extract hash patterns from URL
        hash_pattern1 = re.search(r'([a-f0-9]{16})[^a-f0-9]?([a-f0-9]{8})', url)
        hash_pattern2 = re.findall(r'[a-f0-9]{8,}', url)
        
        # List all Naver image files
        naver_files = glob.glob(os.path.join(naver_dir, "naver_*.jpg")) + glob.glob(os.path.join(naver_dir, "naver_*.png"))
        
        if not naver_files:
            logger.warning(f"No Naver image files found in directory: {naver_dir}")
            return None
            
        # Try exact hash match first (most reliable)
        if hash_pattern1:
            hash1, hash2 = hash_pattern1.groups()
            hash_combined = f"{hash1}_{hash2}"
            
            # Look for exact pattern match
            for file_path in naver_files:
                filename = os.path.basename(file_path)
                if hash_combined in filename:
                    # Prefer _nobg.png files if available
                    if '_nobg.png' in filename:
                        logger.debug(f"Found exact Naver hash match (nobg): {filename}")
                        return file_path
                
            # Second pass for non-nobg version if nobg not found
            for file_path in naver_files:
                filename = os.path.basename(file_path)
                if hash_combined in filename:
                    logger.debug(f"Found exact Naver hash match: {filename}")
                    return file_path
                    
            # Try individual hashes
            for file_path in naver_files:
                filename = os.path.basename(file_path)
                if hash1 in filename or hash2 in filename:
                    logger.debug(f"Found partial Naver hash match: {filename}")
                    return file_path
        
        # Try pattern match with individual hashes found in URL
        if hash_pattern2:
            for hash_segment in hash_pattern2:
                if len(hash_segment) >= 8:  # Only use reasonably long hashes
                    for file_path in naver_files:
                        filename = os.path.basename(file_path)
                        if hash_segment in filename:
                            logger.debug(f"Found Naver hash segment match: {filename}")
                            return file_path
        
        # Last resort - use URL hash
        url_hash = hashlib.md5(url.encode()).hexdigest()[:12]
        for file_path in naver_files:
            filename = os.path.basename(file_path)
            # Look for similar patterns
            if any(segment in filename for segment in url_hash.split('_')):
                logger.debug(f"Found Naver URL hash match: {filename}")
                return file_path
                
        logger.warning(f"No matching Naver image file found for URL: {url[:50]}...")
        return None
    
    # Track the count of images per column
    img_counts = {col: 0 for col in columns_to_process}
    err_counts = {col: 0 for col in columns_to_process}
    
    logger.debug(f"Processing {len(columns_to_process)} image columns")
    
    # For each image column in the DataFrame
    for col_idx, column in enumerate(columns_to_process):
        is_kogift_image = 'kogift' in column.lower() or '고려기프트' in column  # Track whether it's a Kogift column
        is_naver_image = 'naver' in column.lower() or '네이버' in column       # Add tracking for Naver columns
        
        # Excel column letter for this column (e.g., 'A', 'B', ...)
        excel_col = get_column_letter(df.columns.get_loc(column) + 1)
        
        # For each row in the DataFrame
        for row_idx, cell_value in enumerate(df[column]):
            img_path = None  # Initialize image path
            fallback_img_path = default_img_path  # Use default fallback
            
            # Skip empty cells (None, NaN, empty strings)
            if pd.isna(cell_value) or cell_value == "":
                continue
                
            # Skip cells with placeholder dash
            if cell_value == "-":
                continue
                
            # Special handling for Naver images - skip placeholders and non-image URLs
            if is_naver_image:
                # Skip any Naver data that doesn't have a valid image URL
                if isinstance(cell_value, dict):
                    url = cell_value.get('url', '')
                    if not url or not isinstance(url, str) or not 'phinf.pstatic.net' in url:
                        logger.debug(f"Skipping Naver value that doesn't have a valid image URL: {str(cell_value)[:50]}...")
                        continue
                elif isinstance(cell_value, str) and not 'phinf.pstatic.net' in cell_value:
                    # Skip strings that aren't Naver image URLs
                    logger.debug(f"Skipping Naver string that isn't an image URL: {cell_value[:50]}...")
                    continue
            
            # Handle dictionary format (most complete info)
            if isinstance(cell_value, dict):
                # Log for debugging which is very useful for Naver images
                if is_naver_image:
                    logger.debug(f"Processing Naver image data: {str(cell_value)[:100]}...")
                
                # Try local path first, then URL
                if 'local_path' in cell_value and cell_value['local_path']:
                    img_path = cell_value['local_path']
                    
                    # IMPROVED: Special handling for Naver images - log and verify paths
                    if is_naver_image:
                        logger.debug(f"Found Naver local_path: {img_path}")
                        
                        # Verify the path exists and is absolute
                        if not os.path.isabs(img_path):
                            abs_path = os.path.abspath(img_path)
                            logger.debug(f"Converting relative Naver path to absolute: {img_path} -> {abs_path}")
                            img_path = abs_path
                        
                        # Verify the file exists - if not, try more alternatives
                        if not os.path.exists(img_path):
                            logger.warning(f"Naver image path doesn't exist: {img_path}")
                            
                            # IMPROVED: More comprehensive fallback strategy
                            url_to_try = cell_value.get('url')
                            if url_to_try and 'phinf.pstatic.net' in url_to_try:
                                # Try to find alternative file based on URL hash pattern
                                alt_path = find_naver_image_from_url(url_to_try)
                                if alt_path:
                                    logger.info(f"Found alternative Naver image from URL: {alt_path}")
                                    img_path = alt_path
                                    # Update the cell value with correct path for future use
                                    cell_value['local_path'] = alt_path
                                    df.at[row_idx, column] = cell_value
                            
                            if not os.path.exists(img_path):
                                # Try alternative extensions
                                base_path = os.path.splitext(img_path)[0]
                                for ext in ['.jpg', '.jpeg', '.png', '.gif']:
                                    alt_path = f"{base_path}{ext}"
                                    if os.path.exists(alt_path):
                                        logger.info(f"Found alternative Naver image path: {alt_path}")
                                        img_path = alt_path
                                        break
                                else:
                                    # If no alternative found, try looking for _nobg version
                                    nobg_path = f"{base_path}_nobg.png"
                                    if os.path.exists(nobg_path):
                                        logger.info(f"Found _nobg version of Naver image: {nobg_path}")
                                        img_path = nobg_path
                
                # IMPROVED: If local_path failed, only try URL-based approach for Naver images with phinf.pstatic.net URLs
                if not img_path or not os.path.exists(img_path):
                    url = cell_value.get('url')
                    if url and isinstance(url, str) and url.startswith(('http://', 'https://')):
                        # For Naver, only process actual image URLs (phinf.pstatic.net)
                        if is_naver_image and not 'phinf.pstatic.net' in url:
                            logger.debug(f"Skipping Naver URL that isn't a phinf.pstatic.net image URL: {url[:50]}...")
                            continue
                            
                        logger.debug(f"Local path failed or missing, trying URL-based approach: {url[:50]}...")
                        
                        if is_naver_image:
                            # For Naver images, use the hash-based lookup function
                            alt_path = find_naver_image_from_url(url)
                            if alt_path:
                                logger.info(f"Found Naver image via URL hash: {alt_path}")
                                img_path = alt_path
                                # Update the cell value with correct path for future reference
                                cell_value['local_path'] = alt_path
                                df.at[row_idx, column] = cell_value
            
            # If we've found a potentially valid path, try to load the image
            if img_path and os.path.exists(img_path):
                try:
                    attempted_embeddings += 1
                    if is_naver_image:
                        naver_attempted += 1
                    elif is_kogift_image:
                        kogift_attempted += 1
                    
                    # Process the image safely - resize for Excel
                    processed_img_path = safe_load_image(img_path, 
                                                         max_height=IMAGE_STANDARD_SIZE[1], 
                                                         max_width=IMAGE_STANDARD_SIZE[0])
                    
                    if processed_img_path and os.path.exists(processed_img_path):
                        # Calculate Excel cell position (1-indexed for openpyxl)
                        cell_position = f"{excel_col}{row_idx + 2}"  # +2 for header row and 1-indexing
                        
                        try:
                            # Create the openpyxl image object
                            img_obj = Image(processed_img_path)
                            
                            # Default anchor is cell top-left with row/col offsets at 0
                            img_obj.anchor = cell_position
                            
                            # Add image to worksheet
                            worksheet.add_image(img_obj)
                            
                            # Update success counts
                            successful_embeddings += 1
                            img_counts[column] += 1
                            
                            if is_naver_image:
                                naver_successful += 1
                                logger.debug(f"Successfully added Naver image at {cell_position}: {img_path}")
                            elif is_kogift_image:
                                kogift_successful += 1
                                logger.debug(f"Successfully added Kogift image at {cell_position}: {img_path}")
                            else:
                                logger.debug(f"Successfully added image at {cell_position}: {img_path}")
                                
                        except Exception as img_err:
                            logger.error(f"Error adding image to worksheet at {cell_position}: {str(img_err)}")
                            err_counts[column] += 1
                            
                            # Try one more time with the original image (unprocessed)
                            try:
                                img_obj = Image(img_path)
                                img_obj.anchor = cell_position
                                worksheet.add_image(img_obj)
                                logger.info(f"Successfully added original image as fallback at {cell_position}")
                                successful_embeddings += 1
                                img_counts[column] += 1
                            except Exception as orig_err:
                                logger.error(f"Error adding original image: {str(orig_err)}")
                    else:
                        logger.warning(f"Failed to process image at {img_path}")
                        err_counts[column] += 1
                except Exception as e:
                    logger.error(f"Error processing image at {img_path}: {str(e)}")
                    err_counts[column] += 1
            elif img_path:
                logger.warning(f"Image path not found: {img_path}")
                err_counts[column] += 1
    
    # Log success/error statistics
    logger.info(f"Image embedding complete: {successful_embeddings}/{attempted_embeddings} successful")
    logger.info(f"Naver images: {naver_successful}/{naver_attempted} successful")
    logger.info(f"Kogift images: {kogift_successful}/{kogift_attempted} successful")
    
    for col in columns_to_process:
        logger.info(f"Column '{col}': {img_counts[col]} images added, {err_counts[col]} errors")
    
    return successful_embeddings

def _apply_conditional_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Applies conditional formatting (e.g., yellow fill for price difference < -1)."""
    logger.debug("Applying conditional formatting.")

    # Find price difference columns (both regular and percentage) using new names
    price_diff_cols = [
        col for col in df.columns
        if col in ['가격차이(2)', '가격차이(3)', '가격차이(2)(%)', '가격차이(3)(%)'] # Include percentage columns
    ]

    if not price_diff_cols:
        logger.debug("No price difference columns found for conditional formatting.")
        return

    # Define yellow fill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # First check if these columns actually exist in the DataFrame
    existing_diff_cols = [col for col in price_diff_cols if col in df.columns]
    if not existing_diff_cols:
        logger.warning("None of the price difference columns exist in the DataFrame. Skipping conditional formatting.")
        return

    # Add detailed logging for debugging
    logger.info(f"가격차이 조건부 서식 적용 시작 (음수 강조): {existing_diff_cols}")
    logger.info(f"총 확인할 행 수: {worksheet.max_row - 1}")  # Subtract 1 for header row
    
    # Log column data types for debugging
    for col in existing_diff_cols:
        logger.info(f"열 '{col}' 데이터 타입: {df[col].dtype}")
        # Try to count negative values in each column
        try:
            if df[col].dtype in ['int64', 'float64']:
                # For numeric columns, count directly
                neg_count = (df[col] < -1).sum()
                logger.info(f"열 '{col}'에서 -1 미만인 값: {neg_count}개")
            else:
                # For non-numeric columns, try to convert first
                try:
                    neg_count = (pd.to_numeric(df[col], errors='coerce') < -1).sum()
                    logger.info(f"열 '{col}'에서 -1 미만인 값: {neg_count}개 (변환 후)")
                except:
                    logger.warning(f"열 '{col}'의 값을 숫자로 변환할 수 없습니다.")
        except Exception as e:
            logger.warning(f"열 '{col}'에서 음수 값 계산 중 오류: {e}")
    
    rows_highlighted = 0
    rows_checked = 0
    errors = 0

    # Get the column indices in the Excel worksheet
    col_indices = {}
    for i, header in enumerate(df.columns, 1):
        col_indices[header] = i

    # Process each row
    for row_idx in range(2, worksheet.max_row + 1):  # Excel is 1-indexed, row 1 is header
        rows_checked += 1
        highlight_row = False
        
        # Check each price difference column
        for diff_col in existing_diff_cols:
            # Get the Excel column index
            if diff_col not in col_indices:
                continue
                
            col_idx = col_indices[diff_col]
            
            # Get the cell value directly from the worksheet
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            
            # Skip empty cells
            if cell_value is None or cell_value == '' or cell_value == '-':
                continue
                
            try:
                # Handle different types of values
                numeric_value = None
                
                if isinstance(cell_value, (int, float)):
                    # Direct numeric value
                    numeric_value = float(cell_value)
                elif isinstance(cell_value, str):
                    # Strip any whitespace, commas, currency symbols
                    clean_value = cell_value.strip().replace(',', '').replace(' ', '')
                    
                    # Handle parentheses format for negative numbers like (100)
                    if clean_value.startswith('(') and clean_value.endswith(')'):
                        clean_value = '-' + clean_value[1:-1]
                        
                    # Attempt conversion to float if it's not just a dash or empty
                    if clean_value and clean_value != '-':
                        try:
                            numeric_value = float(clean_value)
                        except ValueError:
                            # If conversion fails, skip this cell
                            continue
                
                # If we successfully got a numeric value and it's < -1, highlight the row
                if numeric_value is not None and numeric_value < -1:
                    highlight_row = True
                    logger.debug(f"음수 가격차이 발견: 행 {row_idx}, 열 '{diff_col}', 값 {numeric_value} < -1")
                    break
                    
            except Exception as e:
                logger.warning(f"행 {row_idx}, 열 '{diff_col}' 처리 중 오류: {e}")
                errors += 1
                continue
                
        # Apply highlighting to the entire row if needed
        if highlight_row:
            rows_highlighted += 1
            for col_idx in range(1, worksheet.max_column + 1):
                try:
                    # Apply yellow fill to all cells in the row
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow_fill
                except Exception as e:
                    logger.error(f"셀 서식 적용 오류 (행 {row_idx}, 열 {col_idx}): {e}")
                    errors += 1

    # Log summary of highlighting results
    logger.info(f"조건부 서식 적용 완료: {rows_highlighted}개 행에 가격차이 < -1 하이라이팅 적용됨 (검사 행: {rows_checked}, 오류: {errors})")

def _setup_page_layout(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    """Sets up page orientation, print area, freeze panes, etc."""
    logger.debug("Setting up page layout.")
    try:
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0 # Fit to width primarily
        worksheet.print_options.horizontalCentered = True
        # worksheet.print_options.verticalCentered = True # Optional
        worksheet.print_options.gridLines = False # Typically false for final reports
        worksheet.freeze_panes = 'A2'  # Freeze header row
        # Set print area to used range (optional, helps if there's stray data)
        # worksheet.print_area = worksheet.dimensions
        logger.debug("Page layout settings applied.")
    except Exception as e:
        logger.error(f"Failed to set page layout options: {e}")

def _add_hyperlinks_to_worksheet(worksheet, df, hyperlinks_as_formulas=False):
    """
    Adds hyperlinks to URL cells in the worksheet.
    If hyperlinks_as_formulas=True, use Excel formulas for hyperlinks.
    Otherwise, use openpyxl's Hyperlink object.
    """
    try:
        # Define columns that should contain hyperlinks
        link_columns = [col for col in df.columns if any(term in col.lower() for term in ['링크', 'link', 'url'])]
        
        # Process each URL column
        total_urls_processed = 0
        
        for col in link_columns:
            if col in df.columns:
                col_idx = list(df.columns).index(col) + 1  # 1-based indexing for Excel
                
                # Loop through each cell in this column
                for row_idx, value in enumerate(df[col], 2):  # Start from row 2 (after header)
                    # Handle Series objects
                    if isinstance(value, pd.Series):
                        # Take the first non-empty value
                        for item in value:
                            if pd.notna(item) and item not in ['-', '']:
                                value = item
                                break
                        else:
                            value = ''
                    
                    # Skip empty values
                    if pd.isna(value) or value in ['', '-', 'None', 'nan']:
                        continue
                        
                    # Convert to string
                    url = str(value)
                    
                    # Extract URL from dictionary if needed
                    if isinstance(value, dict) and 'url' in value:
                        url = value['url']
                    
                    # Skip non-URL values
                    if not ('http://' in url or 'https://' in url or 'file:///' in url):
                        continue
                        
                    # Clean URL if needed
                    url = url.strip()
                    
                    try:
                        # Cell to apply hyperlink
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        if hyperlinks_as_formulas:
                            # Use Excel HYPERLINK formula
                            display_text = url
                            if len(display_text) > 50:
                                display_text = display_text[:47] + "..."
                            
                            cell.value = f'=HYPERLINK("{url}","{display_text}")'
                        else:
                            # Use openpyxl hyperlink object
                            cell.hyperlink = url
                            cell.value = url
                            
                            # Style for hyperlink
                            cell.font = Font(color="0563C1", underline="single")
                        
                        total_urls_processed += 1
                    except Exception as hyperlink_err:
                        logger.warning(f"Error adding hyperlink in row {row_idx}, col {col}: {hyperlink_err}")
                        # Keep original text if hyperlink fails
                        cell.value = url
                        
        logger.info(f"Processed link columns as plain text. Found {total_urls_processed} URLs across link columns.")
    except Exception as e:
        logger.warning(f"Error processing hyperlinks: {e}")
        logger.debug(traceback.format_exc())

def _add_header_footer(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    """Adds standard header and footer."""
    try:
        # Check if header_footer attribute exists (some versions don't support it)
        if hasattr(worksheet, 'header_footer'):
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
            worksheet.header_footer.center_header.text = "가격 비교 결과"
            worksheet.header_footer.right_header.text = f"생성일: {current_date}"
            worksheet.header_footer.left_footer.text = "해오름 RPA 가격 비교"
            worksheet.header_footer.right_footer.text = "페이지 &P / &N"
            logger.debug("Added header and footer to worksheet")
        else:
            logger.warning("Header/footer not supported in this Excel version - skipping")
    except Exception as e:
        logger.warning(f"Could not set header/footer: {e}")

def _apply_table_format(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    """Applies Excel table formatting to the data range."""
    # 테이블 서식 적용 함수 비우기 - 필터 적용 방지
    logger.debug("Table formatting skipped as requested.")
    return

def verify_image_data(img_value, img_col_name):
    """Helper function to verify and format image data for the Excel output."""
    try:
        # If the value is a string that looks like a dictionary (from JSON)
        if isinstance(img_value, str) and img_value.startswith('{') and img_value.endswith('}'):
            try:
                # Convert string representation to actual dictionary
                import ast
                img_dict = ast.literal_eval(img_value)
                if isinstance(img_dict, dict):
                    # Return the parsed dictionary for further processing
                    return img_dict
            except (SyntaxError, ValueError):
                # If parsing fails, treat as a regular string
                pass

        # Handle dictionary format (expected for all image sources)
        if isinstance(img_value, dict):
            # If there's a local_path, verify it exists
            if 'local_path' in img_value and img_value['local_path']:
                local_path = img_value['local_path']
                if os.path.exists(local_path) and os.path.getsize(local_path) > 0:
                    return img_value  # Return the valid dictionary

            # If no valid local_path but URL exists, keep the dictionary for the URL
            if 'url' in img_value and img_value['url']:
                 return img_value  # Return dictionary with just URL

            return '-'  # No valid path or URL

        # Handle string path/URL
        elif isinstance(img_value, str) and img_value and img_value != '-':
            img_value = img_value.strip()
            
            # Determine source from column name
            source_map = {
                '본사': 'haoreum',
                '고려': 'kogift',
                '네이버': 'naver'
            }
            
            source = 'other'
            for key, value in source_map.items():
                if key in img_col_name:
                    source = value
                    break

            # For URL strings (not file paths)
            if img_value.startswith(('http:', 'https:')):
                # Return a dictionary format for consistency
                return {'url': img_value, 'source': source}

            # Fix backslashes in path
            if '\\' in img_value:
                img_value = img_value.replace('\\', '/')

            # For file path strings (absolute paths preferred)
            if os.path.isabs(img_value) and os.path.exists(img_value) and os.path.getsize(img_value) > 0:
                # Convert file path to dictionary format for consistency
                img_value_str = img_value.replace(os.sep, '/')
                placeholder_url = f"file:///{img_value_str}"
                return {
                    'url': placeholder_url, 
                    'local_path': img_value, 
                    'original_path': img_value, 
                    'source': source
                }

            # Handle relative paths by checking multiple base directories
            elif not os.path.isabs(img_value):
                # Try different base paths based on source type
                base_paths = []
                if source == 'haoreum':
                    base_paths = [
                        IMAGE_MAIN_DIR / HAEREUM_DIR_NAME,
                        IMAGE_MAIN_DIR / 'Target' / HAEREUM_DIR_NAME, # Assuming Target exists
                        IMAGE_MAIN_DIR # Fallback
                    ]
                elif source == 'kogift':
                    base_paths = [
                        IMAGE_MAIN_DIR / KOGIFT_DIR_NAME,
                        IMAGE_MAIN_DIR / 'Target' / KOGIFT_DIR_NAME,
                        IMAGE_MAIN_DIR
                    ]
                elif source == 'naver':
                    base_paths = [
                        IMAGE_MAIN_DIR / NAVER_DIR_NAME,
                        IMAGE_MAIN_DIR / 'Target' / NAVER_DIR_NAME,
                        IMAGE_MAIN_DIR
                    ]
                else: # source == 'other'
                    base_paths = [
                        IMAGE_MAIN_DIR / OTHER_DIR_NAME,
                        IMAGE_MAIN_DIR / 'Target' / OTHER_DIR_NAME,
                        IMAGE_MAIN_DIR # General fallback
                    ]
                
                # Try each base path for resolving the relative path
                for base_path in base_paths:
                    try:
                        abs_path = (base_path / img_value).resolve()
                        if abs_path.exists() and abs_path.stat().st_size > 0:
                            abs_path_str = str(abs_path).replace('\\', '/')
                            placeholder_url = f"file:///{abs_path_str}"
                            return {
                                'url': placeholder_url, 
                                'local_path': str(abs_path), 
                                'original_path': str(abs_path), 
                                'source': source
                            }
                    except Exception:
                        continue  # Try next base path
                
                # If we reach here, all base paths failed
                if source == 'haoreum':
                    # For Haoreum, try to check if the image might be in a common format
                    try:
                        # Try standard haereum image format
                        standard_paths = [
                            # Common Haoreum image patterns using the base dir constant
                            IMAGE_MAIN_DIR / HAEREUM_DIR_NAME / f"haoreum_{os.path.basename(img_value)}",
                            IMAGE_MAIN_DIR / HAEREUM_DIR_NAME / f"haoreum_{img_value}",
                            IMAGE_MAIN_DIR / HAEREUM_DIR_NAME / os.path.basename(img_value)
                        ]
                        
                        for std_path in standard_paths:
                            std_path_str = str(std_path).replace('\\', '/') # Ensure correct format for URL
                            if os.path.exists(std_path_str) and os.path.getsize(std_path_str) > 0:
                                placeholder_url = f"file:///{std_path_str}"
                                return {
                                    'url': placeholder_url,
                                    'local_path': std_path_str,
                                    'original_path': img_value, # Keep original value as original_path
                                    'source': 'haoreum'
                                }
                    except Exception:
                        pass  # Ignore errors in this speculative search

            # If all attempts fail, return the original string for further handling
            return {'original_path': img_value, 'source': source}

        return '-'  # None, NaN, empty string, etc.
    except Exception as e:
        logging.warning(f"Error verifying image data '{str(img_value)[:100]}...' for column {img_col_name}: {e}")
        return '-'  # Return placeholder on error

def _prepare_data_for_excel(df: pd.DataFrame, skip_images=False) -> pd.DataFrame:
    """
    Prepares the DataFrame for Excel output: column order, formatting.
    """
    # Make a copy to avoid modifying the original
    df = df.copy()

    # 1) Rename columns EARLY so that original names are preserved before we drop/reorder columns
    df.rename(columns=COLUMN_RENAME_MAP, inplace=True, errors='ignore')

    # 2) Ensure all required columns from FINAL_COLUMN_ORDER exist
    for col in FINAL_COLUMN_ORDER:
        if col not in df.columns:
            df[col] = ""
            logger.debug(f"Added missing column '{col}' to DataFrame before ordering.")

    # 3) Re-order columns based on FINAL_COLUMN_ORDER (keep only expected columns)
    df = df[[col for col in FINAL_COLUMN_ORDER if col in df.columns]]

    # For upload file, modify image column values to be web URLs or empty
    if skip_images:
        # Image columns now use new names from FINAL_COLUMN_ORDER / IMAGE_COLUMNS constant
        # final_image_columns = ['해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)'] # Already defined
        image_columns = [col for col in df.columns if col in IMAGE_COLUMNS] # Use the constant

        for col in image_columns:
            # Replace image dict/path with web URL or empty string for upload file
            df[col] = df[col].apply(
                lambda x:
                    # Case 1: Input is a dictionary with 'url' key
                    x['url'] if isinstance(x, dict) and 'url' in x and isinstance(x['url'], str) and x['url'].startswith(('http://', 'https://'))
                    # Case 2: Input is a string that is already a web URL
                    else (x if isinstance(x, str) and x.startswith(('http://', 'https://'))
                    # Case 3: Anything else (dict without web URL, local path, file://, other types, None)
                    else '')
                if pd.notna(x) else ''
            )
        logger.debug(f"Processed image columns for upload file, keeping only web URLs: {image_columns}")

    # Format numeric columns (prices, quantities) using new names
    # numeric_keywords removed, using specific lists instead
    for col in df.columns:
        if any(keyword in col for keyword in ['단가', '가격', '수량']):
            try:
                # Attempt conversion, handle errors gracefully
                original_dtype = df[col].dtype
                df[col] = pd.to_numeric(df[col], errors='coerce')
                # Only fillna if conversion was successful (result is numeric)
                if pd.api.types.is_numeric_dtype(df[col]):
                    df[col] = df[col].fillna('') # Fill NaN with empty string for Excel
                else:
                     # If coercion failed, maybe revert or log? Keep original if not numeric.
                     df[col] = df[col].astype(original_dtype) # Revert if conversion failed badly
                     df[col] = df[col].fillna('') # Still fill NaNs
            except Exception as e:
                logging.warning(f"Error formatting numeric column '{col}': {str(e)}")
                df[col] = df[col].fillna('') # Ensure NaNs are handled even on error
    
    logger.debug(f"Final columns for Excel output: {df.columns.tolist()}")
    return df

def safe_excel_operation(func):
    """
    데코레이터: Excel 작업 중 발생할 수 있는 예외를 안전하게 처리합니다.
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logging.error(f"Excel operation failed in {func.__name__}: {str(e)}", exc_info=True)
            return False
    return wrapper

# --- Main Public Function --- #

def flatten_nested_image_dicts(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten any nested dictionaries in image data structures to prevent Excel conversion errors.
    """
    df_result = df.copy()
    
    # Define image-related columns (can be customized)
    image_cols = ['본사 이미지', '고려기프트 이미지', '네이버 이미지', '해오름 이미지 URL']
    # Add any columns that might contain image data
    image_cols.extend([col for col in df.columns if '이미지' in col])
    # Remove duplicates while preserving order
    image_cols = list(dict.fromkeys(image_cols))
    
    # Only process columns that actually exist in the dataframe
    image_cols = [col for col in image_cols if col in df.columns]
    
    if not image_cols:
        return df_result  # No processing needed
    
    for col in image_cols:
        for idx in df_result.index:
            value = df_result.loc[idx, col]
            
            # Process dictionaries
            if isinstance(value, dict):
                # Handle nested URL
                if 'url' in value and isinstance(value['url'], dict):
                    if 'url' in value['url']:
                        value['original_nested_url'] = value['url']  # Save original for reference
                        value['url'] = value['url']['url']  # Extract the inner URL
                
                # Ensure all dictionary values are strings if needed
                for k, v in list(value.items()):
                    if isinstance(v, dict):
                        value[k] = str(v)
            
            # Process the special case where the entire cell is a URL dict
            elif isinstance(value, str) and value.startswith('{') and value.endswith('}'):
                try:
                    # Try to parse as JSON (though this is rarely the issue)
                    import json
                    json_value = json.loads(value.replace("'", '"'))
                    if isinstance(json_value, dict) and 'url' in json_value:
                        df_result.at[idx, col] = json_value['url']
                except:
                    # If parsing fails, just keep the original
                    pass
    
    return df_result


@safe_excel_operation
def create_split_excel_outputs(df_finalized: pd.DataFrame, output_path_base: str, input_filename: str = None) -> tuple:
    """
    Create separate Excel output files for normal view and upload.
    
    Args:
        df_finalized: Final DataFrame with all data
        output_path_base: Base path for output files
        input_filename: Input filename for reference (optional)
    
    Returns:
        Tuple of (result_path, upload_path)
    """
    try:
        # Setup output file paths
        if input_filename:
            # Extract base name without extension
            input_base = os.path.splitext(os.path.basename(input_filename))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_path = os.path.join(output_path_base, f"{input_base}_result_{timestamp}.xlsx")
            upload_path = os.path.join(output_path_base, f"{input_base}_upload_{timestamp}.xlsx")
        else:
            # Use a timestamp if no input filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_path = os.path.join(output_path_base, f"result_{timestamp}.xlsx")
            upload_path = os.path.join(output_path_base, f"upload_{timestamp}.xlsx")
        
        # Set up configuration for naver image handling
        try:
            config = configparser.ConfigParser()
            config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'config.ini')
            config.read(config_path, encoding='utf-8')
            
            # Get lenient validation setting from config
            lenient_naver_validation = config.getboolean('ImageFiltering', 'lenient_naver_validation', fallback=True)
        except Exception as e:
            logger.warning(f"Could not read lenient_naver_validation from config: {e}. Using default (True).")
            lenient_naver_validation = True
        
        # Ensure result path exists
        result_dir = os.path.dirname(result_path)
        if not os.path.exists(result_dir):
            os.makedirs(result_dir)
            logger.info(f"Created directory: {result_dir}")
            
        # Create copy of DataFrame for processing
        df_for_excel = _prepare_data_for_excel(df_finalized)
        
        # -----------------------------------------
        # 1. Create Result File (with all data and images)
        # -----------------------------------------
        try:
            with pd.ExcelWriter(
                result_path, 
                engine='openpyxl',
                mode='w'
            ) as writer:
                # Convert DataFrame to Excel
                df_for_excel.to_excel(writer, index=False)
                
                # Get the worksheet
                worksheet = writer.sheets['Sheet1']
                
                # Set lenient validation flag on worksheet
                worksheet._lenient_naver_validation = lenient_naver_validation
                
                # Apply Excel styles and formatting
                apply_excel_styles(worksheet, df_for_excel)
                
                # FIXED: Ensure filter is removed before adding images
                if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
                    worksheet.auto_filter.ref = None
                    logger.info("Removed filter from result Excel file before adding images")
            
            # Save the workbook first without images
            logger.info(f"Created initial result file (no images): {result_path}")
            
            # Now add images to the saved file if any image columns exist
            try:
                workbook_with_images = openpyxl.load_workbook(result_path)
                worksheet_with_images = workbook_with_images.active
                
                # Set lenient validation flag on worksheet with images too
                worksheet_with_images._lenient_naver_validation = lenient_naver_validation
                
                # Process image columns
                images_added = 0
                
                # Find columns that have image data
                image_columns = []
                for col_idx, column in enumerate(df_for_excel.columns, start=1):
                    col_name = str(column)
                    if '이미지' in col_name:
                        image_columns.append(col_idx)
                
                if image_columns:
                    for row_idx, row in enumerate(df_for_excel.itertuples(index=False), start=2):
                        for col_idx in image_columns:
                            col_name = df_for_excel.columns[col_idx-1]
                            img_value = df_for_excel.iloc[row_idx-2, col_idx-1]
                            
                            if not isinstance(img_value, dict):
                                continue
                                
                            # Check for local_path in the dictionary
                            img_path = img_value.get('local_path', None)
                            has_url = 'url' in img_value and img_value['url']
                            
                            if img_path and os.path.isfile(img_path):
                                try:
                                    # Create and add image to cell
                                    img = Image.open(img_path)
                                    
                                    # Resize image if too large (max 150x150 pixels)
                                    max_width, max_height = 150, 150
                                    if img.width > max_width or img.height > max_height:
                                        # Calculate new dimensions maintaining aspect ratio
                                        width_ratio = max_width / img.width
                                        height_ratio = max_height / img.height
                                        ratio = min(width_ratio, height_ratio)
                                        
                                        new_width = int(img.width * ratio)
                                        new_height = int(img.height * ratio)
                                        
                                        # Resize the image
                                        img = img.resize((new_width, new_height), RESAMPLING_FILTER)
                                    
                                    # Save to a BytesIO object to create openpyxl image
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format=img.format or 'PNG')
                                    img_byte_arr.seek(0)
                                    
                                    # Create openpyxl image
                                    excel_img = openpyxl.drawing.image.Image(img_byte_arr)
                                    
                                    # Add to worksheet at the cell position
                                    cell_coord = f"{get_column_letter(col_idx)}{row_idx}"
                                    excel_img.anchor = cell_coord
                                    worksheet_with_images.add_image(excel_img)
                                    
                                    # Count image
                                    images_added += 1
                                    
                                    # Adjust row height to fit image
                                    row_height = min(max_height, max(112.5, worksheet_with_images.row_dimensions[row_idx].height))
                                    worksheet_with_images.row_dimensions[row_idx].height = row_height
                                except Exception as e:
                                    logger.error(f"Error adding image from {img_path}: {e}")
                                    # Fallback to displaying URL if available
                                    if has_url:
                                        cell = worksheet_with_images.cell(row=row_idx, column=col_idx)
                                        cell.value = img_value['url']
                                        cell.hyperlink = img_value['url']
                                        cell.font = Font(color="0563C1", underline="single")
                                        logger.debug(f"Falling back to URL display for failed image: {img_value['url'][:50]}...")
                            else:
                                # No valid local file, try to add as URL if available
                                try:
                                    if has_url:
                                        cell = worksheet_with_images.cell(row=row_idx, column=col_idx)
                                        cell.value = img_value['url']
                                        cell.hyperlink = img_value['url']
                                        cell.font = Font(color="0563C1", underline="single")
                                        logger.debug(f"Falling back to URL display for failed image: {img_value['url'][:50]}...")
                                except Exception as e:
                                    logger.error(f"Error creating/configuring image object for {img_path}: {e}")
                                    # Fallback to displaying URL if available
                                    if has_url:
                                        cell = worksheet_with_images.cell(row=row_idx, column=col_idx)
                                        cell.value = img_value['url']
                                        cell.hyperlink = img_value['url']
                                        cell.font = Font(color="0563C1", underline="single")
                
                # FIXED: Ensure filter is removed after image addition too
                if hasattr(worksheet_with_images, 'auto_filter') and worksheet_with_images.auto_filter:
                    worksheet_with_images.auto_filter.ref = None
                    logger.info("Removed filter from result Excel file after adding images")
                
                # Clean Naver images and related data for the workbook with images as well
                clean_naver_images_and_data(worksheet_with_images, df_for_excel)
                
                # Save the workbook with images
                workbook_with_images.save(result_path)
                logger.info(f"Successfully added {images_added} images to result file")
            except Exception as img_err:
                logger.error(f"Error adding images to result file: {img_err}")
                # Continue with the file without images
            
            result_success = True
            logger.info(f"Successfully created result file: {result_path}")
            
        except Exception as e:
            logger.error(f"Error creating result file: {e}")
            logger.debug(traceback.format_exc())
            result_success = False
        
        # -----------------------------------------
        # 2. Create Upload File (with links only and different column names)
        # -----------------------------------------
        try:
            logger.info(f"Preparing data for upload file: {upload_path}")
            
            # Create a deep copy of the original DataFrame to avoid modifying it
            df_upload = pd.DataFrame()
            
            # FIX: First extract image URLs from the DataFrame before column mapping
            # Create a new DataFrame with the image URLs
            df_with_image_urls = df_finalized.copy()
            
            # Process each image column to extract only web URLs
            for img_col in IMAGE_COLUMNS:
                if img_col in df_finalized.columns:
                    logger.info(f"Extracting image URLs from {img_col} column...")
                    
                    # Map the column names to the upload file column names
                    upload_img_col = COLUMN_MAPPING_FINAL_TO_UPLOAD.get(img_col, img_col) # Use mapping, fallback to original if not found

                    # Create the target upload column if it doesn't exist in the intermediate df
                    if upload_img_col not in df_with_image_urls.columns:
                        df_with_image_urls[upload_img_col] = ""
                    
                    # Track URL extraction results for this column
                    urls_found = 0
                    fallback_urls_generated = 0
                    url_errors = 0
                    
                    # Process all rows to extract image URLs
                    for idx in df_finalized.index:
                        value = df_finalized.at[idx, img_col]
                        
                        # Default to empty string (not "-") to avoid showing placeholders in cells
                        image_url = ""
                        url_source = "unknown"
                        
                        try:
                            # Check dictionary structure and extract 'url' key if it's a web URL
                            if isinstance(value, dict):
                                # 1. Check for product_url first (for Naver)
                                if 'product_url' in value and isinstance(value['product_url'], str) and value['product_url'].startswith(('http://', 'https://')):
                                    image_url = value['product_url'].strip()
                                    url_source = "direct_from_product_url_key"
                                    urls_found += 1
                                    logger.debug(f"Found product URL in {img_col} at idx {idx} using 'product_url' key: {image_url[:50]}...")
                                # 2. Check for original_crawled_url added by our improvements
                                elif 'original_crawled_url' in value and isinstance(value['original_crawled_url'], str) and value['original_crawled_url'].startswith(('http://', 'https://')):
                                    image_url = value['original_crawled_url'].strip()
                                    url_source = "from_original_crawled_url_key"
                                    urls_found += 1
                                    logger.debug(f"Found original crawled URL in {img_col} at idx {idx}: {image_url[:50]}...")
                                # 3. Check for regular 'url' key (including placeholders, but preferring real URLs)
                                elif 'url' in value and isinstance(value['url'], str) and value['url'].startswith(('http://', 'https://')):
                                    # Try to use real URLs, but for placeholders check for original_crawled_url
                                    if value['url'].startswith('http://placeholder.url/'):
                                        # For placeholder URLs, check if we have an original crawled URL to use instead
                                        if 'original_crawled_url' in value and isinstance(value['original_crawled_url'], str) and value['original_crawled_url'].startswith(('http://', 'https://')) and not value['original_crawled_url'].startswith('http://placeholder.url/'):
                                            image_url = value['original_crawled_url'].strip()
                                            url_source = "from_original_crawled_url_key"
                                            urls_found += 1
                                            logger.debug(f"Found original crawled URL in {img_col} at idx {idx}: {image_url[:50]}...")
                                        else:
                                            # Don't use placeholder URLs
                                            # Try to construct a better URL based on image source
                                            source = value.get('source', '').lower()
                                            
                                            # For Kogift images
                                            if source == 'kogift' or '고려' in img_col:
                                                # If we have a product_id, try to construct URL
                                                if 'product_id' in value and value['product_id']:
                                                    product_id = value['product_id']
                                                    constructed_url = f"https://koreagift.com/ez/upload/mall/shop_{product_id}.jpg"
                                                    image_url = constructed_url
                                                    url_source = "constructed_kogift_url"
                                                    urls_found += 1
                                                    logger.debug(f"Constructed URL for Kogift image: {image_url}")
                                                # If we have product link in a different column, use it
                                                elif '고려기프트 상품링크' in df_finalized.columns:
                                                    product_link = df_finalized.at[idx, '고려기프트 상품링크']
                                                    if isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                                                        # Try to extract product ID from link
                                                        import re
                                                        no_match = re.search(r'no=(\d+)', product_link)
                                                        if no_match:
                                                            product_id = no_match.group(1)
                                                            constructed_url = f"https://koreagift.com/ez/upload/mall/shop_{product_id}.jpg"
                                                            image_url = constructed_url
                                                            url_source = "constructed_from_product_link"
                                                            urls_found += 1
                                                            logger.debug(f"Constructed URL from product link: {image_url}")
                                                        else:
                                                            # If we can't extract product ID, use product link as fallback
                                                            image_url = product_link
                                                            url_source = "kogift_product_link_fallback"
                                                            urls_found += 1
                                                            logger.debug(f"Using product link for Kogift image: {image_url[:50]}...")
                                            
                                            # For Naver images
                                            elif source == 'naver' or '네이버' in img_col:
                                                # Try to use product_id to construct URL
                                                if 'product_id' in value and value['product_id']:
                                                    product_id = value['product_id']
                                                    constructed_url = f"https://shopping-phinf.pstatic.net/main_{product_id}/{product_id}.jpg"
                                                    image_url = constructed_url
                                                    url_source = "constructed_naver_url"
                                                    urls_found += 1
                                                    logger.debug(f"Constructed URL for Naver image: {image_url}")
                                                # Check if we have product link in a different column
                                                elif '네이버 쇼핑 링크' in df_finalized.columns:
                                                    product_link = df_finalized.at[idx, '네이버 쇼핑 링크']
                                                    if isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                                                        image_url = product_link
                                                        url_source = "naver_product_link_fallback"
                                                        urls_found += 1
                                                        logger.debug(f"Using product link for Naver image: {image_url[:50]}...")
                                            
                                            # If still no URL, use original placeholder as last resort
                                            if not image_url:
                                                image_url = value['url'].strip()
                                                url_source = "placeholder_url_key"
                                                logger.debug(f"Using placeholder URL in {img_col} at idx {idx}: {image_url[:50]}...")
                                    else:
                                        image_url = value['url'].strip()
                                        url_source = "direct_from_url_key"
                                        urls_found += 1
                                        logger.debug(f"Found web URL in {img_col} at idx {idx} using 'url' key: {image_url[:50]}...")
                                # 4. Check for original_url key for original crawled URLs
                                elif 'original_url' in value and isinstance(value['original_url'], str) and value['original_url'].startswith(('http://', 'https://')):
                                    image_url = value['original_url'].strip()
                                    url_source = "from_original_url_key"
                                    urls_found += 1
                                    logger.debug(f"Found original URL in {img_col} at idx {idx} using 'original_url' key: {image_url[:50]}...")
                                # 5. Check if original_path is a URL
                                elif 'original_path' in value and isinstance(value['original_path'], str) and value['original_path'].startswith(('http://', 'https://')):
                                    image_url = value['original_path'].strip()
                                    url_source = "from_original_path_as_url"
                                    urls_found += 1
                                    logger.debug(f"Found URL in original_path for {img_col} at idx {idx}: {image_url[:50]}...")
                                # 6. Try to generate a URL based on the source and other available information
                                elif not image_url:
                                    source = value.get('source', '').lower()
                                    
                                    # For Haereum images, try to extract product code
                                    if source == 'haereum' or '본사' in img_col:
                                        if 'p_idx' in value:
                                            product_code = value['p_idx']
                                            constructed_url = f"https://www.jclgift.com/upload/product/bimg3/{product_code}b.jpg"
                                            image_url = constructed_url
                                            url_source = "constructed_haereum_url"
                                            urls_found += 1
                                            logger.debug(f"Constructed URL for Haereum image: {image_url}")
                                        elif 'original_path' in value and isinstance(value['original_path'], str):
                                            path = value['original_path']
                                            # Look for product code pattern (e.g., BBCA0009349, CCBK0001873)
                                            code_match = re.search(r'([A-Z]{4}\d{7})', path)
                                            if code_match:
                                                product_code = code_match.group(1)
                                                constructed_url = f"https://www.jclgift.com/upload/product/bimg3/{product_code}b.jpg"
                                                image_url = constructed_url
                                                url_source = "constructed_haereum_url_from_path"
                                                urls_found += 1
                                                logger.debug(f"Constructed URL for Haereum image from path: {image_url}")
                                            
                                    # For Kogift images
                                    elif source == 'kogift' or '고려' in img_col:
                                        if '고려기프트 상품링크' in df_finalized.columns:
                                            product_link = df_finalized.at[idx, '고려기프트 상품링크']
                                            if isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                                                # Try to extract product ID from link
                                                import re
                                                no_match = re.search(r'no=(\d+)', product_link)
                                                if no_match:
                                                    product_id = no_match.group(1)
                                                    constructed_url = f"https://koreagift.com/ez/upload/mall/shop_{product_id}.jpg"
                                                    image_url = constructed_url
                                                    url_source = "constructed_from_product_link"
                                                    urls_found += 1
                                                    logger.debug(f"Constructed URL from product link: {image_url}")
                                                else:
                                                    # If we can't extract product ID, use product link as fallback
                                                    image_url = product_link
                                                    url_source = "kogift_product_link_fallback"
                                                    urls_found += 1
                                                    logger.debug(f"Using product link for Kogift image: {image_url[:50]}...")
                                    
                                    # For Naver images
                                    elif source == 'naver' or '네이버' in img_col:
                                        if '네이버 쇼핑 링크' in df_finalized.columns:
                                            product_link = df_finalized.at[idx, '네이버 쇼핑 링크']
                                            if isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                                                image_url = product_link
                                                url_source = "naver_product_link_fallback"
                                                urls_found += 1
                                                logger.debug(f"Using product link for Naver image: {image_url[:50]}...")
                                
                                # 7. Check other potential keys if still no URL found
                                if not image_url:
                                    for url_key in ['image_url', 'src', 'product_link']:
                                        fallback_url = value.get(url_key)
                                        if fallback_url and isinstance(fallback_url, str) and fallback_url.startswith(('http://', 'https://')):
                                            image_url = fallback_url.strip()
                                            url_source = f"fallback_from_{url_key}"
                                            urls_found += 1
                                            logger.debug(f"Found web URL in {img_col} at idx {idx} using fallback key '{url_key}': {image_url[:50]}...")
                                            break # Stop checking keys once a valid URL is found
                            
                            # Special handling for non-dictionary values (direct strings)
                            elif isinstance(value, str) and value and value != '-':
                                if value.startswith(('http://', 'https://')):
                                    # Direct URL string
                                    image_url = value.strip()
                                    url_source = "direct_string_url"
                                    urls_found += 1
                                    logger.debug(f"Found direct string URL in {img_col} at idx {idx}: {image_url[:50]}...")
                                elif os.path.exists(value):
                                    # Convert local file path to URL
                                    if '본사' in img_col or 'haereum' in img_col.lower():
                                        # Try to extract product code from path
                                        code_match = re.search(r'([A-Z]{4}\d{7})', value)
                                        if code_match:
                                            product_code = code_match.group(1)
                                            image_url = f"https://www.jclgift.com/upload/product/bimg3/{product_code}b.jpg"
                                            url_source = "converted_haereum_path_to_url"
                                            urls_found += 1
                                            logger.debug(f"Converted Haereum path to URL: {image_url}")
                                    elif '고려' in img_col or 'kogift' in img_col.lower():
                                        # Try to extract product info from path
                                        basename = os.path.basename(value)
                                        if basename.startswith('shop_'):
                                            product_id = basename.replace('shop_', '').split('.')[0]
                                            image_url = f"https://koreagift.com/ez/upload/mall/{basename}"
                                            url_source = "converted_kogift_path_to_url"
                                            urls_found += 1
                                            logger.debug(f"Converted Kogift path to URL: {image_url}")
                            
                            # Special fallback for specific columns if URL is still empty
                            if not image_url:
                                # For Kogift images, try using the product link
                                if '고려' in img_col and '고려기프트 상품링크' in df_finalized.columns:
                                    product_link = df_finalized.at[idx, '고려기프트 상품링크']
                                    if isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                                        image_url = product_link
                                        url_source = "final_kogift_product_link_fallback"
                                        urls_found += 1
                                        logger.debug(f"Using product link as final fallback for Kogift: {image_url[:50]}...")
                                
                                # For Naver images, try using the product link
                                elif '네이버' in img_col and '네이버 쇼핑 링크' in df_finalized.columns:
                                    product_link = df_finalized.at[idx, '네이버 쇼핑 링크']
                                    if isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                                        image_url = product_link
                                        url_source = "final_naver_product_link_fallback"
                                        urls_found += 1
                                        logger.debug(f"Using product link as final fallback for Naver: {image_url[:50]}...")
                        except Exception as e:
                            logger.error(f"이미지 URL 추출 중 오류 발생 (행 {idx+1}, {img_col}): {str(e)[:100]}")
                            image_url = ""  # Reset on error
                            url_errors += 1
                        
                        # Store the extracted image URL in the intermediate DataFrame under the UPLOAD column name
                        df_with_image_urls.at[idx, upload_img_col] = image_url if image_url else "" # Use empty string if no valid URL

                    # Log summary for this column
                    logger.info(f"URL 추출 결과 ({upload_img_col}): 총 {urls_found}개 URL 추출 성공, {url_errors}개 오류")

            # Special handling for Naver image column - recover original URLs and handle placeholders
            df_with_image_urls = prepare_naver_image_urls_for_upload(df_with_image_urls)
            
            # Special handling for Kogift image column - recover original URLs
            df_with_image_urls = prepare_kogift_image_urls_for_upload(df_with_image_urls)

            # Check if we extracted any image URLs
            for img_col in ['해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)']: # Use upload file column names
                if img_col in df_with_image_urls.columns:
                    url_count = (df_with_image_urls[img_col].astype(str).str.startswith(('http://', 'https://'))).sum()
                    logger.info(f"Extracted {url_count} image URLs for {img_col} in upload data")
                    
                    # Log a few examples if any URLs were found
                    if url_count > 0:
                        sample_urls = df_with_image_urls[df_with_image_urls[img_col].astype(str).str.startswith(('http://', 'https://'))][img_col].head(3).tolist()
                        logger.info(f"Sample image URLs for {img_col}: {sample_urls}")

            # Map columns from result format to upload format 
            df_upload = pd.DataFrame() # Start with an empty DataFrame for the upload file
            
            for target_col in UPLOAD_COLUMN_ORDER:
                # Find corresponding source column from the original result format
                source_col = None
                for result_col, upload_col in COLUMN_MAPPING_FINAL_TO_UPLOAD.items():
                    if upload_col == target_col:
                        source_col = result_col
                        break
                
                # Determine where to get the data:
                # - If it's an upload image link column, get it from df_with_image_urls[target_col]
                # - Otherwise, get it from df_finalized[source_col]
                
                if target_col in ['해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)']:
                    # Get the already processed image URL from df_with_image_urls
                    if target_col in df_with_image_urls.columns:
                        df_upload[target_col] = df_with_image_urls[target_col]
                    else:
                        df_upload[target_col] = '' # Should not happen, but safety check
                        logger.warning(f"Processed image URL column '{target_col}' not found in intermediate df.")
                elif source_col and source_col in df_finalized.columns:
                    # Get non-image data from the original finalized DataFrame
                    df_upload[target_col] = df_finalized[source_col]
                else:
                    # If no matching column found, add an empty column
                    df_upload[target_col] = ''
                    logger.warning(f"Could not find source column for upload column '{target_col}' or source column missing.")

            # Log image columns in the final upload file to confirm extraction worked
            for img_col in ['해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)']:
                if img_col in df_upload.columns:
                    non_empty = df_upload[img_col].astype(str).str.strip().str.len() > 0
                    non_empty_urls = df_upload[img_col].astype(str).str.startswith(('http://', 'https://'))
                    count = non_empty.sum()
                    url_count = non_empty_urls.sum()
                    logger.info(f"Upload file: {img_col} column has {count} non-empty values, {url_count} are URLs")
                    
                    # Log sample values
                    if url_count > 0:
                        sample_values = df_upload.loc[non_empty_urls, img_col].head(3).tolist()
                        logger.info(f"Sample URL values in final upload df: {sample_values}")
                    elif count > 0:
                         sample_non_urls = df_upload.loc[non_empty & ~non_empty_urls, img_col].head(3).tolist()
                         logger.warning(f"Sample non-URL values in final upload df {img_col}: {sample_non_urls}")

            # Create new workbook for upload file (now with properly extracted image URLs)
            workbook_upload = openpyxl.Workbook()
            worksheet_upload = workbook_upload.active
            worksheet_upload.title = "제품 가격 비교 (업로드용)"
            
            logger.info(f"Writing upload file (with image links): {upload_path} with {len(df_upload)} rows.")
            
            # Write header
            for col_idx, col_name in enumerate(df_upload.columns, 1):
                worksheet_upload.cell(row=1, column=col_idx, value=col_name)
            
            # Write data (now with properly extracted image URLs)
            for row_idx, row in enumerate(df_upload.itertuples(), 2):
                for col_idx, value in enumerate(row[1:], 1):  # Skip the index
                    # Handle NaN/None values
                    if pd.isna(value) or value is None:
                        cell_value = ""
                    else:
                        cell_value = value
                    
                    # Write the cell value
                    worksheet_upload.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Add direct verification of image URL columns for logging
                if row_idx <= 5:  # Log only the first 5 rows
                    img_idx_start = len(UPLOAD_COLUMN_ORDER) - 3
                    for i in range(3):  # Check all 3 image columns
                        col_idx = img_idx_start + i + 1
                        cell_value = worksheet_upload.cell(row=row_idx, column=col_idx).value
                        col_name = UPLOAD_COLUMN_ORDER[img_idx_start + i]
                        if cell_value and isinstance(cell_value, str) and cell_value.startswith(('http://', 'https://')):
                            logger.debug(f"Row {row_idx}, {col_name}: Valid URL found: {cell_value[:50]}...")
                        else:
                            logger.debug(f"Row {row_idx}, {col_name}: No valid URL found. Value: '{cell_value}'")
            
            # Apply upload file specific formatting (new function with requested formatting)
            _apply_upload_file_formatting(worksheet_upload, df_upload.columns.tolist())
            
            # Add hyperlinks to all image URL cells
            try:
                logger.info("Adding hyperlinks to image URLs in upload file...")
                
                # Define upload image columns
                upload_image_cols = ['해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)']
                
                # Get the column indices for these columns
                col_indices = {}
                for i, col_name in enumerate(df_upload.columns, 1):
                    if col_name in upload_image_cols:
                        col_indices[col_name] = i
                
                # Add hyperlinks to the cells
                for row_idx in range(2, len(df_upload) + 2):  # Start from row 2 (after header)
                    for col_name, col_idx in col_indices.items():
                        cell = worksheet_upload.cell(row=row_idx, column=col_idx)
                        url = cell.value
                        
                        # Only add hyperlink if the cell contains a valid URL
                        if isinstance(url, str) and url.strip() and url.startswith(('http://', 'https://')):
                            cell.hyperlink = url
                            cell.font = Font(color="0563C1", underline="single")
                
                logger.info("Hyperlinks added to upload file successfully")
            except Exception as e:
                logger.warning(f"Error adding hyperlinks to upload file: {e}")
            
            # Save upload file
            workbook_upload.save(upload_path)
            upload_success = True
            logger.info(f"Successfully created upload file (with image links): {upload_path}")
                
            return result_success, upload_success, result_path, upload_path
        
        except Exception as upload_err:
            logger.error(f"Error creating upload file: {upload_err}")
            logger.debug(traceback.format_exc())
            upload_success = False
        
        return result_success, upload_success, result_path, upload_path
        
    except Exception as main_error:
        logger.error(f"Unexpected error in create_split_excel_outputs: {main_error}")
        logger.debug(traceback.format_exc())
        return False, False, None, None

@safe_excel_operation
def create_final_output_excel(df: pd.DataFrame, output_path: str) -> bool:
    """
    Create a combined Excel output file with images and various formatting.
    Unlike create_split_excel_outputs, this creates a single Excel file with advanced formatting.
    It's kept for potential direct use but create_split_excel_outputs is preferred
    for most use cases.

    Args:
        df: DataFrame with the data
        output_path: Path where to save the Excel file

    Returns:
        bool: True if successful, False otherwise
    """
    if df is None:
        logger.error("Cannot create Excel file: Input DataFrame is None.")
        return False

    logger.info(f"Starting creation of single final Excel output: {output_path}")
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # 1. Finalize the DataFrame (Rename, Order, Clean)
    df_finalized = finalize_dataframe_for_excel(df) # Use the refactored function

    if df_finalized.empty and not df.empty: # Check if finalization failed or cleared data
        logger.error("DataFrame became empty after finalization step. Cannot save Excel.")
        return False
    elif df_finalized.empty and df.empty:
        logger.warning("Input DataFrame was empty. Saving Excel file with only headers.")
        # Allow proceeding to create an empty file with headers

    # 2. Check if file is locked
    if os.path.exists(output_path):
        try:
            with open(output_path, 'a+b'):
                 pass # Check lock
        except (IOError, PermissionError) as lock_err:
             logger.error(f"Output file {output_path} is locked: {lock_err}. Cannot save.")
             # Optional: Could try alternative path like in split function
             # timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
             # alternative_path = f"{os.path.splitext(output_path)[0]}_{timestamp}{os.path.splitext(output_path)[1]}"
             # logger.warning(f"Attempting alternative path: {alternative_path}")
             # output_path = alternative_path
             # But for now, just fail if locked.
             return False


    # 3. Save finalized data to Excel using openpyxl engine
    try:
        logger.info(f"Attempting to write final Excel: {output_path} with {len(df_finalized)} rows.")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_finalized.to_excel(writer, index=False, sheet_name='Results', na_rep='')
            worksheet = writer.sheets['Results']
            workbook = writer.book # Get workbook if needed later
            logger.info(f"Data ({worksheet.max_row -1} rows) written to sheet 'Results'. Applying formatting...")

            # --- Apply Full Formatting ---
            _apply_column_widths(worksheet, df_finalized)
            _apply_cell_styles_and_alignment(worksheet, df_finalized)
            if not df_finalized.empty: # Avoid processing images on empty df
                _process_image_columns(worksheet, df_finalized)
                _adjust_image_cell_dimensions(worksheet, df_finalized)
            else:
                 logger.info("Skipping image processing and dimension adjustment for empty DataFrame.")
            _add_hyperlinks_to_worksheet(worksheet, df_finalized)
            _apply_conditional_formatting(worksheet, df_finalized)
            _setup_page_layout(worksheet)
            _add_header_footer(worksheet)
            # _apply_table_format(worksheet) # Keep disabled

        logger.info(f"Successfully created and formatted final Excel file: {output_path}")
        return True

    except PermissionError as pe:
        logger.error(f"Permission denied writing final Excel file '{output_path}'. Is it open? Error: {pe}", exc_info=True)
        return False
    except Exception as e:
        logger.error(f"Error creating final Excel file '{output_path}': {e}", exc_info=True)
        return False

def apply_excel_styles(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """
    Apply Excel styles to the worksheet.
    This is a wrapper around _apply_cell_styles_and_alignment for backward compatibility.
    """
    _apply_cell_styles_and_alignment(worksheet, df)

def _adjust_image_cell_dimensions(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Adjusts row heights and column widths for cells containing images."""
    logger.debug("Adjusting dimensions for image cells...")
    
    # Get image column indices using the IMAGE_COLUMNS constant
    image_cols = {col: idx for idx, col in enumerate(df.columns, 1) if col in IMAGE_COLUMNS}
    
    if not image_cols:
        return

    # Increase column widths for image columns to accommodate larger images
    for col_name, col_idx in image_cols.items():
        try:
            col_letter = get_column_letter(col_idx)
            # FIXED: Use larger column width for image columns
            worksheet.column_dimensions[col_letter].width = 85  # Increased from 80
        except Exception as e:
            logger.error(f"Error adjusting column width for {col_name}: {e}")
    
    # Create a set of rows that need height adjustment
    rows_with_images = set()
    
    try:
        # Find rows that have actual images (not error messages or empty cells)
        for row_idx in range(2, worksheet.max_row + 1):
            for col_name, col_idx in image_cols.items():
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # If the cell is empty, it likely has an image
                    if cell.value == "" or cell.value is None:
                        rows_with_images.add(row_idx)
                        break
                        
                    # Check for image data in dictionary format
                    cell_value = str(cell.value) if cell.value else ""
                    if cell_value and cell_value.startswith('{') and cell_value.endswith('}'):
                        try:
                            import ast
                            img_dict = ast.literal_eval(cell_value)
                            if isinstance(img_dict, dict) and ('local_path' in img_dict or 'url' in img_dict):
                                rows_with_images.add(row_idx)
                                break
                        except:
                            pass
                            
                    # Check for path-like strings
                    if (cell_value and cell_value != '-' and 
                        not any(err_msg in cell_value for err_msg in ERROR_MESSAGE_VALUES) and
                        ('\\' in cell_value or '/' in cell_value or '.jpg' in cell_value.lower() or 
                         '.png' in cell_value.lower() or '.jpeg' in cell_value.lower() or
                         'http' in cell_value.lower())):
                        rows_with_images.add(row_idx)
                        break
                except Exception as e:
                    logger.error(f"Error checking cell at row {row_idx}, column {col_idx}: {e}")
    except Exception as e:
        logger.error(f"Error finding rows with images: {e}")
    
    # Apply increased height to rows with images
    for row_idx in rows_with_images:
        try:
            # FIXED: Set larger row height to accommodate bigger images
            worksheet.row_dimensions[row_idx].height = 400  # Increased from 380
            
            # Center-align all cells in this row for better appearance with images
            for col_idx in range(1, worksheet.max_column + 1):
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    # Preserve horizontal alignment, set vertical to center
                    current_alignment = cell.alignment
                    cell.alignment = Alignment(
                        horizontal=current_alignment.horizontal,
                        vertical="center",
                        wrap_text=current_alignment.wrap_text
                    )
                except Exception as e:
                    logger.error(f"Error adjusting cell alignment at row {row_idx}, column {col_idx}: {e}")
        except Exception as e:
            logger.error(f"Error adjusting row height for row {row_idx}: {e}")
    
    logger.debug(f"Adjusted dimensions for {len(rows_with_images)} rows with images")

# --- Refactored Data Finalization ---
def finalize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    Finalizes the DataFrame for Excel output: Renames columns, ensures all required columns exist,
    sets the final column order, and applies basic type formatting.
    Assumes image data (paths or dicts) is already present.
    """
    if df is None:
        logger.error("Input DataFrame is None, cannot finalize.")
        # Return empty df with correct columns to avoid downstream errors
        return pd.DataFrame(columns=FINAL_COLUMN_ORDER)

    logger.info(f"Finalizing DataFrame for Excel. Input shape: {df.shape}")
    logger.debug(f"Input columns: {df.columns.tolist()}")
    
    # Check for duplicate column names - this can cause the 'dtype' error
    duplicate_cols = df.columns[df.columns.duplicated()].tolist()
    if duplicate_cols:
        logger.warning(f"Found {len(duplicate_cols)} duplicate column names: {duplicate_cols}")
        # Create a new DataFrame with deduplicated columns
        # For each duplicate, keep only the first instance
        unique_cols = []
        
        for col in df.columns:
            if col in unique_cols:
                # Skip this column as we already have it
                continue
            unique_cols.append(col)
        
        # Create new DataFrame with only unique columns
        df = df[unique_cols]
        logger.info(f"Removed duplicate columns. New shape: {df.shape}")
    
    # Step 1: Create a new DataFrame to avoid modifying the original
    df_final = df.copy()
    
    # Step 2: Rename columns to the target names
    df_final = df_final.rename(columns=COLUMN_RENAME_MAP, errors='ignore')
    logger.debug(f"Columns after rename: {df_final.columns.tolist()}")
    
    # FIXED: Debug logging for Kogift price data
    logger.info("Checking Kogift price data before column processing:")
    kogift_price_col = '판매단가(V포함)(2)'
    if kogift_price_col in df_final.columns:
        price_count = df_final[kogift_price_col].notnull().sum()
        logger.info(f"Found {price_count} non-null values in '{kogift_price_col}' column")
        
        # Log a few sample values
        if price_count > 0:
            sample_values = df_final[kogift_price_col].head(3).tolist()
            logger.info(f"Sample Kogift price values: {sample_values}")
    else:
        logger.warning(f"Column '{kogift_price_col}' not found in dataframe!")
    
    # Step 3: Create an output DataFrame with columns in the proper order
    output_df = pd.DataFrame()
    
    # Identify which columns in the final_order exist in the input
    available_cols = [col for col in FINAL_COLUMN_ORDER if col in df_final.columns]
    
    # Log which columns from FINAL_COLUMN_ORDER are missing
    missing_cols = [col for col in FINAL_COLUMN_ORDER if col not in df_final.columns]
    if missing_cols:
        logger.warning(f"The following columns from FINAL_COLUMN_ORDER are missing: {missing_cols}")
    
    # Copy data from original to new dataframe
    for col in available_cols:
        try:
            output_df[col] = df_final[col]
        except Exception as e:
            logger.error(f"Error copying column '{col}' during finalization: {e}")
            output_df[col] = None # Add empty column on error

    # Step 4: Add missing columns with empty values
    for col in FINAL_COLUMN_ORDER:
        if col not in output_df.columns:
            output_df[col] = None # Add missing column with None values
            logger.debug(f"Added missing column '{col}' with None values")
    
    # FIXED: Ensure the Kogift price column exists and has data copied correctly
    # First, check for the specific price column again:
    if kogift_price_col in df_final.columns and kogift_price_col in output_df.columns:
        # Explicitly copy the price column data again to ensure it's not lost
        output_df[kogift_price_col] = df_final[kogift_price_col]
        logger.info(f"Explicitly copied '{kogift_price_col}' data to ensure preservation")
        
        # Verify the copy worked
        kogift_price_count_after = output_df[kogift_price_col].notnull().sum()
        logger.info(f"After explicit copy: {kogift_price_count_after} non-null values in '{kogift_price_col}'")
    
    # FIXED: Special handling for price columns to avoid data loss
    # Check for alternate column names that might hold price data
    price_alternates = {
        '판매가(V포함)(2)': ['판매단가2(VAT포함)', '판매단가(V포함)(2)', '고려가격', '고려기프트판매가'], # Added original processing name
        '판매단가(V포함)(3)': ['판매단가3 (VAT포함)', '판매단가(V포함)(3)', '네이버가격', '네이버판매가'] # Added original processing name
    }

    logger.info("Checking for and consolidating data from alternate price columns...")
    for target_col, alt_cols in price_alternates.items():
        if target_col in output_df.columns:
            # Count valid data points in the target column currently
            target_valid_count = output_df[target_col].apply(lambda x: pd.notna(x) and x not in ['-', '']).sum()
            logger.debug(f"Target '{target_col}' currently has {target_valid_count} valid entries.")

            # Check each alternate column found in the *renamed* df_final
            best_alt_col = None
            best_alt_count = target_valid_count # Start with current count

            for alt_col_potential in alt_cols:
                # Check if this alternate exists *after renaming*
                if alt_col_potential in df_final.columns:
                    alt_valid_count = df_final[alt_col_potential].apply(lambda x: pd.notna(x) and x not in ['-', '']).sum()
                    logger.debug(f"  Checking alternate '{alt_col_potential}': Found {alt_valid_count} valid entries.")
                    # If this alternate has more valid data, consider it
                    if alt_valid_count > best_alt_count:
                        best_alt_count = alt_valid_count
                        best_alt_col = alt_col_potential

            # If a better alternate was found, copy its data to the target column
            if best_alt_col:
                logger.info(f"Found better data in alternate column '{best_alt_col}' ({best_alt_count} valid). Copying to '{target_col}'.")
                # Ensure the source column exists before copying
                if best_alt_col in df_final.columns:
                     # Copy data, converting potential errors during copy
                     try:
                          output_df[target_col] = pd.to_numeric(df_final[best_alt_col], errors='coerce')
                     except Exception as copy_err:
                          logger.error(f"Error coercing/copying from {best_alt_col} to {target_col}: {copy_err}")
                          # Fallback: copy raw data if numeric conversion fails
                          output_df[target_col] = df_final[best_alt_col]
                else:
                     logger.warning(f"Attempted to copy from non-existent column '{best_alt_col}'")

    # Step 5: Format numeric columns (Ensure this runs AFTER alternate consolidation)
    # Get image columns for exclusion from numeric formatting
    image_cols = [col for col in output_df.columns if col in IMAGE_COLUMNS]

    logger.info("Applying numeric formatting to relevant columns...")
    for col in output_df.columns:
        # Skip image columns
        if col in image_cols:
            continue

        # Check if column should be numeric (Prices, Quantities, Differences)
        # Use final column names here
        is_numeric_col = (
            col in PRICE_COLUMNS or
            col in QUANTITY_COLUMNS or
            col in PERCENTAGE_COLUMNS or
            col in ['가격차이(2)', '가격차이(3)'] # Explicitly include difference columns
        )

        if is_numeric_col:
            # Log before conversion
            pre_conversion_sample = output_df[col].head(3).tolist()
            pre_conversion_dtype = output_df[col].dtype
            logger.debug(f"Converting column '{col}' (dtype: {pre_conversion_dtype}, sample: {pre_conversion_sample}) to numeric.")

            try:
                # Store original data before attempting conversion
                original_data = output_df[col].copy()
                
                # Attempt conversion to numeric, coercing errors to NaN
                output_df[col] = pd.to_numeric(output_df[col], errors='coerce')
                
                # Log after conversion
                post_conversion_sample = output_df[col].head(3).tolist()
                post_conversion_dtype = output_df[col].dtype
                nan_count = output_df[col].isna().sum()
                logger.debug(f"  -> Post-conversion '{col}' (dtype: {post_conversion_dtype}, sample: {post_conversion_sample}, NaNs: {nan_count})")

                # If conversion resulted in all NaNs, consider reverting
                if nan_count == len(output_df[col]):
                    logger.warning(f"Numeric conversion resulted in all NaN for column '{col}'. Reverting to original data.")
                    output_df[col] = original_data

            except Exception as e:
                logger.warning(f"Error converting column '{col}' to numeric: {e}. Keeping original data.")
                # Optionally revert if needed, but pd.to_numeric usually handles errors well with coerce

    # Step 6: Replace NaN/NaT with None for Excel compatibility
    output_df = output_df.replace({pd.NA: None, np.nan: None, pd.NaT: None})

    # Step 7: Set default values for empty cells ('-' for non-image, handling None)
    logger.info("Setting default values for empty cells ('-')...")
    for col in output_df.columns:
        if col not in image_cols:
            # Apply '-' default to cells that are None or empty strings after NaN replacement
            # Ensure we don't overwrite 0 or False
            output_df[col] = output_df[col].apply(
                lambda x: '-' if (x is None or x == '') else x
            )

    # Final verification of key data
    logger.info(f"DataFrame finalized. Output shape: {output_df.shape}")
    logger.debug(f"Final columns: {output_df.columns.tolist()}")

    # Final verification of price data
    for price_col in ['판매단가(V포함)', '판매가(V포함)(2)', '판매단가(V포함)(3)', '가격차이(2)', '가격차이(3)']:
        if price_col in output_df.columns:
            try:
                 # Count non-empty, non-'', non-None values
                 non_empty_count = output_df[price_col].apply(lambda x: pd.notna(x) and x not in ['-', '']).sum()
                 # Log sample non-empty values if they exist
                 sample_values = output_df.loc[output_df[price_col].apply(lambda x: pd.notna(x) and x not in ['-', '']), price_col].head(3).tolist()
                 logger.info(f"Final check: Column '{price_col}' has {non_empty_count} valid entries. Sample: {sample_values}")
            except Exception as e:
                 logger.error(f"Error during final check for column '{price_col}': {e}")
                 logger.info(f"Final check: Column '{price_col}' dtype: {output_df[price_col].dtype}")


    return output_df

def _apply_basic_excel_formatting(worksheet, column_list):
    """
    Applies basic Excel formatting to the worksheet:
    - Sets column widths
    - Applies header styles
    - Applies basic cell formatting
    """
    try:
        # 1. Set column widths based on content type
        for col_idx, col_name in enumerate(column_list, 1):
            # Default width based on column type
            if '이미지' in col_name or 'image' in col_name.lower():
                width = 30  # Image columns
            elif 'URL' in col_name or '링크' in col_name or 'link' in col_name.lower():
                width = 40  # URL columns
            elif '상품명' in col_name or '제품명' in col_name:
                width = 35  # Product name columns
            elif '코드' in col_name or 'code' in col_name.lower():
                width = 15  # Code columns
            else:
                width = 20  # Default width
            
            # Set column width
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
        
        # 2. Apply header style
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(bold=True, size=11)
        header_style.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header style to first row
        for col_idx in range(1, len(column_list) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.style = header_style
        
        # Make header row taller
        worksheet.row_dimensions[1].height = 30
        
        # 3. Apply basic data cell formatting
        data_style = NamedStyle(name='data_style')
        data_style.alignment = Alignment(vertical='center', wrap_text=True)
        data_style.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get the number of rows in the worksheet (excluding header)
        max_row = worksheet.max_row
        
        # Apply data style to all data cells
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, len(column_list) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.style = data_style
                
                # Specific formatting for certain column types
                col_name = column_list[col_idx - 1]
                
                # Price columns - right align and format as number
                if '단가' in col_name or '가격' in col_name or 'price' in col_name.lower():
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                # Code/ID columns - center align
                elif '코드' in col_name or 'ID' in col_name or 'id' in col_name.lower():
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                # URL/Link columns - left align
                elif 'URL' in col_name or '링크' in col_name or 'link' in col_name.lower():
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                # Regular text columns - left align
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 4. Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # FIXED: Remove the auto-filter functionality
        # Explicitly remove any existing filter
        if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
            worksheet.auto_filter.ref = None
        
        logger.debug(f"Applied basic Excel formatting to worksheet (header + {max_row-1} data rows)")
        
    except Exception as e:
        logger.warning(f"Error applying basic Excel formatting: {e}")
        logger.debug(traceback.format_exc())

# Add a new function specifically for upload file formatting
def _apply_upload_file_formatting(worksheet, column_list):
    """
    Applies specific formatting for the upload Excel file:
    - Headers with gray background and 2 lines display
    - Content with wrap text
    - Specific cell dimensions and borders
    """
    try:
        # 1. Set standard column width (7) for all columns
        for col_idx in range(1, len(column_list) + 1):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 7
        
        # Adjust specific columns that need different widths
        special_width_columns = {
            '상품명': 35,  # 상품명 needs more width
            '상품코드': 12, # Product code
            '카테고리(중분류)': 15,  # Category
            '해오름(이미지링크)': 40, # Image URLs
            '고려기프트(이미지링크)': 40,
            '네이버쇼핑(이미지링크)': 40,
            '본사링크': 30, # Product links
            '고려 링크': 30,
            '네이버 링크': 30
        }
        
        for idx, col_name in enumerate(column_list, 1):
            column_letter = get_column_letter(idx)
            # Check if this column needs special width
            for special_name, width in special_width_columns.items():
                if special_name in col_name:
                    worksheet.column_dimensions[column_letter].width = width
                    break
        
        # 2. Set row heights - header row = 34.5, data rows = 16.9
        worksheet.row_dimensions[1].height = 34.5  # Header row height
        
        # Set data row heights
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 16.9
        
        # 3. Apply header formatting - gray background and wrap text
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx in range(1, len(column_list) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            # Apply gray background
            cell.fill = gray_fill
            # Enable text wrapping for 2-line display
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Add borders
            cell.border = thin_border
            # Bold font
            cell.font = Font(bold=True, size=10)
        
        # 4. Apply data row formatting - wrap text and borders
        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in range(1, len(column_list) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # Enable text wrapping (fit to cell)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                # Add borders
                cell.border = thin_border
                
                # Adjust alignment based on column content
                col_name = column_list[col_idx - 1]
                # Right-align numeric columns
                if any(term in col_name for term in ['단가', '가격차이', '기본수량']):
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                # Center-align code/ID columns
                elif any(term in col_name for term in ['코드', 'Code', '구분']):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Left-align everything else
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 5. Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Remove any existing auto-filter
        if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
            worksheet.auto_filter.ref = None
        
        logger.info(f"Applied upload file specific formatting to worksheet with {worksheet.max_row} rows.")
        
    except Exception as e:
        logger.warning(f"Error applying upload file formatting: {e}")
        logger.debug(traceback.format_exc())

def extract_naver_image_urls(df: pd.DataFrame, column_name: str) -> pd.DataFrame:
    """
    Extracts Naver image URLs from a DataFrame column and filters out unreliable 'front' URLs.
    
    Args:
        df: DataFrame containing image data
        column_name: Name of the column containing image data
        
    Returns:
        DataFrame with processed image URLs
    """
    if column_name not in df.columns:
        logger.warning(f"Column '{column_name}' not found in DataFrame")
        return df
    
    result_df = df.copy()
    front_url_count = 0
    
    for idx in df.index:
        value = df.at[idx, column_name]
        
        # Skip non-dictionary values
        if not isinstance(value, dict):
            continue
        
        # Check if this is a Naver image dict with a URL
        if 'source' in value and value['source'] == 'naver' and 'url' in value:
            url = value['url']
            
            # Check if it's an unreliable front URL
            if is_unreliable_naver_url(url):
                front_url_count += 1
                
                # Either remove the URL or the entire image data
                if 'local_path' in value and os.path.exists(value['local_path']):
                    # Keep the local image but remove the URL
                    value['url'] = ''
                    logger.info(f"Row {idx}: Removed unreliable 'front' URL but kept local image")
                    result_df.at[idx, column_name] = value
                else:
                    # Remove the entire image data
                    result_df.at[idx, column_name] = '-'
                    logger.info(f"Row {idx}: Removed unreliable 'front' URL image data entirely")
    
    if front_url_count > 0:
        logger.warning(f"Removed or modified {front_url_count} unreliable 'front' URLs in '{column_name}' column")
    
    return result_df

def use_naver_product_links_for_upload(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replace Naver image URLs with the actual product website links for the upload file.
    
    Args:
        df: DataFrame with image information
        
    Returns:
        DataFrame with Naver image URLs replaced with product website links
    """
    # Create a copy to avoid modifying the original DataFrame
    result_df = df.copy()
    
    # We need to find both the Naver image column and the Naver product link column
    naver_img_col = '네이버 이미지'
    naver_link_col = '네이버 쇼핑 링크'
    
    # Ensure both columns exist
    if naver_img_col not in df.columns or naver_link_col not in df.columns:
        logger.warning(f"Either the Naver image column '{naver_img_col}' or link column '{naver_link_col}' is missing. Cannot replace URLs.")
        return df
    
    # Track processed items
    replaced_count = 0
    processed_count = 0
    
    # Process each row
    for idx in df.index:
        try:
            # Get the image data
            img_value = df.at[idx, naver_img_col]
            # Get the product link value
            product_link = df.at[idx, naver_link_col]
            
            processed_count += 1
            
            # Skip if no product link exists
            if pd.isna(product_link) or product_link in ['', '-', 'None', None]:
                continue
            
            # Handle dictionary format for image value
            if isinstance(img_value, dict):
                # Check if this is a Naver image dictionary
                if 'source' in img_value and img_value['source'] == 'naver':
                    # Create a copy of the image data
                    new_img_value = img_value.copy()
                    # Replace the URL with the product link
                    if product_link and isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                        new_img_value['url'] = product_link
                        new_img_value['product_url'] = product_link  # Also store as product_url for clarity
                        result_df.at[idx, naver_img_col] = new_img_value
                        replaced_count += 1
                        logger.debug(f"Row {idx}: Replaced Naver image URL with product link: {product_link[:50]}...")
            # If the image value is a string URL
            elif isinstance(img_value, str) and img_value.startswith(('http://', 'https://')):
                # Only replace if we have a valid product link
                if product_link and isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                    result_df.at[idx, naver_img_col] = {
                        'url': product_link,
                        'source': 'naver',
                        'product_url': product_link,
                        'original_image_url': img_value  # Keep the original image URL for reference
                    }
                    replaced_count += 1
                    logger.debug(f"Row {idx}: Replaced Naver image string URL with product link: {product_link[:50]}...")
        except Exception as e:
            logger.error(f"Error processing row {idx} in use_naver_product_links_for_upload: {e}")
            
    logger.info(f"Replaced {replaced_count}/{processed_count} Naver image URLs with product website links")
    return result_df


# Function to handle extraction and replacement of image URLs for upload file
def prepare_naver_image_urls_for_upload(df_with_image_urls: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare Naver image URLs for the upload file by prioritizing actual image URLs over placeholders.
    
    Args:
        df_with_image_urls: DataFrame with extracted image URLs
        
    Returns:
        DataFrame with processed Naver image URLs
    """
    if df_with_image_urls.empty:
        return df_with_image_urls
        
    # Naver image column in upload format
    naver_img_col = '네이버쇼핑(이미지링크)'
    
    # Check if necessary columns exist
    if naver_img_col not in df_with_image_urls.columns:
        logger.warning(f"Naver image column '{naver_img_col}' not found in DataFrame. Skipping preparation.")
        return df_with_image_urls
    
    # Track processed items
    replaced_count = 0
    placeholder_fixed = 0
    processed_count = 0
    
    # Create a copy of the DataFrame
    result_df = df_with_image_urls.copy()
    
    # Process each row
    for idx in df_with_image_urls.index:
        try:
            # Get the image URL value
            img_url = df_with_image_urls.at[idx, naver_img_col]
            processed_count += 1
            
            # Handle any placeholder or empty URLs - clear instead of replacing with fallbacks
            if not img_url or (isinstance(img_url, str) and (img_url.startswith('http://placeholder.url/') or not img_url.strip())):
                # First try to get data from the original Naver image column
                naver_img_key = '네이버 이미지'
                if naver_img_key in df_with_image_urls.columns:
                    original_data = df_with_image_urls.at[idx, naver_img_key]
                    
                    # Try only original_crawled_url that points to an actual image
                    url_found = False
                    
                    if isinstance(original_data, dict):
                        # Only use original_crawled_url if it's a phinf.pstatic.net URL (actual image)
                        original_crawled_url = original_data.get('original_crawled_url')
                        if original_crawled_url and isinstance(original_crawled_url, str) and original_crawled_url.startswith(('http://', 'https://')) and 'phinf.pstatic.net' in original_crawled_url:
                            result_df.at[idx, naver_img_col] = original_crawled_url
                            placeholder_fixed += 1
                            logger.info(f"Row {idx}: Recovered actual Naver image URL: {original_crawled_url[:50]}...")
                            url_found = True
                    
                    # If no valid URL found, clear the value
                    if not url_found:
                        result_df.at[idx, naver_img_col] = ""
                        logger.warning(f"Row {idx}: No valid Naver image URL found. Clearing value.")
                else:
                    # If no original data column, just clear the value
                    result_df.at[idx, naver_img_col] = ""
                    logger.warning(f"Row {idx}: No Naver image data column. Clearing value.")
            
            # For non-placeholder URLs that are already valid, keep them only if they're actual image URLs
            elif isinstance(img_url, str) and img_url.startswith(('http://', 'https://')):
                # Only keep if it's a phinf.pstatic.net URL (actual image)
                if 'phinf.pstatic.net' in img_url and not img_url.startswith('http://placeholder.url/'):
                    logger.debug(f"Row {idx}: Already has valid Naver image URL: {img_url[:50]}...")
                else:
                    # It's a shopping link or other non-image URL, clear it
                    result_df.at[idx, naver_img_col] = ""
                    logger.warning(f"Row {idx}: URL is not a Naver image URL. Clearing value: {img_url[:50]}...")
            else:
                # Any other non-string or invalid value
                result_df.at[idx, naver_img_col] = ""
                logger.warning(f"Row {idx}: Invalid Naver image value. Clearing.")
        
        except Exception as e:
            logger.error(f"Error processing row {idx} in prepare_naver_image_urls_for_upload: {e}")
            # On error, clear the value rather than keeping potentially bad data
            result_df.at[idx, naver_img_col] = ""
    
    logger.info(f"Processed {processed_count} Naver image URLs for upload file. Fixed {placeholder_fixed} placeholder URLs.")
    return result_df

def is_unreliable_naver_url(url: str) -> bool:
    """
    Checks if a URL is an unreliable Naver 'front' URL.
    
    Args:
        url: The URL to check
        
    Returns:
        True if it's an unreliable front URL, False otherwise
    """
    if not url or not isinstance(url, str):
        return False
    
    # Check for the problematic pattern
    if "pstatic.net/front/" in url:
        logger.warning(f"Detected unreliable 'front' URL: {url}")
        return True
    
    return False

def clean_naver_data_if_link_missing(worksheet, df):
    """
    네이버 이미지 링크 누락 시 관련 데이터 제거
    
    이 함수는 네이버 이미지 URL이 없는 경우에만 관련 데이터를 제거합니다.
    로컬 이미지 파일이 있는 경우는 데이터를 유지합니다.
    """
    try:
        # Check if we should be more lenient with Naver data validation
        lenient_validation = getattr(worksheet, "_lenient_naver_validation", True)
        
        # If lenient validation is enabled, skip cleaning
        if lenient_validation:
            logger.info("Using lenient Naver data validation in clean_naver_data_if_link_missing - keeping all data")
            return
        
        if '네이버쇼핑(이미지링크)' not in df.columns:
            logger.warning("Column '네이버쇼핑(이미지링크)' not found in DataFrame - skipping clean_naver_data_if_link_missing")
            return
            
        naver_col_idx = df.columns.get_loc('네이버쇼핑(이미지링크)') + 1
        data_cols_to_clear = [
            '기본수량(3)', '판매단가(V포함)(3)', 
            '가격차이(3)', '가격차이(3)(%)'
        ]
        
        for row_idx, row in df.iterrows():
            naver_info = row['네이버쇼핑(이미지링크)']
            
            # Check for valid URL or local image file
            has_valid_link = False
            
            if isinstance(naver_info, dict):
                # Check URL
                if 'url' in naver_info and naver_info['url'] and isinstance(naver_info['url'], str):
                    if naver_info['url'].startswith(('http://', 'https://')):
                        has_valid_link = True
                
                # Check local path
                if not has_valid_link and 'local_path' in naver_info and naver_info['local_path']:
                    if os.path.exists(str(naver_info['local_path'])):
                        has_valid_link = True
                        logger.info(f"Row {row_idx+2}: No valid Naver URL but found valid local image path - keeping data")
            
            if not has_valid_link:
                # 이미지 제거
                cell = worksheet.cell(row=row_idx+2, column=naver_col_idx)
                cell.value = '-'
                
                # 관련 데이터 클리어
                for col in data_cols_to_clear:
                    if col in df.columns:
                        col_idx = df.columns.get_loc(col) + 1
                        worksheet.cell(row=row_idx+2, column=col_idx).value = '-'
    except Exception as e:
        logger.error(f"Error in clean_naver_data_if_link_missing: {e}")
        # Continue execution even if this function fails

def prepare_kogift_image_urls_for_upload(df_with_image_urls: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare Kogift image URLs for the upload file by recovering original URLs from placeholders.
    
    Args:
        df_with_image_urls: DataFrame with extracted image URLs
        
    Returns:
        DataFrame with processed Kogift image URLs
    """
    if df_with_image_urls.empty:
        return df_with_image_urls
        
    # Kogift image column in upload format
    kogift_img_col = '고려기프트(이미지링크)'
    # Kogift link column in upload format 
    kogift_link_col = '고려 링크'
    
    # Check if necessary columns exist
    if kogift_img_col not in df_with_image_urls.columns:
        logger.warning(f"Kogift image column '{kogift_img_col}' not found in DataFrame. Skipping preparation.")
        return df_with_image_urls
    
    # Track processed items
    placeholder_fixed = 0
    processed_count = 0
    
    # Create a copy of the DataFrame
    result_df = df_with_image_urls.copy()
    
    # Process each row
    for idx in df_with_image_urls.index:
        try:
            # Get the image URL value
            img_url = df_with_image_urls.at[idx, kogift_img_col]
            # Get the product link value if available
            product_link = df_with_image_urls.at[idx, kogift_link_col] if kogift_link_col in df_with_image_urls.columns else None
            
            processed_count += 1
            
            # Handle any placeholder or empty URLs
            if not img_url or (isinstance(img_url, str) and (img_url.startswith('http://placeholder.url/') or not img_url.strip())):
                # Log that we're processing this row
                logger.debug(f"Row {idx}: Processing Kogift placeholder URL")
                
                # Initialize URL found flag
                url_found = False
                
                # Check if we have useful information in the original Kogift image column
                kogift_img_key = '고려기프트 이미지'
                if kogift_img_key in df_with_image_urls.columns:
                    original_data = df_with_image_urls.at[idx, kogift_img_key]
                    if isinstance(original_data, dict):
                        # 1. First try original_crawled_url (highest priority)
                        original_crawled_url = original_data.get('original_crawled_url')
                        if original_crawled_url and isinstance(original_crawled_url, str) and original_crawled_url.startswith(('http://', 'https://')) and not original_crawled_url.startswith('http://placeholder.url/'):
                            result_df.at[idx, kogift_img_col] = original_crawled_url
                            placeholder_fixed += 1
                            logger.info(f"Row {idx}: Recovered original crawled URL for Kogift placeholder: {original_crawled_url[:50]}...")
                            url_found = True
                        
                        # 2. Try original_url
                        elif not url_found:
                            original_url = original_data.get('original_url')
                            if original_url and isinstance(original_url, str) and original_url.startswith(('http://', 'https://')) and not original_url.startswith('http://placeholder.url/'):
                                result_df.at[idx, kogift_img_col] = original_url
                                placeholder_fixed += 1
                                logger.info(f"Row {idx}: Recovered original URL for Kogift placeholder: {original_url[:50]}...")
                                url_found = True
                            
                        # 3. Try direct url (might be a valid URL even if flagged as placeholder)
                        elif not url_found:
                            url = original_data.get('url')
                            if url and isinstance(url, str) and url.startswith(('http://', 'https://')) and not url.startswith('http://placeholder.url/'):
                                result_df.at[idx, kogift_img_col] = url
                                placeholder_fixed += 1
                                logger.info(f"Row {idx}: Using direct URL from Kogift image data: {url[:50]}...")
                                url_found = True
                        
                        # 4. Try to construct URL from product_id in the image data
                        elif not url_found and 'product_id' in original_data:
                            product_id = original_data.get('product_id')
                            if product_id:
                                try:
                                    constructed_url = f"https://koreagift.com/ez/upload/mall/shop_{product_id}.jpg"
                                    result_df.at[idx, kogift_img_col] = constructed_url
                                    placeholder_fixed += 1
                                    logger.info(f"Row {idx}: Constructed Kogift image URL from product_id in image data: {constructed_url}")
                                    url_found = True
                                except Exception as e:
                                    logger.warning(f"Row {idx}: Failed to construct URL from product_id in image data: {e}")
                        
                # 5. Try to use product link and extract product_id
                if not url_found and product_link and isinstance(product_link, str) and product_link.startswith(('http://', 'https://')):
                    try:
                        # Extract product code from URL (common pattern in Kogift links)
                        import re
                        # Try different patterns for extracting product ID
                        product_id = None
                        
                        # Look for 'product_id='
                        product_id_match = re.search(r'product_id=(\d+)', product_link)
                        if product_id_match:
                            product_id = product_id_match.group(1)
                        
                        # Also try 'no=' which appears in some Kogift URLs
                        elif 'no=' in product_link:
                            no_match = re.search(r'no=(\d+)', product_link)
                            if no_match:
                                product_id = no_match.group(1)
                        
                        # If found product_id, construct image URL
                        if product_id:
                            constructed_url = f"https://koreagift.com/ez/upload/mall/shop_{product_id}.jpg"
                            result_df.at[idx, kogift_img_col] = constructed_url
                            placeholder_fixed += 1
                            logger.info(f"Row {idx}: Constructed Kogift image URL from product ID in link: {constructed_url}")
                            url_found = True
                        # If we couldn't extract product_id, use the product link as fallback
                        else:
                            result_df.at[idx, kogift_img_col] = product_link
                            placeholder_fixed += 1
                            logger.info(f"Row {idx}: Using product link as fallback for Kogift image: {product_link[:50]}...")
                            url_found = True
                    except Exception as e:
                        logger.warning(f"Row {idx}: Failed to process Kogift product link: {e}")
                        # Still use the product link as fallback even if processing failed
                        result_df.at[idx, kogift_img_col] = product_link
                        placeholder_fixed += 1
                        logger.info(f"Row {idx}: Using product link as fallback despite error: {product_link[:50]}...")
                        url_found = True
                
                # 6. Log if we couldn't find any valid URL
                if not url_found:
                    logger.warning(f"Row {idx}: Failed to find any valid URL for Kogift image, placeholder remains")
            
            # For non-placeholder URLs that are already valid, keep them
            elif isinstance(img_url, str) and img_url.startswith(('http://', 'https://')) and not img_url.startswith('http://placeholder.url/'):
                logger.debug(f"Row {idx}: Already has valid Kogift image URL: {img_url[:50]}...")
            
        except Exception as e:
            logger.error(f"Error processing row {idx} in prepare_kogift_image_urls_for_upload: {e}")
    
    logger.info(f"Processed {processed_count} Kogift image URLs for upload file. Fixed {placeholder_fixed} placeholder URLs.")
    return result_df
