#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Fix Kogift Images and Pricing in Excel Files
-------------------------------------------
This script fixes issues with Kogift images and pricing in Excel files by:
1. Reading generated Excel files
2. Updating pricing based on correct quantity tiers from Kogift data
3. Fixing image paths and URLs as needed
4. Preserving hyperlinks and other formatting

Usage:
    python fix_kogift_images.py --input [input_excel_file] --output [output_excel_file]
"""

import os
import sys
import logging
import argparse
import json
import re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import ast
import shutil
import configparser
from urllib.parse import urlparse
from typing import Optional

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('fix_kogift_images.log')
    ]
)
logger = logging.getLogger('fix_kogift_images')

def get_config(config_path='../config.ini'):
    """Load configuration from config.ini file."""
    # Try to find config.ini relative to this script, then one level up.
    script_dir = Path(__file__).parent
    paths_to_try = [
        script_dir / config_path,
        script_dir.parent / 'config.ini'
    ]
    
    conf = configparser.ConfigParser()
    loaded_path = None
    for p_try in paths_to_try:
        if p_try.exists():
            conf.read(p_try, encoding='utf-8')
            loaded_path = p_try
            break
            
    if not loaded_path:
        # Fallback for when script is run from a different context (e.g. main_rpa.py)
        # In this case, assume config.ini is in the root of the project (one level above PythonScript)
        project_root_config = Path(os.getcwd()).parent / 'config.ini'
        if project_root_config.exists():
             conf.read(project_root_config, encoding='utf-8')
             loaded_path = project_root_config
        else: # Final fallback
            default_config_path = Path('config.ini') # current working directory
            if default_config_path.exists():
                conf.read(default_config_path, encoding='utf-8')
                loaded_path = default_config_path
            else:
                 logger.error(f"Config file not found at {paths_to_try} or {project_root_config} or {default_config_path}")
                 raise FileNotFoundError(f"Config file not found.")
    logger.info(f"Loaded config from: {loaded_path}")
    return conf

def find_local_image_by_url(url: str, base_image_dir: Path) -> Optional[str]:
    """Attempts to find a local image file based on its URL filename."""
    if not url or not isinstance(url, str) or not base_image_dir.exists():
        return None
    
    try:
        url_filename = Path(urlparse(url).path).name
        if not url_filename:
            return None

        # Search for the filename (and common variations) in the base_image_dir
        # This is a simple search, can be expanded with glob or recursive search if needed
        possible_files = [
            base_image_dir / url_filename,
            base_image_dir / url_filename.lower(),
        ]
        # Check for common image extensions if original URL filename doesn't have one or is generic
        if '.' not in url_filename:
            for ext in ['.jpg', '.png', '.jpeg', '.gif']:
                possible_files.append(base_image_dir / (url_filename + ext))
                possible_files.append(base_image_dir / (url_filename.lower() + ext))

        for p_file in possible_files:
            if p_file.exists() and p_file.is_file():
                logger.info(f"Found matching local image for URL '{url}' at '{p_file}'")
                return str(p_file)
        
        # Fallback: search with glob for partial matches if direct name fails
        # (e.g. if downloaded file has a prefix or slightly different name)
        # Use the part of the filename without extension for broader matching
        url_filename_stem = Path(url_filename).stem
        if url_filename_stem:
            for ext_pattern in ['*.jpg', '*.png', '*.jpeg', '*.gif']:
                for found_file in base_image_dir.glob(f"*{url_filename_stem}*{ext_pattern}"):
                    if found_file.is_file():
                        logger.info(f"Found glob matching local image for URL '{url}' at '{found_file}' (stem: {url_filename_stem})")
                        return str(found_file)

    except Exception as e:
        logger.error(f"Error while trying to find local image for URL '{url}' in '{base_image_dir}': {e}")
    
    return None

def find_appropriate_price(quantity_prices, target_quantity):
    """
    Find the appropriate price tier for the given quantity.
    
    Args:
        quantity_prices: Dictionary of quantity-price information
        target_quantity: Target quantity to match
        
    Returns:
        tuple: (price, price_with_vat, exact_match, actual_quantity, note)
    """
    if not quantity_prices:
        return None, None, False, None, "No quantity prices available"
    
    # Ensure all keys are integers (sometimes they're stored as strings)
    qty_prices = {}
    for k, v in quantity_prices.items():
        try:
            qty_prices[int(k)] = v
        except (ValueError, TypeError):
            continue # Skip non-integer keys
    
    # Get available quantities, sorted in ascending order
    quantities = sorted(qty_prices.keys())
    if not quantities:
        return None, None, False, None, "No valid quantity tiers found"
    
    min_quantity = min(quantities) # 최소 티어 수량
    max_quantity = max(quantities) # 최대 티어 수량
    # logger.info(f"테이블 최소 수량: {min_quantity}개") # Original log line, can be kept or removed.

    # New logic begins
    # 1. 정확히 일치하는 수량이 있는 경우
    if target_quantity in quantities:
        logger.info(f"수량 {target_quantity}개 정확히 일치: {qty_prices[target_quantity].get('price', 0)}원")
        price_info = qty_prices[target_quantity]
        return (
            price_info.get('price', 0),
            price_info.get('price_with_vat', 0),
            True,
            target_quantity,
            "정확히 일치하는 수량"
        )

    # 2. 주문 수량이 모든 가격 티어보다 큰 경우: 가장 큰 티어의 가격 적용
    if target_quantity > max_quantity:
        logger.info(f"주문 수량({target_quantity})이 최대 티어 수량({max_quantity})보다 큽니다. 최대 티어의 가격을 적용합니다.")
        price_info = qty_prices[max_quantity]
        return (
            price_info.get('price', 0),
            price_info.get('price_with_vat', 0),
            False,
            max_quantity,
            f"최대 구간({max_quantity}개) 가격 적용 (요청 수량 초과)"
        )
        
    # 3. 주문 수량이 특정 티어 사이에 있거나, 최소 티어보다 작은 경우:
    #    주문 수량보다 크거나 같은 티어 중 가장 작은 티어의 가격을 적용
    #    (예: 티어 [200, 300, 500] / 주문 100 -> 200개 가격 / 주문 250 -> 300개 가격)
    
    # target_quantity 보다 크거나 같은 티어들을 찾음
    higher_or_equal_tiers = [q for q in quantities if q >= target_quantity]
    
    if higher_or_equal_tiers:
        chosen_tier = min(higher_or_equal_tiers) # 그 중 가장 작은 티어 선택
        note = f"구간 가격({chosen_tier}개) 적용"
        if target_quantity < chosen_tier: # 요청 수량이 선택된 티어보다 작을 경우 (예: 100개 요청 -> 200개 티어 선택)
            note = f"최소 적용 가능 구간({chosen_tier}개) 가격 적용"
            
        logger.info(f"주문 수량({target_quantity})에 대해 {note}: {qty_prices[chosen_tier].get('price', 0)}원")
        price_info = qty_prices[chosen_tier]
        return (
            price_info.get('price', 0),
            price_info.get('price_with_vat', 0),
            False, # target_quantity와 chosen_tier가 다를 수 있음
            chosen_tier,
            note
        )
        
    # 위의 로직으로 대부분 커버되지만, 예외적 상황(예: quantities가 비어있지 않으나 higher_or_equal_tiers가 빈 경우 등)을 위한 폴백.
    # 현재 로직 상으로는 이 부분에 도달하기 어려움.
    # 만약 target_quantity가 min_quantity보다 작다면, higher_or_equal_tiers에는 min_quantity가 포함되어 위에서 처리됨.
    logger.warning(f"주문 수량({target_quantity})에 대한 가격 티어를 결정하는 데 예외적인 상황 발생. 사용 가능한 티어: {quantities}. 최소 티어 가격을 사용합니다.")
    # Fallback to the smallest available tier if other logic fails. (min_quantity is already defined)
    price_info = qty_prices[min_quantity]
    return (
        price_info.get('price', 0),
        price_info.get('price_with_vat', 0),
        False,
        min_quantity, # Fallback to min_quantity tier
        f"폴백: 최소 구간({min_quantity}개) 가격 적용"
    )

def parse_complex_value(value):
    """Parse string representations of dictionaries or complex objects."""
    if isinstance(value, dict):
        return value
    
    if isinstance(value, str):
        value = value.strip()
        if value.startswith('{') and value.endswith('}'):
            try:
                return ast.literal_eval(value)
            except (SyntaxError, ValueError):
                pass
    return value

def extract_quantity_prices_from_row(row, temp_kogift_col='_temp_kogift_quantity_prices'):
    """
    Extract quantity-based pricing information from a row.
    Returns dict {quantity: {'price': price, 'price_with_vat': price_with_vat}} or None.
    """
    # First, check the primary column for actual crawled price tiers
    actual_tiers_col = '고려기프트_실제가격티어'
    if actual_tiers_col in row and pd.notna(row[actual_tiers_col]) and row[actual_tiers_col] != '-':
        try:
            data_str = row[actual_tiers_col]
            if isinstance(data_str, str):
                # Parse the string representation of dict
                parsed_data = ast.literal_eval(data_str)
                if isinstance(parsed_data, dict):
                    # Convert string keys to integers
                    result = {}
                    for k, v in parsed_data.items():
                        try:
                            result[int(k)] = v
                        except ValueError:
                            logger.warning(f"Invalid quantity key '{k}' in {actual_tiers_col}, skipping")
                    
                    if result:
                        logger.info(f"Successfully parsed {len(result)} price tiers from '{actual_tiers_col}'")
                        return result
            elif isinstance(data_str, dict):
                # Already a dict
                logger.info(f"Using dict data directly from '{actual_tiers_col}'")
                return {int(k): v for k, v in data_str.items() if str(k).isdigit()}
        except Exception as e:
            logger.warning(f"Error parsing '{actual_tiers_col}': {e}")
    
    # Then check the temp column
    if temp_kogift_col in row and pd.notna(row[temp_kogift_col]) and row[temp_kogift_col] != '-':
        data = parse_complex_value(row[temp_kogift_col])
        if isinstance(data, dict) and data:
            logger.info(f"Found quantity_prices data in temp column: {list(data.keys())} tiers")
            # Convert string keys to integers
            return {int(k): v for k, v in data.items() if str(k).isdigit()}

    # Check standard column names
    standard_cols = [
        'kogift_quantity_prices', 'kogift_data', 'kogift_price_data', 
        'kogift_product_data', 'quantity_prices', 
        '고려기프트_데이터', '고려기프트_가격정보', '고려기프트_수량가격',
        '수량별가격', '가격정보', 'price_data', 'pricing_info',
        # 추가 칼럼
        '고려기프트_가격티어', '고려기프트_가격', '고려기프트_수량별가격정보',
        '고려_가격티어', '고려_수량별가격', '판매단가2_티어',
        'kogift_price_tiers', 'kogift_prices', 'quantity_price_data',
        '고려가격', '고려수량가격'
    ]
    
    for col in standard_cols:
        if col in row and pd.notna(row[col]) and row[col] != '-':
            raw_value = row[col]
            logger.debug(f"Checking column '{col}' with value type: {type(raw_value)}")
            
            # Try to parse the value
            data = parse_complex_value(raw_value)
            
            if isinstance(data, dict) and data:
                # Check if it's a direct quantity-price mapping
                if all(str(k).isdigit() for k in data.keys()):
                    logger.info(f"Found direct quantity-price mapping in column '{col}'")
                    return {int(k): v for k, v in data.items()}
                
                # Check for nested structure
                if 'quantity_prices' in data:
                    nested_data = data['quantity_prices']
                    if isinstance(nested_data, dict):
                        logger.info(f"Found nested quantity_prices in column '{col}'")
                        return {int(k): v for k, v in nested_data.items() if str(k).isdigit()}
                
                # Check if values are price dictionaries
                first_value = next(iter(data.values()), None)
                if isinstance(first_value, dict) and ('price' in first_value or 'price_with_vat' in first_value):
                    logger.info(f"Found price dictionary structure in column '{col}'")
                    return {int(k): v for k, v in data.items() if str(k).isdigit()}
            
            # Try to parse as string representation of dict
            elif isinstance(raw_value, str):
                try:
                    # Remove any BOM or special characters
                    cleaned_value = raw_value.strip().strip('\ufeff')
                    if cleaned_value.startswith('{') and cleaned_value.endswith('}'):
                        parsed_data = ast.literal_eval(cleaned_value)
                        if isinstance(parsed_data, dict):
                            logger.info(f"Parsed string representation of dict in column '{col}'")
                            return {int(k): v for k, v in parsed_data.items() if str(k).isdigit()}
                except Exception as e:
                    logger.debug(f"Failed to parse string as dict in column '{col}': {e}")
            
            # 추가: 단일 가격/수량 정보만 있는 경우 처리
            elif isinstance(raw_value, (int, float)) and '기본수량(2)' in row and pd.notna(row['기본수량(2)']):
                try:
                    quantity = int(row['기본수량(2)'])
                    price = float(raw_value)
                    logger.info(f"Found single price {price} for quantity {quantity} in column '{col}'")
                    return {quantity: {'price': price, 'price_with_vat': price}}
                except (ValueError, TypeError) as e:
                    logger.debug(f"Failed to parse single price/quantity: {e}")

    # 추가: 직접 수량-가격 조합 검색
    # 열 이름에서 숫자 패턴 (예: '수량_100', '수량_300') 검색
    quantity_price_pairs = {}
    quantity_pattern = re.compile(r'(수량|개수|갯수|qty|quantity)_(\d+)', re.IGNORECASE)
    price_pattern = re.compile(r'(가격|단가|price)_(\d+)', re.IGNORECASE)
    
    # 수량 패턴 먼저 찾기
    quantity_cols = {}
    for col_name in row.index:
        if isinstance(col_name, str):
            qty_match = quantity_pattern.search(col_name)
            if qty_match:
                try:
                    qty = int(qty_match.group(2))
                    quantity_cols[qty] = col_name
                except ValueError:
                    continue
    
    # 각 수량에 맞는 가격 찾기
    for qty, qty_col in quantity_cols.items():
        # 1. 같은 숫자를 가진 가격 열 찾기 (예: '수량_100'에 대한 '가격_100')
        price_col = None
        for col_name in row.index:
            if isinstance(col_name, str):
                price_match = price_pattern.search(col_name)
                if price_match and price_match.group(2) == str(qty):
                    price_col = col_name
                    break
        
        # 2. 가격 열이 있으면 값 추출
        if price_col and pd.notna(row[price_col]) and row[price_col] != '-':
            try:
                price = float(row[price_col])
                quantity_price_pairs[qty] = {'price': price, 'price_with_vat': price}
            except (ValueError, TypeError):
                logger.debug(f"Failed to parse price from column '{price_col}'")
    
    if quantity_price_pairs:
        logger.info(f"Found {len(quantity_price_pairs)} quantity-price pairs from pattern matching")
        return quantity_price_pairs

    # Check if there's a link but no data - provide more helpful message
    kogift_link_cols = ['고려기프트 상품링크', '고려기프트상품링크', '고려기프트 링크', 'kogift_link', '고려 링크']
    has_kogift_link = False
    for link_col in kogift_link_cols:
        if link_col in row and pd.notna(row[link_col]) and row[link_col] != '-':
            has_kogift_link = True
            break
    
    if has_kogift_link:
        product_name = row.get('상품명', 'Unknown')
        logger.warning(f"Product '{product_name}': Has Kogift link but no quantity-price data found. Data might not have been crawled properly.")
    else:
        logger.debug("No Kogift link found, likely not a Kogift product.")
    
    return None

def fix_excel_kogift_images(input_file, output_file=None, config_obj=None):
    """
    Fix Kogift images and pricing in Excel files.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file (optional)
        config_obj: Optional pre-loaded ConfigParser object
        
    Returns:
        str: Path to output file if successful, None otherwise
    """
    try:
        logger.info(f"Reading Excel file: {input_file}")
        
        if config_obj:
            config = config_obj
        else:
            config = get_config()

        image_main_dir_str = config.get('Paths', 'image_main_dir', fallback='C:\\\\RPA\\\\Image\\\\Main')
        kogift_image_base_dir = Path(image_main_dir_str) / 'Kogift'

        if not kogift_image_base_dir.exists():
            logger.error(f"CRITICAL: Kogift image base directory does not exist: {kogift_image_base_dir}")
            # Decide if to create it or fail
            try:
                kogift_image_base_dir.mkdir(parents=True, exist_ok=True)
                logger.info(f"Created Kogift image base directory: {kogift_image_base_dir}")
            except Exception as e:
                logger.error(f"Failed to create Kogift image base directory {kogift_image_base_dir}: {e}")
                # Proceeding without a valid kogift_image_base_dir will likely cause issues
                # For now, we set it to None as per original logic's fallback, but with a critical error logged.
                # This allows the price fixing part to potentially still run.
                # return None # Optionally, hard fail if image directory is critical
        
        # Set output file path if not specified
        if not output_file:
            input_path = Path(input_file)
            output_file = str(input_path.parent / f"{input_path.stem}_fixed{input_path.suffix}")
        
        # Read the Excel file
        df = pd.read_excel(input_file)
        logger.info(f"Successfully read Excel file with {len(df)} rows")
        
        # Check if this is a result or upload file
        is_result_file = "result" in os.path.basename(input_file).lower()
        is_upload_file = "upload" in os.path.basename(input_file).lower()
        file_type = "result" if is_result_file else "upload" if is_upload_file else "unknown"
        logger.info(f"Detected file type: {file_type}")
        
        # Make a copy of the workbook with openpyxl to preserve formatting and hyperlinks
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
        
        # Use the dynamically determined kogift_image_base_dir
        # kogift_image_dir is now a Path object or None if creation failed
        active_kogift_image_dir = str(kogift_image_base_dir) if kogift_image_base_dir and kogift_image_base_dir.exists() else None

        if not active_kogift_image_dir:
            logger.warning(f"Kogift image directory is not available. Image fixing will be skipped or limited.")
            # kogift_image_dir = None # Ensure it's None if not active_kogift_image_dir
        
        # Map column names (accounting for variations in column names)
        column_mapping = {
            # 기본 수량 및 가격 칼럼
            '기본수량(1)': ['기본수량(1)', '기본수량', '수량', '본사 기본수량', '본사수량'],
            '판매단가(V포함)': ['판매단가(V포함)', '판매단가1(VAT포함)', '판매단가(VAT포함)', '본사 판매단가'],
            
            # 링크 관련 칼럼
            '본사상품링크': ['본사상품링크', '본사링크', '해오름링크', '본사 링크'],
            '고려기프트 상품링크': ['고려기프트 상품링크', '고려기프트상품링크', '고려기프트 링크', '고려 링크', 'kogift_link'],
            '네이버 쇼핑 링크': ['네이버 쇼핑 링크', '네이버 링크', '네이버쇼핑링크', '네이버 링크'],
            
            # 고려기프트 관련 칼럼
            '기본수량(2)': ['기본수량(2)', '고려 기본수량', '고려기프트 기본수량', '고려기프트수량'],
            '판매가(V포함)(2)': ['판매가(V포함)(2)', '판매단가(V포함)(2)', '고려 판매가(V포함)', '고려기프트 판매가', '판매단가2(VAT포함)', '고려기프트 판매단가'],
            '가격차이(2)': ['가격차이(2)', '고려 가격차이', '고려기프트 가격차이'],
            '가격차이(2)(%)': ['가격차이(2)(%)', '고려 가격차이(%)', '고려 가격 차이(%)', '가격차이(2) %'],
            
            # 네이버 관련 칼럼
            '기본수량(3)': ['기본수량(3)', '네이버 기본수량'],
            '판매단가(V포함)(3)': ['판매단가(V포함)(3)', '판매단가3 (VAT포함)', '네이버 판매단가'],
            '가격차이(3)': ['가격차이(3)', '네이버 가격차이'],
            '가격차이(3)(%)': ['가격차이(3)(%)', '네이버가격차이(%)', '네이버 가격차이(%)'],
            '공급사명': ['공급사명', '네이버 공급사명'],
            
            # 이미지 관련 칼럼
            '본사 이미지': ['본사 이미지', '해오름(이미지링크)', '해오름이미지', '본사이미지URL', '해오름 이미지 URL'],
            '고려기프트 이미지': ['고려기프트 이미지', '고려기프트이미지', '고려 이미지', 'kogift_image', '고려기프트(이미지링크)'],
            '네이버 이미지': ['네이버 이미지', '네이버쇼핑(이미지링크)', '네이버이미지']
        }
        
        # Find which variant of each column exists in the DataFrame
        columns_found = {}
        for key, variants in column_mapping.items():
            for variant in variants:
                if variant in df.columns:
                    columns_found[key] = variant
                    logger.debug(f"Found column '{variant}' for key '{key}'")
                    break
            if key not in columns_found:
                logger.warning(f"Could not find any variant for column '{key}'")
        
        # Log found columns
        logger.info(f"Found column mappings: {columns_found}")
        logger.info(f"Total columns in DataFrame: {len(df.columns)}")
        logger.debug(f"All DataFrame columns: {list(df.columns)}")
        
        # For upload files, the structure may be different and may not have all required columns
        required_columns_by_type = {
            'result': ['기본수량(1)', '고려기프트 상품링크', '고려기프트 이미지'],
            'upload': ['본사 기본수량', '고려 링크', '고려기프트(이미지링크)']  # upload 파일의 실제 칼럼명 사용
        }
        
        # Get required columns for this file type
        required_columns = required_columns_by_type.get(file_type, ['기본수량(1)', '고려기프트 상품링크'])
        
        # Check for required columns based on file type
        missing_columns = []
        if file_type == 'result':
            # result 파일의 경우 매핑된 표준 칼럼명으로 체크
            missing_columns = [col for col in required_columns_by_type['result'] if col not in columns_found]
            if missing_columns:
                logger.warning(f"Result 파일에서 필요한 칼럼이 없습니다: {missing_columns}. 가능한 칼럼으로 진행합니다.")
        elif file_type == 'upload':
            # upload 파일의 경우 실제 칼럼명으로 체크 (엑셀에 있는 그대로)
            upload_columns = required_columns_by_type['upload']
            missing_upload_columns = []
            for col in upload_columns:
                found = False
                for variants in column_mapping.values():
                    if col in variants and any(variant in df.columns for variant in variants):
                        found = True
                        break
                if not found:
                    missing_upload_columns.append(col)
            
            if missing_upload_columns:
                logger.warning(f"Upload 파일에서 필요한 칼럼이 없습니다: {missing_upload_columns}. 가능한 칼럼으로 진행합니다.")
            missing_columns = missing_upload_columns
        else:
            logger.warning(f"알 수 없는 파일 타입: {file_type}. 기본 칼럼 요구사항으로 진행합니다.")
            missing_columns = [col for col in required_columns if col not in columns_found]
            if missing_columns:
                logger.warning(f"필요한 칼럼이 없습니다: {missing_columns}. 가능한 칼럼으로 진행합니다.")
        
        # Find column indices for updating (1-indexed for openpyxl)
        column_indices = {}
        for col_idx, cell in enumerate(sheet[1], 1):  # 1-indexed columns
            column_indices[cell.value] = col_idx
        
        # Log found column indices
        logger.info(f"Found column indices: {column_indices}")
        
        # Map the actual column names in the Excel file to our expected column names
        real_column_indices = {}
        for expected_col, column_idx in column_indices.items():
            # Try to map each column in the excel file to our expected columns
            for key, variants in column_mapping.items():
                if expected_col in variants:
                    real_column_indices[key] = column_idx
                    break
                    
        logger.info(f"Mapped column indices: {real_column_indices}")
        
        # Get the actual column names to use based on what's in the DataFrame
        quantity_col = columns_found.get('기본수량(1)')
        base_price_col = columns_found.get('판매단가(V포함)')
        kogift_link_col = columns_found.get('고려기프트 상품링크')
        quantity2_col = columns_found.get('기본수량(2)')
        price2_col = columns_found.get('판매가(V포함)(2)')
        price_diff_col = columns_found.get('가격차이(2)')
        price_diff_pct_col = columns_found.get('가격차이(2)(%)')
        kogift_image_col = columns_found.get('고려기프트 이미지')
        
        # 칼럼을 찾지 못한 경우 로그 남기기
        if not quantity_col:
            logger.warning("기본수량(1) 칼럼을 찾을 수 없습니다.")
        if not kogift_link_col:
            logger.warning("고려기프트 상품링크 칼럼을 찾을 수 없습니다.")
        if not price2_col:
            logger.warning("판매가(V포함)(2) 칼럼을 찾을 수 없습니다. 가격 정보를 업데이트할 수 없습니다.")
        
        # Process each row that has Kogift data
        update_count = 0
        price_diffs_updated = 0
        wrong_image_count = 0
        filtered_by_missing_data = 0
        filtered_by_missing_image = 0
        
        # 특별히 관심 있는 수량 값들 추적
        small_quantity_handling = {}  # 수량이 작은 행 처리 결과 추적
        
        # 행 별로 처리
        for idx, row in df.iterrows():
            # Kogift 링크가 있고 기본수량 칼럼이 있는 행만 처리
            has_kogift_link = False
            if kogift_link_col and kogift_link_col in row:
                has_kogift_link = not pd.isna(row[kogift_link_col]) and row[kogift_link_col] != '-'
            
            # 필수 상품 정보 (수량과 기본 가격)가 있는지 체크
            has_required_data = False
            if quantity_col and quantity_col in row and pd.notna(row[quantity_col]) and row[quantity_col] != '-':
                if base_price_col and base_price_col in row and pd.notna(row[base_price_col]) and row[base_price_col] != '-':
                    has_required_data = True
            
            # 케이스 1: 이미지 링크만 있고 상품 정보가 완전히 없는 경우에만 필터링
            # 그렇지 않으면 가능한 많은 정보를 유지하도록 변경
            if has_kogift_link and not has_required_data:
                # 링크는 있지만 상품 정보는 완전히 없어야 함
                # 수량이 있고 가격만 없는 경우는 계속 진행
                if quantity_col and not (quantity_col in row and pd.notna(row[quantity_col]) and row[quantity_col] != '-'):
                    logger.warning(f"Row {idx+1}: Kogift link exists but completely missing product data (no quantity). Marking for review but preserving data.")
                    filtered_by_missing_data += 1
                    # 데이터를 제거하지 않고 로그만 남김
                    
                    # 원래 코드는 여기서 모든 데이터를 지웠지만, 이제는 보존함:
                    # xl_row = idx + 2
                    # 고려기프트 링크, 이미지, 가격 정보 지우기
            
            # 케이스 2: 상품 정보는 있지만 Kogift 링크가 없는 경우
            # 링크가 없더라도 다른 정보가 있으면 유지
            if has_required_data and not has_kogift_link:
                # 이미지나 가격 정보가 있는지 체크
                has_kogift_image = False
                if kogift_image_col in columns_found and kogift_image_col in row:
                    cell_value = row[kogift_image_col]
                    if isinstance(cell_value, dict) or (isinstance(cell_value, str) and cell_value.strip() and cell_value != '-'):
                        has_kogift_image = True
                
                has_kogift_price = False
                if price2_col in columns_found and pd.notna(row.get(price2_col)) and row.get(price2_col) != '-':
                    has_kogift_price = True
                
                # 이미지나 가격 정보가 있지만 링크가 없으면 경고만 기록하고 데이터는 유지
                if has_kogift_image or has_kogift_price:
                    logger.warning(f"Row {idx+1}: Missing Kogift link but has Kogift image or price data. Keeping data for review.")
                    filtered_by_missing_image += 1
                    # 데이터를 제거하지 않고 로그만 남김
            
            # 링크도 없고 상품 정보도 없으면 처리할 필요 없음
            if not has_kogift_link and not has_required_data:
                continue
            
            # 여기부터는 기존 로직 계속 (링크도 있고 상품 정보도 있는 정상 케이스)
            
            # 이미지 데이터 검증
            if kogift_image_col and kogift_image_col in row:
                image_data = parse_complex_value(row[kogift_image_col])
                
                # Upload 파일 형식 처리 (이미지 링크 문자열)
                if is_upload_file and isinstance(image_data, str) and (image_data.startswith('http') or '/' in image_data):
                    # Upload 파일의 경우 이미지 링크 문자열을 그대로 유지
                    logger.debug(f"Row {idx+1}: Upload file image link preserved: {image_data[:50]}...")
                    # 이미지 처리 필요 없음 - 링크 형식 유지
                    correct_path_found = True  # 처리 완료로 표시
                
                # Result 파일 형식 처리 (이미지 데이터 딕셔너리)
                elif isinstance(image_data, dict):
                    local_path = image_data.get('local_path') or image_data.get('image_path')
                    image_url = image_data.get('url')
                    
                    correct_path_found = False
                    if local_path and isinstance(local_path, str) and Path(local_path).exists():
                        if active_kogift_image_dir and local_path.replace('\\\\', '/').startswith(active_kogift_image_dir.replace('\\\\', '/')):
                            correct_path_found = True
                        elif not active_kogift_image_dir: # No dir to check against, assume path is fine if it exists
                             correct_path_found = True

                    if not correct_path_found:
                        logger.warning(f"Row {idx+1}: Kogift image local_path '{local_path}' is invalid or not in correct directory.")
                        new_local_path = None
                        if image_url and active_kogift_image_dir: # Try to find by URL if dir exists
                            logger.info(f"Attempting to find local Kogift image for URL: {image_url}")
                            new_local_path = find_local_image_by_url(image_url, kogift_image_base_dir)
                        
                        if new_local_path:
                            logger.info(f"Row {idx+1}: Found replacement Kogift image path: {new_local_path}")
                            image_data['local_path'] = new_local_path
                            image_data['original_path'] = new_local_path # Update original_path too
                            
                            # Excel 셀 업데이트 (1-indexed row, 칼럼 인덱스)
                            kogift_image_idx = real_column_indices.get('고려기프트 이미지')
                            if kogift_image_idx:
                                if is_result_file:
                                    # Result 파일의 경우 딕셔너리를 JSON 문자열로 변환
                                    try:
                                        json_str = json.dumps(image_data, ensure_ascii=False)
                                        sheet.cell(row=idx+2, column=kogift_image_idx).value = json_str
                                        logger.debug(f"Updated Excel cell with new image data: {new_local_path}")
                                    except Exception as e:
                                        logger.error(f"Error updating Excel cell: {e}")
                                else:
                                    # Upload 파일의 경우 URL만 사용
                                    sheet.cell(row=idx+2, column=kogift_image_idx).value = image_url or new_local_path
                            
                            correct_path_found = True 
                        else:
                            logger.warning(f"Row {idx+1}: Could not find a valid local Kogift image. Will use URL only.")
                            wrong_image_count += 1
                            
                            # URL만 있는 경우, 파일 형식에 따라 다르게 처리
                            if image_url:
                                kogift_image_idx = real_column_indices.get('고려기프트 이미지')
                                if kogift_image_idx:
                                    if is_upload_file:
                                        # Upload 파일의 경우 URL 문자열만 저장
                                        sheet.cell(row=idx+2, column=kogift_image_idx).value = image_url
                                    else:
                                        # Result 파일의 경우 URL만 포함된 딕셔너리 저장
                                        image_data = {'url': image_url, 'source': 'kogift'}
                                        try:
                                            json_str = json.dumps(image_data, ensure_ascii=False)
                                            sheet.cell(row=idx+2, column=kogift_image_idx).value = json_str
                                        except Exception as e:
                                            logger.error(f"Error updating Excel cell with URL: {e}")

            # 기본수량 확인
            base_quantity = None
            if quantity_col and quantity_col in row and pd.notna(row[quantity_col]):
                try:
                    # 수량을 정수로 변환
                    base_quantity = int(row[quantity_col])
                    logger.info(f"Processing row {idx+1}: Product name: {row.get('상품명', 'Unknown')} with quantity {base_quantity}")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid base quantity in row {idx+1}: {row.get(quantity_col)}")
                    continue
            else:
                logger.debug(f"No base quantity found for row {idx+1}")
                continue
            
            # 특별히 주시하는 경우: 수량이 100과 같이 작은 경우
            is_special_case = base_quantity < 200
            
            # 크롤링된 수량-가격 정보 추출 시도
            quantity_prices = extract_quantity_prices_from_row(row.copy()) # Pass a copy to avoid SettingWithCopyWarning if row is a slice
            
            if not quantity_prices:
                logger.warning(f"Row {idx+1} (Product: '{row.get('상품명', 'Unknown')}'): No valid crawled quantity-price data found. Continuing with available data.")
                # Don't skip - continue with any available data
                # continue
                # Just continue to the next row, but don't exit the loop entirely
                continue
            
            # 로그 출력
            if quantity_prices:
                logger.info(f"Row {idx+1}: 사용 가능한 수량 티어: {sorted(quantity_prices.keys())}")
            else:
                logger.warning(f"Row {idx+1}: 사용 가능한 수량-가격 정보 없음")
                continue
            
            if is_special_case:
                logger.info(f"!! 특별 케이스 발견 !! - Row {idx+1}: 수량이 {base_quantity}개인 경우 처리")
                
                # 이전 가격 정보 저장 (수정 확인용)
                old_price = None
                price2_idx = real_column_indices.get('판매가(V포함)(2)')
                if price2_idx:
                    old_cell = sheet.cell(row=idx+2, column=price2_idx)
                    old_price = old_cell.value
                    logger.info(f"   현재 가격: {old_price}")
            
            # 적절한 가격 티어 찾기
            price, price_with_vat, exact_match, actual_quantity, note = find_appropriate_price(
                quantity_prices, base_quantity
            )
            
            if is_special_case:
                # 특별 케이스인 경우 처리 결과 저장
                small_quantity_handling[idx] = {
                    'row': idx+1,
                    'product_name': row.get('상품명', 'Unknown'),
                    'base_quantity': base_quantity,
                    'available_tiers': sorted(quantity_prices.keys()),
                    'selected_tier': actual_quantity,
                    'price': price,
                    'price_with_vat': price_with_vat,
                    'note': note,
                    'old_price': old_price
                }
                logger.info(f"   해결 결과: 티어 {actual_quantity} 선택, 가격 {price}원 (부가세 포함: {price_with_vat}원)")
                logger.info(f"   처리 내용: {note}")
            
            if price_with_vat:
                # Calculate row in Excel (1-indexed and header row)
                xl_row = idx + 2
                
                # Update quantity column
                quantity2_idx = real_column_indices.get('기본수량(2)')
                if quantity2_idx:
                    sheet.cell(row=xl_row, column=quantity2_idx).value = base_quantity
                
                # Update price column
                price2_idx = real_column_indices.get('판매가(V포함)(2)')
                if price2_idx:
                    current_price = sheet.cell(row=xl_row, column=price2_idx).value
                    sheet.cell(row=xl_row, column=price2_idx).value = price_with_vat
                    logger.info(f"Row {idx+1}: 가격 업데이트: {current_price} -> {price_with_vat}")
                
                # Update price difference if possible
                price_diff_idx = real_column_indices.get('가격차이(2)')
                price_diff_pct_idx = real_column_indices.get('가격차이(2)(%)')
                
                # 본사 가격 찾기 (판매단가(V포함) 또는 판매단가1(VAT포함) 칼럼 이름 사용)
                base_price = None
                base_price_col_name = columns_found.get('판매단가(V포함)')
                
                if base_price_col_name and base_price_col_name in row and pd.notna(row[base_price_col_name]):
                    try:
                        base_price = float(row[base_price_col_name])
                    except (ValueError, TypeError):
                        logger.warning(f"행 {idx+1}: 본사 가격 '{row[base_price_col_name]}'를 숫자로 변환할 수 없습니다.")
                
                # 가격 차이 계산 및 업데이트
                if price_diff_idx and base_price is not None:
                    try:
                        price_diff = price_with_vat - base_price
                        sheet.cell(row=xl_row, column=price_diff_idx).value = price_diff
                        
                        # 음수 가격 차이일 경우 빨간색 배경 적용
                        if price_diff < 0:
                            sheet.cell(row=xl_row, column=price_diff_idx).fill = PatternFill(
                                start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                            )
                        
                        # 퍼센트 가격 차이 계산 및 업데이트
                        if price_diff_pct_idx and base_price != 0:
                            pct_diff = (price_diff / base_price) * 100
                            sheet.cell(row=xl_row, column=price_diff_pct_idx).value = round(pct_diff, 1)
                            
                            # 음수 퍼센트 가격 차이일 경우 빨간색 배경 적용
                            if pct_diff < 0:
                                sheet.cell(row=xl_row, column=price_diff_pct_idx).fill = PatternFill(
                                    start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                                )
                        
                        price_diffs_updated += 1
                        logger.debug(f"행 {idx+1}: 가격차이={price_diff:.1f}, 가격차이(%)={pct_diff:.1f}% 업데이트 완료")
                    except (ValueError, TypeError, NameError) as e:
                        logger.debug(f"행 {idx+1}: 가격차이 계산 중 오류: {e}")
                        
                # 특정 열에 대한 인덱스를 찾지 못한 경우 로그
                if not price_diff_idx and base_price is not None:
                    logger.debug(f"행 {idx+1}: 가격차이 열을 찾을 수 없어 가격차이 업데이트 불가")
                if not price_diff_pct_idx and base_price is not None:
                    logger.debug(f"행 {idx+1}: 가격차이(%) 열을 찾을 수 없어 퍼센트 가격차이 업데이트 불가")
                
                update_count += 1
                logger.debug(f"Updated row {idx+1}: Quantity {base_quantity}, Price {price_with_vat}, Tier {actual_quantity}")

        # 특별 케이스 처리 결과 요약
        if small_quantity_handling:
            logger.info("\n===== 적은 수량 특별 처리 결과 요약 =====")
            for case_idx, case_data in small_quantity_handling.items():
                logger.info(f"행 #{case_data['row']}: {case_data['product_name']}")
                logger.info(f"  수량: {case_data['base_quantity']}, 가능한 티어: {case_data['available_tiers']}")
                logger.info(f"  기존 가격: {case_data['old_price']} -> 새 가격: {case_data['price_with_vat']} (티어 {case_data['selected_tier']})")
                logger.info(f"  비고: {case_data['note']}")
                logger.info("-" * 40)
            logger.info("========================================")
        
        # 필터링 결과 로그
        if filtered_by_missing_data > 0 or filtered_by_missing_image > 0:
            logger.info(f"잠재적 문제 항목: 이미지만 있고 상품 데이터 없는 행 {filtered_by_missing_data}개, 상품 데이터만 있고 이미지 없는 행 {filtered_by_missing_image}개가 있습니다.")
            logger.info("이 항목들은 이전에는 제거되었지만, 현재는 데이터를 유지합니다.")
        
        # Save the modified workbook
        workbook.save(output_file)
        logger.info(f"성공적으로 {update_count}개 행의 가격 정보가 수정되었습니다. (가격차이 계산: {price_diffs_updated}개)")
        logger.info(f"수정된 엑셀 파일 저장 경로: {output_file}")
        
        # 최종 확인 로그 추가
        if update_count == 0:
            logger.warning("!! 주의 !! - 업데이트된 행이 없습니다. 칼럼 매핑을 확인하세요.")
        
        # 문제가 발생한 경우 알림
        missing_column_list = []
        for key_col in ['기본수량(1)', '판매단가(V포함)', '고려기프트 상품링크', '판매가(V포함)(2)']:
            if key_col not in columns_found:
                missing_column_list.append(key_col)
        
        if missing_column_list:
            logger.warning(f"!! 주의 !! - 일부 중요 칼럼을 찾지 못했습니다: {missing_column_list}")
            logger.warning("이로 인해 일부 행이 처리되지 않았을 수 있습니다.")
        
        return output_file
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}", exc_info=True)
        return None

def main():
    """Standalone script to fix Kogift images and pricing in Excel files"""
    parser = argparse.ArgumentParser(description='Fix Kogift images and pricing in Excel files')
    parser.add_argument('--input', '-i', required=True, help='Input Excel file path')
    parser.add_argument('--output', '-o', help='Output Excel file path (optional)')
    
    args = parser.parse_args()
    
    # Load config for standalone execution
    try:
        config = get_config()
    except FileNotFoundError:
        return 1 # Exit if config is not found

    # Validate input file
    input_file = args.input
    if not os.path.exists(input_file):
        logger.error(f"Input file not found: {input_file}")
        return 1
    
    # Set output file if not specified
    output_file = args.output
    
    logger.info(f"Starting Kogift fix process")
    logger.info(f"Input file: {input_file}")
    logger.info(f"Output file: {output_file or 'Will be auto-generated'}")
    
    # Call the fix function
    result = fix_excel_kogift_images(input_file, output_file, config_obj=config)
    
    if result:
        logger.info(f"Successfully fixed Kogift images and pricing. Output saved to: {result}")
        print(f"✅ Successfully fixed Kogift images and pricing in Excel file.")
        print(f"✅ Output saved to: {result}")
        return 0
    else:
        logger.error("Failed to fix Kogift images and pricing")
        print("❌ Failed to fix Kogift images and pricing. Check the log for details.")
        return 1

if __name__ == "__main__":
    sys.exit(main()) 