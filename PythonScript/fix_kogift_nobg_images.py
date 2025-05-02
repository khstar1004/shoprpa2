#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Fix Kogift No-Background Images
------------------------------
This script addresses an issue where Kogift images with _nobg suffix aren't properly 
loaded in the result Excel files, even though the images exist.

Usage:
    python fix_kogift_nobg_images.py [path_to_excel_file]
"""

import os
import sys
import logging
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import re
from pathlib import Path
import argparse

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger()

def scan_kogift_images(base_dir='C:\\RPA\\Image\\Main\\Kogift'):
    """
    Scan Kogift image directory and build mappings between regular and _nobg versions
    
    Args:
        base_dir: Base directory for Kogift images
        
    Returns:
        Dictionary mapping regular image filenames to _nobg paths
    """
    if not os.path.exists(base_dir):
        logger.warning(f"Directory does not exist: {base_dir}")
        return {}
    
    # Dictionary to store mappings between regular images and their _nobg versions
    image_mappings = {}
    
    logger.info(f"Scanning directory: {base_dir}")
    
    # Regular images without _nobg
    regular_images = {}
    # Images with _nobg
    nobg_images = {}
    
    # First pass: collect all images
    for file in os.listdir(base_dir):
        if file.lower().endswith(('.jpg', '.jpeg', '.png')):
            full_path = os.path.join(base_dir, file)
            
            # Skip small files
            if os.path.getsize(full_path) < 1000:
                continue
                
            if '_nobg' in file.lower():
                nobg_images[file] = full_path
            else:
                regular_images[file] = full_path
    
    logger.info(f"Found {len(regular_images)} regular images and {len(nobg_images)} _nobg images")
    
    # Second pass: create mappings
    for reg_file, reg_path in regular_images.items():
        # Get base filename without extension
        base_name = os.path.splitext(reg_file)[0]
        
        # Look for corresponding _nobg versions
        nobg_variant = f"{base_name}_nobg.png"
        if nobg_variant in nobg_images:
            image_mappings[reg_file] = nobg_images[nobg_variant]
            logger.debug(f"Mapped {reg_file} → {nobg_variant}")
    
    # Add hash-based mappings for kogift_ files
    for nobg_file, nobg_path in nobg_images.items():
        if not nobg_file.lower().startswith('kogift_'):
            continue
            
        # Extract the hash part if present
        hash_match = re.search(r'kogift_.*?_([a-f0-9]{8,})_nobg', nobg_file.lower())
        if hash_match:
            hash_val = hash_match.group(1)
            # Add mappings for hash-based patterns
            image_mappings[f"kogift_{hash_val}.jpg"] = nobg_path
            image_mappings[f"kogift_{hash_val}.png"] = nobg_path
    
    logger.info(f"Created {len(image_mappings)} mappings between regular images and _nobg versions")
    return image_mappings

def fix_excel_kogift_images(excel_file, output_file=None):
    """
    Fix Kogift images in an Excel file by replacing them with _nobg versions if available
    
    Args:
        excel_file: Path to the Excel file
        output_file: Optional output file path (if not provided, will overwrite input)
        
    Returns:
        Path to the output file if successful, None otherwise
    """
    try:
        # Validate input file
        if not os.path.exists(excel_file):
            logger.error(f"Input file not found: {excel_file}")
            return None
            
        # Set output file if not specified
        if not output_file:
            output_file = excel_file
            
        logger.info(f"Processing Excel file: {excel_file}")
        
        # Scan for Kogift images
        kogift_images = scan_kogift_images()
        if not kogift_images:
            logger.warning("No Kogift image mappings found. Make sure the Kogift folder exists and contains images.")
            return None
            
        # Load the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Find the Kogift image column
        kogift_col = None
        kogift_col_title = '고려기프트 이미지'
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value == kogift_col_title:
                kogift_col = col
                break
                
        if not kogift_col:
            logger.warning(f"Column '{kogift_col_title}' not found in the Excel file.")
            return None
            
        logger.info(f"Found Kogift image column at position {kogift_col}")
        
        # Process each row
        fixed_count = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=kogift_col)
            cell_value = cell.value
            
            # Skip if cell is empty
            if not cell_value:
                continue
                
            # Remove any existing images
            for img in list(ws._images):
                if img.anchor.to_string() == f"{get_column_letter(kogift_col)}{row}":
                    ws._images.remove(img)
            
            # Check if the value corresponds to a file in our mapping
            img_added = False
            if isinstance(cell_value, str):
                # Try direct filename match
                file_name = os.path.basename(cell_value)
                if file_name in kogift_images:
                    nobg_path = kogift_images[file_name]
                    logger.debug(f"Row {row}: Found mapping for {file_name}")
                    
                    # Add nobg image
                    try:
                        img = Image(nobg_path)
                        # Adjust image size
                        img.width = 150
                        img.height = 150
                        ws.add_image(img, f"{get_column_letter(kogift_col)}{row}")
                        img_added = True
                        fixed_count += 1
                    except Exception as e:
                        logger.error(f"Failed to add image {nobg_path}: {e}")
                elif '_nobg' not in cell_value and 'kogift_' in cell_value.lower():
                    # Try to check if there's a hash-based match
                    hash_match = re.search(r'kogift_.*?_([a-f0-9]{8,})\.', cell_value.lower())
                    if hash_match:
                        hash_val = hash_match.group(1)
                        hash_key = f"kogift_{hash_val}.jpg"
                        if hash_key in kogift_images:
                            nobg_path = kogift_images[hash_key]
                            logger.debug(f"Row {row}: Found hash-based mapping for {hash_key}")
                            
                            # Add nobg image
                            try:
                                img = Image(nobg_path)
                                # Adjust image size
                                img.width = 150
                                img.height = 150
                                ws.add_image(img, f"{get_column_letter(kogift_col)}{row}")
                                img_added = True
                                fixed_count += 1
                            except Exception as e:
                                logger.error(f"Failed to add image {nobg_path}: {e}")
                                
            if not img_added:
                logger.debug(f"Row {row}: No mapping found for {cell_value}")
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Fixed {fixed_count} Kogift images in {excel_file}")
        logger.info(f"Saved output to {output_file}")
        
        return output_file
    except Exception as e:
        logger.error(f"Error fixing Excel file: {e}")
        return None

def main():
    parser = argparse.ArgumentParser(description='Fix Kogift _nobg images in Excel files')
    parser.add_argument('excel_file', nargs='?', help='Path to the Excel file to fix')
    parser.add_argument('--output', '-o', help='Output file path (optional)')
    parser.add_argument('--dir', '-d', help='Directory containing Excel files to process')
    
    args = parser.parse_args()
    
    if args.dir:
        # Process all Excel files in directory
        if not os.path.isdir(args.dir):
            logger.error(f"Directory not found: {args.dir}")
            return 1
            
        success_count = 0
        fail_count = 0
        
        for file in os.listdir(args.dir):
            if file.endswith(('.xlsx', '.xls')):
                file_path = os.path.join(args.dir, file)
                logger.info(f"Processing file: {file_path}")
                
                if fix_excel_kogift_images(file_path):
                    success_count += 1
                else:
                    fail_count += 1
                    
        logger.info(f"Processed {success_count + fail_count} files: {success_count} successful, {fail_count} failed")
        return 0 if fail_count == 0 else 1
    elif args.excel_file:
        # Process single file
        if not os.path.isfile(args.excel_file):
            logger.error(f"File not found: {args.excel_file}")
            return 1
            
        output_file = args.output if args.output else args.excel_file
        if fix_excel_kogift_images(args.excel_file, output_file):
            logger.info("Success!")
            return 0
        else:
            logger.error("Failed to fix Kogift images")
            return 1
    else:
        # No arguments provided, show usage
        parser.print_help()
        return 1

if __name__ == "__main__":
    sys.exit(main()) 