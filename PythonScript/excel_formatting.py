import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
import traceback
from typing import List, Dict, Any, Optional

# Import constants
from excel_constants import (
    PRICE_COLUMNS, QUANTITY_COLUMNS, PERCENTAGE_COLUMNS,
    IMAGE_COLUMNS, ERROR_MESSAGE_VALUES
)
from excel_style_constants import (
    HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT,
    LEFT_ALIGNMENT, CENTER_ALIGNMENT, RIGHT_ALIGNMENT,
    DEFAULT_FONT, DEFAULT_BORDER, NEGATIVE_PRICE_FILL,
    LINK_FONT, UPLOAD_HEADER_FILL, COLUMN_WIDTH_SETTINGS,
    UPLOAD_HEADER_HEIGHT, UPLOAD_DATA_ROW_HEIGHT
)

# Initialize logger
logger = logging.getLogger(__name__)

def _apply_column_widths(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Sets appropriate column widths based on column names/types."""
    logger.debug(f"Applying column widths. DataFrame columns: {df.columns.tolist()}")
    for idx, col_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        width = COLUMN_WIDTH_SETTINGS['default']  # Default width

        col_name_str = str(col_name)  # Ensure col_name is string for checks

        # Determine width based on column name patterns
        if col_name_str in IMAGE_COLUMNS:
            width = COLUMN_WIDTH_SETTINGS['image']
        elif '상품명' in col_name_str:
            width = COLUMN_WIDTH_SETTINGS['name']
        elif col_name_str in PRICE_COLUMNS:
            width = COLUMN_WIDTH_SETTINGS['price']
        elif col_name_str in PERCENTAGE_COLUMNS:
            width = COLUMN_WIDTH_SETTINGS['percent']
        elif col_name_str in QUANTITY_COLUMNS:
            width = COLUMN_WIDTH_SETTINGS['quantity']
        elif 'Code' in col_name_str or '코드' in col_name_str:
            width = COLUMN_WIDTH_SETTINGS['code']
        elif '카테고리' in col_name_str or '분류' in col_name_str:
            width = COLUMN_WIDTH_SETTINGS['category']
        elif col_name_str in ['구분', '담당자', '업체명', '업체코드']:
            width = COLUMN_WIDTH_SETTINGS['text_short']
        elif '링크' in col_name_str or 'link' in col_name_str.lower() or 'url' in col_name_str.lower():
            width = COLUMN_WIDTH_SETTINGS['link']

        worksheet.column_dimensions[column_letter].width = width
    logger.debug("Finished applying column widths.")

def _apply_cell_styles_and_alignment(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Applies formatting (font, border, alignment) to header and data cells."""
    logger.debug("Applying cell styles and alignments.")
    # Header Styling
    for cell in worksheet[1]:  # First row is header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        
        # Enable wrap text for headers to display in 2 lines
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.alignment = header_alignment
        
        cell.border = DEFAULT_BORDER

    # Data Cell Styling
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = DEFAULT_FONT
            cell.border = DEFAULT_BORDER

            # Apply alignment based on column type
            col_name_str = str(col_name)
            is_pct_col = col_name_str in ['가격차이(2)(%)', '가격차이(3)(%)']  # Explicit check for percentage columns

            # Check if the cell value is likely numeric (ignoring error messages)
            is_numeric_value = False
            cell_value_str = str(cell.value)
            if cell_value_str not in ERROR_MESSAGE_VALUES and cell_value_str != '-':
                 # Basic check if it looks like a number
                 try:
                      float(cell_value_str.replace(',', '').replace('%',''))
                      is_numeric_value = True
                 except ValueError:
                      is_numeric_value = False

            # Apply right alignment to numbers and specifically formatted percentage strings
            quantity_columns = ['기본수량(1)', '기본수량(2)', '판매가(V포함)(2)', '기본수량(3)']
            if col_name_str in quantity_columns:
                cell.alignment = RIGHT_ALIGNMENT
            elif is_pct_col or ((col_name_str in PRICE_COLUMNS or col_name_str in QUANTITY_COLUMNS) and is_numeric_value):
                cell.alignment = RIGHT_ALIGNMENT
            # Update checks for center alignment based on new names
            elif col_name_str in IMAGE_COLUMNS or '코드' in col_name_str or 'Code' in col_name_str or col_name_str == '구분':
                 cell.alignment = CENTER_ALIGNMENT
            else:
                cell.alignment = LEFT_ALIGNMENT  # Default left align for text/links/errors
    logger.debug("Finished applying cell styles.")

def _apply_conditional_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Applies conditional formatting (e.g., yellow fill for price difference < -1)."""
    logger.debug("Applying conditional formatting.")

    # Find price difference columns (both regular and percentage) using new names
    price_diff_cols = [
        col for col in df.columns
        if col in ['가격차이(2)', '가격차이(3)', '가격차이(2)(%)', '가격차이(3)(%)']
    ]

    if not price_diff_cols:
        logger.debug("No price difference columns found for conditional formatting.")
        return

    # Define yellow fill
    yellow_fill = NEGATIVE_PRICE_FILL

    # First check if these columns actually exist in the DataFrame
    existing_diff_cols = [col for col in price_diff_cols if col in df.columns]
    if not existing_diff_cols:
        logger.warning("None of the price difference columns exist in the DataFrame. Skipping conditional formatting.")
        return

    # Add detailed logging for debugging
    logger.info(f"가격차이 조건부 서식 적용 시작 (음수 강조): {existing_diff_cols}")
    logger.info(f"총 확인할 행 수: {worksheet.max_row - 1}")  # Subtract 1 for header row
    
    # Log column data types for debugging
    for col in existing_diff_cols:
        logger.info(f"열 '{col}' 데이터 타입: {df[col].dtype}")
        # Try to count negative values in each column
        try:
            if df[col].dtype in ['int64', 'float64']:
                # For numeric columns, count directly
                neg_count = (df[col] < -1).sum()
                logger.info(f"열 '{col}'에서 -1 미만인 값: {neg_count}개")
            else:
                # For non-numeric columns, try to convert first
                try:
                    neg_count = (pd.to_numeric(df[col], errors='coerce') < -1).sum()
                    logger.info(f"열 '{col}'에서 -1 미만인 값: {neg_count}개 (변환 후)")
                except:
                    logger.warning(f"열 '{col}'의 값을 숫자로 변환할 수 없습니다.")
        except Exception as e:
            logger.warning(f"열 '{col}'에서 음수 값 계산 중 오류: {e}")
    
    rows_highlighted = 0
    rows_checked = 0
    errors = 0

    # Get the column indices in the Excel worksheet
    col_indices = {}
    for i, header in enumerate(df.columns, 1):
        col_indices[header] = i

    # Process each row
    for row_idx in range(2, worksheet.max_row + 1):  # Excel is 1-indexed, row 1 is header
        rows_checked += 1
        highlight_row = False
        
        # Check each price difference column
        for diff_col in existing_diff_cols:
            # Get the Excel column index
            if diff_col not in col_indices:
                continue
                
            col_idx = col_indices[diff_col]
            
            # Get the cell value directly from the worksheet
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            
            # Skip empty cells
            if cell_value is None or cell_value == '' or cell_value == '-':
                continue
                
            try:
                # Handle different types of values
                numeric_value = None
                
                if isinstance(cell_value, (int, float)):
                    # Direct numeric value
                    numeric_value = float(cell_value)
                elif isinstance(cell_value, str):
                    # Strip any whitespace, commas, currency symbols
                    clean_value = cell_value.strip().replace(',', '').replace(' ', '')
                    
                    # Handle parentheses format for negative numbers like (100)
                    if clean_value.startswith('(') and clean_value.endswith(')'):
                        clean_value = '-' + clean_value[1:-1]
                        
                    # Attempt conversion to float if it's not just a dash or empty
                    if clean_value and clean_value != '-':
                        try:
                            numeric_value = float(clean_value)
                        except ValueError:
                            # If conversion fails, skip this cell
                            continue
                
                # If we successfully got a numeric value and it's < -1, highlight the row
                if numeric_value is not None and numeric_value < -1:
                    highlight_row = True
                    logger.debug(f"음수 가격차이 발견: 행 {row_idx}, 열 '{diff_col}', 값 {numeric_value} < -1")
                    break
                    
            except Exception as e:
                logger.warning(f"행 {row_idx}, 열 '{diff_col}' 처리 중 오류: {e}")
                errors += 1
                continue
                
        # Apply highlighting to the entire row if needed
        if highlight_row:
            rows_highlighted += 1
            for col_idx in range(1, worksheet.max_column + 1):
                try:
                    # Apply yellow fill to all cells in the row
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow_fill
                except Exception as e:
                    logger.error(f"셀 서식 적용 오류 (행 {row_idx}, 열 {col_idx}): {e}")
                    errors += 1

    # Log summary of highlighting results
    logger.info(f"조건부 서식 적용 완료: {rows_highlighted}개 행에 가격차이 < -1 하이라이팅 적용됨 (검사 행: {rows_checked}, 오류: {errors})")

def _setup_page_layout(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    """Sets up page orientation, print area, freeze panes, etc."""
    logger.debug("Setting up page layout.")
    try:
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0  # Fit to width primarily
        worksheet.print_options.horizontalCentered = True
        worksheet.print_options.gridLines = False  # Typically false for final reports
        worksheet.freeze_panes = 'A2'  # Freeze header row
        logger.debug("Page layout settings applied.")
    except Exception as e:
        logger.error(f"Failed to set page layout options: {e}")

def _add_hyperlinks_to_worksheet(worksheet, df, hyperlinks_as_formulas=False):
    """
    Adds hyperlinks to URL cells in the worksheet.
    If hyperlinks_as_formulas=True, use Excel formulas for hyperlinks.
    Otherwise, use openpyxl's Hyperlink object.
    """
    try:
        # Define columns that should contain hyperlinks
        link_columns = [col for col in df.columns if any(term in col.lower() for term in ['링크', 'link', 'url'])]
        
        # Process each URL column
        total_urls_processed = 0
        
        for col in link_columns:
            if col in df.columns:
                col_idx = list(df.columns).index(col) + 1  # 1-based indexing for Excel
                
                # Loop through each cell in this column
                for row_idx, value in enumerate(df[col], 2):  # Start from row 2 (after header)
                    # Handle Series objects
                    if isinstance(value, pd.Series):
                        # Take the first non-empty value
                        for item in value:
                            if pd.notna(item) and item not in ['-', '']:
                                value = item
                                break
                        else:
                            value = ''
                    
                    # Skip empty values
                    if pd.isna(value) or value in ['', '-', 'None', 'nan']:
                        continue
                        
                    # Convert to string
                    url = str(value)
                    
                    # Extract URL from dictionary if needed
                    if isinstance(value, dict) and 'url' in value:
                        url = value['url']
                    
                    # Skip non-URL values
                    if not ('http://' in url or 'https://' in url or 'file:///' in url):
                        continue
                        
                    # Clean URL if needed
                    url = url.strip()
                    
                    try:
                        # Cell to apply hyperlink
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        if hyperlinks_as_formulas:
                            # Use Excel HYPERLINK formula
                            display_text = url
                            if len(display_text) > 50:
                                display_text = display_text[:47] + "..."
                            
                            cell.value = f'=HYPERLINK("{url}","{display_text}")'
                        else:
                            # Use openpyxl hyperlink object
                            cell.hyperlink = url
                            cell.value = url
                            
                            # Style for hyperlink
                            cell.font = Font(color="0563C1", underline="single")
                        
                        total_urls_processed += 1
                    except Exception as hyperlink_err:
                        logger.warning(f"Error adding hyperlink in row {row_idx}, col {col}: {hyperlink_err}")
                        # Keep original text if hyperlink fails
                        cell.value = url
                        
        logger.info(f"Processed link columns as plain text. Found {total_urls_processed} URLs across link columns.")
    except Exception as e:
        logger.warning(f"Error processing hyperlinks: {e}")
        logger.debug(traceback.format_exc())

def _add_header_footer(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    """Adds standard header and footer."""
    try:
        # Check if header_footer attribute exists (some versions don't support it)
        if hasattr(worksheet, 'header_footer'):
            from datetime import datetime
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
            worksheet.header_footer.center_header.text = "가격 비교 결과"
            worksheet.header_footer.right_header.text = f"생성일: {current_date}"
            worksheet.header_footer.left_footer.text = "해오름 RPA 가격 비교"
            worksheet.header_footer.right_footer.text = "페이지 &P / &N"
            logger.debug("Added header and footer to worksheet")
        else:
            logger.warning("Header/footer not supported in this Excel version - skipping")
    except Exception as e:
        logger.warning(f"Could not set header/footer: {e}")

def _apply_basic_excel_formatting(worksheet, column_list):
    """
    Applies basic Excel formatting to the worksheet:
    - Sets column widths
    - Applies header styles
    - Applies basic cell formatting
    """
    try:
        # 1. Set column widths based on content type
        for col_idx, col_name in enumerate(column_list, 1):
            # Default width based on column type
            if '이미지' in col_name or 'image' in col_name.lower():
                width = 30  # Image columns
            elif 'URL' in col_name or '링크' in col_name or 'link' in col_name.lower():
                width = 40  # URL columns
            elif '상품명' in col_name or '제품명' in col_name:
                width = 35  # Product name columns
            elif '코드' in col_name or 'code' in col_name.lower():
                width = 15  # Code columns
            else:
                width = 20  # Default width
            
            # Set column width
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
        
        # 2. Apply header style
        header_style = openpyxl.styles.NamedStyle(name='header_style')
        header_style.font = Font(bold=True, size=11)
        header_style.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header style to first row
        for col_idx in range(1, len(column_list) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            for attr_name, attr_value in header_style.__dict__.items():
                if attr_name not in ['name', '_StyleProxy__target'] and hasattr(cell, attr_name):
                    setattr(cell, attr_name, attr_value)
        
        # Make header row taller
        worksheet.row_dimensions[1].height = 30
        
        # 3. Apply basic data cell formatting
        data_style = openpyxl.styles.NamedStyle(name='data_style')
        data_style.alignment = Alignment(vertical='center', wrap_text=True)
        data_style.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get the number of rows in the worksheet (excluding header)
        max_row = worksheet.max_row
        
        # Apply data style to all data cells
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, len(column_list) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Apply border and base alignment to all cells
                cell.border = data_style.border
                cell.alignment = data_style.alignment
                
                # Specific formatting for certain column types
                col_name = column_list[col_idx - 1]
                
                # Price columns - right align and format as number
                if '단가' in col_name or '가격' in col_name or 'price' in col_name.lower():
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                # Code/ID columns - center align
                elif '코드' in col_name or 'ID' in col_name or 'id' in col_name.lower():
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                # URL/Link columns - left align
                elif 'URL' in col_name or '링크' in col_name or 'link' in col_name.lower():
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                # Regular text columns - left align
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 4. Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Explicitly remove any existing filter
        if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
            worksheet.auto_filter.ref = None
        
        logger.debug(f"Applied basic Excel formatting to worksheet (header + {max_row-1} data rows)")
        
    except Exception as e:
        logger.warning(f"Error applying basic Excel formatting: {e}")
        logger.debug(traceback.format_exc())

def _apply_upload_file_formatting(worksheet, column_list):
    """
    Applies specific formatting for the upload Excel file:
    - Headers with gray background and 2 lines display
    - Content with wrap text
    - Specific cell dimensions and borders
    """
    try:
        # 1. Set standard column width (7) for all columns
        for col_idx in range(1, len(column_list) + 1):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 7
        
        # Adjust specific columns that need different widths
        special_width_columns = {
            '상품명': 35,  # 상품명
            '상품코드': 12, # Product code
            '카테고리(중분류)': 15,  # Category
            '해오름(이미지링크)': 40, # Image URLs
            '고려기프트(이미지링크)': 40,
            '네이버쇼핑(이미지링크)': 40,
            '본사링크': 30, # Product links
            '고려 링크': 30,
            '네이버 링크': 30
        }
        
        for idx, col_name in enumerate(column_list, 1):
            column_letter = get_column_letter(idx)
            # Check if this column needs special width
            for special_name, width in special_width_columns.items():
                if special_name in col_name:
                    worksheet.column_dimensions[column_letter].width = width
                    break
        
        # 2. Set row heights - header row = 34.5, data rows = 16.9
        worksheet.row_dimensions[1].height = UPLOAD_HEADER_HEIGHT  # Header row height
        
        # Set data row heights
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = UPLOAD_DATA_ROW_HEIGHT
        
        # 3. Apply header formatting - gray background and wrap text
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx in range(1, len(column_list) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            # Apply gray background
            cell.fill = UPLOAD_HEADER_FILL
            # Enable text wrapping for 2-line display
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Add borders
            cell.border = thin_border
            # Bold font
            cell.font = Font(bold=True, size=10)
        
        # 4. Apply data row formatting - wrap text and borders
        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in range(1, len(column_list) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # Enable text wrapping (fit to cell)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                # Add borders
                cell.border = thin_border
                
                # Adjust alignment based on column content
                col_name = column_list[col_idx - 1]
                # Right-align numeric columns
                if any(term in col_name for term in ['단가', '가격차이', '기본수량']):
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                # Center-align code/ID columns
                elif any(term in col_name for term in ['코드', 'Code', '구분']):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Left-align everything else
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 5. Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Remove any existing auto-filter
        if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
            worksheet.auto_filter.ref = None
        
        logger.info(f"Applied upload file specific formatting to worksheet with {worksheet.max_row} rows.")
        
    except Exception as e:
        logger.warning(f"Error applying upload file formatting: {e}")
        logger.debug(traceback.format_exc()) 

class ExcelFormatter:
    """Class to handle all Excel formatting operations"""
    
    def format_result_file(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                          df: pd.DataFrame) -> None:
        """Format the result Excel file"""
        try:
            logger.info("Applying result file formatting...")
            
            self._apply_column_widths(worksheet, df)
            self._apply_cell_styles_and_alignment(worksheet, df)
            self._add_header_footer(worksheet)
            self._setup_page_layout(worksheet)
            self._apply_conditional_formatting(worksheet, df)
            
            # Remove auto filter
            if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
                worksheet.auto_filter.ref = None
                
            logger.info("Successfully applied result file formatting")
            
        except Exception as e:
            logger.error(f"Error in format_result_file: {e}")
            raise
    
    def format_upload_file(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                          df: pd.DataFrame) -> None:
        """Format the upload Excel file"""
        try:
            logger.info("Applying upload file formatting...")
            
            # Set basic dimensions
            self._set_upload_dimensions(worksheet, df)
            
            # Apply header formatting
            self._apply_upload_header_formatting(worksheet)
            
            # Apply data formatting
            self._apply_upload_data_formatting(worksheet, df)
            
            # Remove auto filter
            if hasattr(worksheet, 'auto_filter') and worksheet.auto_filter:
                worksheet.auto_filter.ref = None
                
            logger.info("Successfully applied upload file formatting")
            
        except Exception as e:
            logger.error(f"Error in format_upload_file: {e}")
            raise
    
    def _set_upload_dimensions(self, worksheet, df: pd.DataFrame) -> None:
        """Set dimensions for upload file"""
        try:
            # Set standard column width
            for col_idx in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = 7
            
            # Set special column widths
            special_widths = {
                '상품명': 35,
                '상품코드': 12,
                '카테고리(중분류)': 15,
                '해오름(이미지링크)': 40,
                '고려기프트(이미지링크)': 40,
                '네이버쇼핑(이미지링크)': 40,
                '본사링크': 30,
                '고려 링크': 30,
                '네이버 링크': 30
            }
            
            for idx, col_name in enumerate(df.columns, 1):
                for special_name, width in special_widths.items():
                    if special_name in col_name:
                        worksheet.column_dimensions[get_column_letter(idx)].width = width
                        break
            
            # Set row heights
            worksheet.row_dimensions[1].height = UPLOAD_HEADER_HEIGHT
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = UPLOAD_DATA_ROW_HEIGHT
                
        except Exception as e:
            logger.error(f"Error setting upload dimensions: {e}")
            raise

    # ... [Rest of the existing methods with added error handling] ... 