import os
import logging
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from PIL import Image as PILImage
from io import BytesIO
import pandas as pd
import traceback
from pathlib import Path
from typing import Optional, Dict, Any, List, Union, Tuple

# Import constants
from excel_constants import (
    IMAGE_COLUMNS, HAEREUM_DIR_NAME, KOGIFT_DIR_NAME, 
    NAVER_DIR_NAME, OTHER_DIR_NAME, IMAGE_MAIN_DIR,
    IMAGE_MAX_SIZE, IMAGE_STANDARD_SIZE, RESAMPLING_FILTER
)
from excel_style_constants import (
    RESULT_IMAGE_WIDTH, RESULT_IMAGE_HEIGHT, RESULT_DATA_ROW_HEIGHT, COLUMN_WIDTH_SETTINGS, LINK_FONT, IMAGE_SETTINGS
)

# Initialize logger
logger = logging.getLogger(__name__)

# Define image processing constants
RESAMPLING_FILTER = Image.Resampling.LANCZOS  # High-quality resampling
MAX_IMAGE_SIZE = IMAGE_SETTINGS['MAX_SIZE']
STANDARD_IMAGE_SIZE = IMAGE_SETTINGS['STANDARD_SIZE']
SUPPORTED_FORMATS = IMAGE_SETTINGS['SUPPORTED_FORMATS']

class ImageProcessor:
    """이미지 처리 및 Excel 파일 내 이미지 관리를 위한 클래스"""
    
    def __init__(self):
        self.max_image_size = (160, 160)  # 최대 이미지 크기 (width, height)
    
    def process_image_for_excel(self, image_path: str) -> Optional[Dict[str, Any]]:
        """Excel에 삽입할 이미지를 처리합니다."""
        try:
            if not os.path.exists(image_path):
                logger.warning(f"Image file not found: {image_path}")
                return None
                
            # 이미지 크기 및 형식 검증
            with PILImage.open(image_path) as img:
                # 이미지 크기 조정
                img.thumbnail(self.max_image_size)
                
                # 이미지 정보 반환
                return {
                    'path': image_path,
                    'size': img.size,
                    'format': img.format
                }
                
        except Exception as e:
            logger.error(f"Error processing image {image_path}: {e}")
            return None
    
    def add_image_to_worksheet(self, 
                             worksheet: openpyxl.worksheet.worksheet.Worksheet,
                             image_path: str,
                             row: int,
                             col: int) -> bool:
        """워크시트에 이미지를 추가합니다."""
        try:
            if not os.path.exists(image_path):
                logger.warning(f"Image file not found: {image_path}")
                return False
                
            # 이미지 객체 생성
            img = Image(image_path)
            img.width, img.height = self.max_image_size
            
            # 이미지 위치 설정
            col_letter = get_column_letter(col)
            img.anchor = f"{col_letter}{row}"
            
            # 워크시트에 이미지 추가
            worksheet.add_image(img)
            
            # 행 높이 조정
            worksheet.row_dimensions[row].height = 120
            
            return True
            
        except Exception as e:
            logger.error(f"Error adding image to worksheet: {e}")
            return False
    
    def extract_image_url(self, image_data: Dict[str, Any]) -> str:
        """이미지 데이터에서 URL을 추출합니다."""
        try:
            if isinstance(image_data, dict):
                return image_data.get('url', '')
            elif isinstance(image_data, str) and image_data.startswith(('http://', 'https://')):
                return image_data
            return ''
        except Exception as e:
            logger.error(f"Error extracting image URL: {e}")
            return ''

def safe_load_image(path, max_height=150, max_width=150):
    """Safely load and resize an image for Excel."""
    try:
        img = PILImage.open(path)
        # Calculate new dimensions preserving aspect ratio
        width, height = img.size
        if width > max_width or height > max_height:
            ratio = min(max_width / width, max_height / height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            img = img.resize((new_width, new_height), RESAMPLING_FILTER)
            
            # Save temporary resized version
            temp_dir = os.path.join(os.path.dirname(path), 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, f"resized_{os.path.basename(path)}")
            img.save(temp_path)
            return temp_path
        return path
    except Exception as e:
        logger.warning(f"Error loading/resizing image {path}: {e}")
        return None

def _process_image_columns(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Process and add images to Excel worksheet."""
    try:
        # Set image dimensions
        image_width = RESULT_IMAGE_WIDTH
        image_height = RESULT_IMAGE_HEIGHT
        
        # Get image columns
        image_cols = [col for col in df.columns if col in IMAGE_COLUMNS]
        
        # Process each row
        for row_idx in range(2, worksheet.max_row + 1):
            # Set row height for image rows
            worksheet.row_dimensions[row_idx].height = RESULT_DATA_ROW_HEIGHT
            
            # Process each image column
            for col in image_cols:
                col_idx = df.columns.get_loc(col) + 1
                col_letter = get_column_letter(col_idx)
                
                # Set column width
                worksheet.column_dimensions[col_letter].width = COLUMN_WIDTH_SETTINGS['image']
                
                # Get image data
                img_value = df.iloc[row_idx - 2][col]
                
                # Skip empty cells
                if pd.isna(img_value) or img_value == '' or img_value == '-':
                    continue
                
                # Get image path
                img_path = None
                if isinstance(img_value, dict):
                    if 'local_path' in img_value and os.path.exists(img_value['local_path']):
                        img_path = img_value['local_path']
                elif isinstance(img_value, str) and os.path.exists(img_value):
                    img_path = img_value
                
                # Add image if path exists
                if img_path:
                    try:
                        img = openpyxl.drawing.image.Image(img_path)
                        img.width = image_width
                        img.height = image_height
                        img.anchor = f"{col_letter}{row_idx}"
                        worksheet.add_image(img)
                        
                        # Clear cell content
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = ""
                        
                        # Add hyperlink if URL exists
                        if isinstance(img_value, dict) and 'url' in img_value:
                            cell.hyperlink = img_value['url']
                            cell.font = LINK_FONT
                    except Exception as e:
                        logger.error(f"Error adding image at row {row_idx}, column {col}: {e}")
                        
                        # Fallback to URL if available
                        if isinstance(img_value, dict) and 'url' in img_value:
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.value = img_value['url']
                            cell.hyperlink = img_value['url']
                            cell.font = LINK_FONT
                            
    except Exception as e:
        logger.error(f"Error in _process_image_columns: {e}")
        logger.debug(traceback.format_exc())

def _adjust_image_cell_dimensions(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """Adjusts row heights and column widths for cells containing images."""
    from excel_constants import ERROR_MESSAGE_VALUES
    
    logger.debug("Adjusting dimensions for image cells...")
    
    # Get image column indices using the IMAGE_COLUMNS constant
    image_cols = {col: idx for idx, col in enumerate(df.columns, 1) if col in IMAGE_COLUMNS}
    
    if not image_cols:
        return

    # Increase column widths for image columns to accommodate larger images
    for col_name, col_idx in image_cols.items():
        try:
            col_letter = get_column_letter(col_idx)
            # Use larger column width for image columns
            worksheet.column_dimensions[col_letter].width = 85  # Increased from 80
        except Exception as e:
            logger.error(f"Error adjusting column width for {col_name}: {e}")
    
    # Create a set of rows that need height adjustment
    rows_with_images = set()
    
    try:
        # Find rows that have actual images (not error messages or empty cells)
        for row_idx in range(2, worksheet.max_row + 1):
            for col_name, col_idx in image_cols.items():
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # If the cell is empty, it likely has an image
                    if cell.value == "" or cell.value is None:
                        rows_with_images.add(row_idx)
                        break
                        
                    # Check for image data in dictionary format
                    cell_value = str(cell.value) if cell.value else ""
                    if cell_value and cell_value.startswith('{') and cell_value.endswith('}'):
                        try:
                            import ast
                            img_dict = ast.literal_eval(cell_value)
                            if isinstance(img_dict, dict) and ('local_path' in img_dict or 'url' in img_dict):
                                rows_with_images.add(row_idx)
                                break
                        except:
                            pass
                            
                    # Check for path-like strings
                    if (cell_value and cell_value != '-' and 
                        not any(err_msg in cell_value for err_msg in ERROR_MESSAGE_VALUES) and
                        ('\\' in cell_value or '/' in cell_value or '.jpg' in cell_value.lower() or 
                         '.png' in cell_value.lower() or '.jpeg' in cell_value.lower() or
                         'http' in cell_value.lower())):
                        rows_with_images.add(row_idx)
                        break
                except Exception as e:
                    logger.error(f"Error checking cell at row {row_idx}, column {col_idx}: {e}")
    except Exception as e:
        logger.error(f"Error finding rows with images: {e}")
    
    # Apply increased height to rows with images
    for row_idx in rows_with_images:
        try:
            # Set larger row height to accommodate bigger images
            worksheet.row_dimensions[row_idx].height = 400  # Increased from 380
            
            # Center-align all cells in this row for better appearance with images
            for col_idx in range(1, worksheet.max_column + 1):
                try:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    # Preserve horizontal alignment, set vertical to center
                    current_alignment = cell.alignment
                    cell.alignment = Alignment(
                        horizontal=current_alignment.horizontal,
                        vertical="center",
                        wrap_text=current_alignment.wrap_text
                    )
                except Exception as e:
                    logger.error(f"Error adjusting cell alignment at row {row_idx}, column {col_idx}: {e}")
        except Exception as e:
            logger.error(f"Error adjusting row height for row {row_idx}: {e}")
    
    logger.debug(f"Adjusted dimensions for {len(rows_with_images)} rows with images")

def find_best_image_file(base_path: str, filename: str) -> Optional[str]:
    """
    Find the best available image file, prioritizing jpg format.
    
    Args:
        base_path: Base directory to search in
        filename: Original filename without extension
        
    Returns:
        Optional[str]: Path to best available image file
    """
    # Remove any existing extension from filename
    filename = os.path.splitext(filename)[0]
    
    # Define priority list for image formats (jpg first)
    formats = [
        '.jpg',       # JPEG (highest priority)
        '.jpeg',      # Alternative JPEG
        '_nobg.png',  # Transparent background version
        '.png',       # Regular PNG
        '.webp',      # WebP
        '.gif'        # GIF
    ]
    
    # Try each format in priority order
    for fmt in formats:
        test_path = os.path.join(base_path, filename + fmt)
        if os.path.exists(test_path) and os.path.getsize(test_path) > 0:
            return test_path
            
    return None

def resize_image_for_excel(img_path: str, max_size: Tuple[int, int] = (160, 160)) -> Optional[str]:
    """
    Resize image while maintaining aspect ratio and quality.
    
    Args:
        img_path: Path to original image
        max_size: Maximum dimensions (width, height)
        
    Returns:
        Optional[str]: Path to resized image
    """
    try:
        with Image.open(img_path) as img:
            # Calculate new dimensions preserving aspect ratio
            width, height = img.size
            ratio = min(max_size[0] / width, max_size[1] / height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            
            # Only resize if needed
            if new_width < width or new_height < height:
                # Use high-quality resampling
                img = img.resize((new_width, new_height), RESAMPLING_FILTER)
                
                # Save with quality preservation
                temp_dir = os.path.join(os.path.dirname(img_path), 'temp')
                os.makedirs(temp_dir, exist_ok=True)
                
                # Use original format if possible
                temp_path = os.path.join(temp_dir, f"resized_{os.path.basename(img_path)}")
                
                # Save with format-specific settings
                if img.format == 'PNG':
                    img.save(temp_path, 'PNG', optimize=True)
                elif img.format in ['JPEG', 'JPG']:
                    img.save(temp_path, 'JPEG', quality=85, optimize=True)
                else:
                    img.save(temp_path, img.format if img.format else 'PNG')
                    
                return temp_path
            
            return img_path
            
    except Exception as e:
        logger.error(f"Error resizing image {img_path}: {e}")
        return None

def process_image_for_excel(image_path: str, is_result_file: bool = True) -> Optional[Dict[str, Any]]:
    """
    Process image for Excel insertion with improved handling.
    
    Args:
        image_path: Path to image file
        is_result_file: Whether this is for result file (True) or upload file (False)
        
    Returns:
        Optional[Dict[str, Any]]: Processed image information
    """
    try:
        if not os.path.exists(image_path):
            return None
            
        # Try to find _nobg version first
        base_path = os.path.dirname(image_path)
        filename = os.path.splitext(os.path.basename(image_path))[0]
        best_image = find_best_image_file(base_path, filename)
        
        if not best_image:
            return None
            
        # For result file, resize image
        if is_result_file:
            resized_path = resize_image_for_excel(best_image)
            if not resized_path:
                return None
                
            return {
                'path': resized_path,
                'original_path': image_path,
                'is_nobg': '_nobg' in best_image.lower()
            }
        else:
            # For upload file, just return path info
            return {
                'path': best_image,
                'original_path': image_path,
                'is_nobg': '_nobg' in best_image.lower()
            }
            
    except Exception as e:
        logger.error(f"Error processing image {image_path}: {e}")
        return None

def add_image_to_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                         image_info: Dict[str, Any],
                         row: int,
                         col: int,
                         is_result_file: bool = True) -> bool:
    """
    Add image to worksheet with improved handling.
    
    Args:
        worksheet: Target worksheet
        image_info: Image information dictionary
        row: Target row
        col: Target column
        is_result_file: Whether this is for result file (True) or upload file (False)
        
    Returns:
        bool: Success status
    """
    try:
        if not image_info or 'path' not in image_info:
            return False
            
        img_path = image_info['path']
        if not os.path.exists(img_path):
            return False
            
        # For result file, add actual image
        if is_result_file:
            try:
                img = openpyxl.drawing.image.Image(img_path)
                
                # Use constants from excel_style_constants
                img.width = RESULT_IMAGE_WIDTH
                img.height = RESULT_IMAGE_HEIGHT
                
                # Position image
                col_letter = get_column_letter(col)
                img.anchor = f"{col_letter}{row}"
                
                # Add image to worksheet
                worksheet.add_image(img)
                
                # Adjust row height
                worksheet.row_dimensions[row].height = RESULT_DATA_ROW_HEIGHT
                
                return True
                
            except Exception as e:
                logger.error(f"Error adding image to worksheet: {e}")
                return False
                
        # For upload file, just add URL or path
        else:
            cell = worksheet.cell(row=row, column=col)
            cell.value = image_info.get('original_path', '')
            return True
            
    except Exception as e:
        logger.error(f"Error in add_image_to_worksheet: {e}")
        return False

__all__ = [
    'ImageProcessor',
    'safe_load_image',
    '_process_image_columns',
    '_adjust_image_cell_dimensions',
    'find_best_image_file',
    'resize_image_for_excel',
    'process_image_for_excel',
    'add_image_to_worksheet'
] 