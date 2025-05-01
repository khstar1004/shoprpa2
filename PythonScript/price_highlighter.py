import os
import logging
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd

# 로깅 설정
logger = logging.getLogger(__name__)

def highlight_negative_price_differences(excel_path, threshold=-1):
    """
    엑셀 파일에서 가격차이 관련 컬럼 값이 지정된 threshold보다 작으면 해당 행 전체를 노란색으로 표시합니다.
    
    Args:
        excel_path (str): 대상 엑셀 파일 경로
        threshold (float, optional): 하이라이팅 적용 기준값. 기본값은 -1.
        
    Returns:
        bool: 성공 여부
    """
    if not os.path.exists(excel_path):
        logger.error(f"엑셀 파일이 존재하지 않습니다: {excel_path}")
        return False
    
    try:
        logger.info(f"가격차이 하이라이팅 시작: {excel_path}")
        
        # 엑셀 파일 로드
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.active
        
        # 노란색 배경 스타일
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 헤더 행에서 가격차이 컬럼 인덱스 찾기
        price_diff_columns = []
        header_row = 1  # 헤더는 1행에 있다고 가정
        
        # 모든 컬럼 명칭 확인
        headers = []
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col_idx).value
            if cell_value:
                headers.append((col_idx, str(cell_value)))
        
        logger.debug(f"엑셀 헤더: {[h[1] for h in headers]}")
        
        # 가격차이 관련 컬럼 찾기 - 정확한 컬럼명 매칭
        exact_price_columns = [
            "가격차이(2)", "가격차이(3)", 
            "가격차이(2)(%)", "가격차이(3)(%)"
        ]
        
        for col_idx, header in headers:
            # 정확한 컬럼명 매칭
            if header in exact_price_columns:
                price_diff_columns.append((col_idx, header))
                logger.debug(f"정확한 가격차이 컬럼 매칭: {header} (열 {col_idx})")
            # 부분 문자열 매칭 (정확한 매칭이 없는 경우 대비)
            elif "가격차이" in header:
                price_diff_columns.append((col_idx, header))
                logger.debug(f"부분 가격차이 컬럼 매칭: {header} (열 {col_idx})")
        
        if not price_diff_columns:
            logger.warning("가격차이 관련 컬럼을 찾을 수 없습니다.")
            return False
        
        # 각 행을 처리하며 가격차이 컬럼 값 확인
        total_rows = worksheet.max_row
        rows_highlighted = 0
        errors = 0
        
        logger.info(f"총 {total_rows-1}개 행에 대해 가격차이 검사 시작 ({len(price_diff_columns)}개 컬럼)")
        
        for row_idx in range(2, total_rows + 1):  # 헤더 이후부터 시작
            highlight_row = False
            negative_values = []
            
            for col_idx, col_name in price_diff_columns:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                # 값이 있고 비교 가능한 경우에만 처리
                if cell_value not in [None, "", "-"]:
                    try:
                        # 문자열을 숫자로 변환 (쉼표, 공백 제거)
                        if isinstance(cell_value, str):
                            # 괄호로 표현된 음수 처리 (예: "(100)" -> "-100")
                            cleaned_value = cell_value.replace(",", "").replace(" ", "")
                            if cleaned_value.startswith("(") and cleaned_value.endswith(")"):
                                cleaned_value = "-" + cleaned_value[1:-1]
                            numeric_value = float(cleaned_value)
                        else:
                            numeric_value = float(cell_value)
                        
                        # 음수(-1 미만) 확인 및 기록
                        if numeric_value < threshold:
                            highlight_row = True
                            negative_values.append(f"{col_name} = {numeric_value}")
                            logger.debug(f"행 {row_idx}: {col_name} = {numeric_value} < {threshold}, 하이라이팅 대상")
                    except (ValueError, TypeError) as e:
                        logger.debug(f"행 {row_idx}, 열 '{col_name}': 숫자 변환 실패 '{cell_value}': {e}")
                        errors += 1
            
            # 조건을 만족하면 행 전체를 노란색으로 표시
            if highlight_row:
                for col_idx in range(1, worksheet.max_column + 1):
                    try:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = yellow_fill
                    except Exception as e:
                        logger.error(f"셀 서식 적용 오류 (행 {row_idx}, 열 {col_idx}): {e}")
                        errors += 1
                rows_highlighted += 1
                logger.info(f"행 {row_idx} 하이라이팅 적용: {', '.join(negative_values)}")
        
        # 결과 저장
        workbook.save(excel_path)
        logger.info(f"가격차이 하이라이팅 완료: {rows_highlighted}개 행에 적용 (오류: {errors}개)")
        return True
    
    except Exception as e:
        logger.error(f"가격차이 하이라이팅 중 오류 발생: {e}", exc_info=True)
        return False


def apply_price_highlighting_to_files(result_path=None, upload_path=None, threshold=-1):
    """
    생성된 결과 엑셀 파일들에 가격차이 하이라이팅을 적용합니다.
    
    Args:
        result_path (str, optional): 결과 엑셀 파일 경로
        upload_path (str, optional): 업로드용 엑셀 파일 경로
        threshold (float, optional): 하이라이팅 적용 기준값. 기본값은 -1.
        
    Returns:
        tuple: (성공한 파일 수, 총 파일 수)
    """
    success_count = 0
    total_files = 0
    
    # 유효한 경로만 처리
    paths_to_process = []
    if result_path and os.path.exists(result_path):
        paths_to_process.append(("결과 파일", result_path))
    if upload_path and os.path.exists(upload_path):
        paths_to_process.append(("업로드 파일", upload_path))
    
    total_files = len(paths_to_process)
    
    if total_files == 0:
        logger.warning("처리할 엑셀 파일이 없습니다.")
        return 0, 0
    
    # 각 파일에 하이라이팅 적용
    for file_type, file_path in paths_to_process:
        logger.info(f"{file_type} 가격차이 하이라이팅 시작: {os.path.basename(file_path)}")
        success = highlight_negative_price_differences(file_path, threshold)
        if success:
            success_count += 1
            logger.info(f"{file_type} 가격차이 하이라이팅 성공: {os.path.basename(file_path)}")
        else:
            logger.error(f"{file_type} 가격차이 하이라이팅 실패: {os.path.basename(file_path)}")
    
    return success_count, total_files 