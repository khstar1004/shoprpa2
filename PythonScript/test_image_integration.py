import os
import pandas as pd
import logging
from pathlib import Path
import configparser
from haereum_images_to_excel import main as haereum_to_excel
from image_integration import integrate_and_filter_images
from enhanced_image_matcher import EnhancedImageMatcher
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import shutil

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

def create_excel_with_images(df, output_file):
    """이미지가 포함된 엑셀 파일 생성"""
    try:
        # 임시 디렉토리 생성
        temp_dir = Path("temp_images")
        temp_dir.mkdir(exist_ok=True)
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        
        # 헤더 작성
        headers = ['번호', '상품명', '파일명', '본사 이미지', '고려기프트 이미지', '네이버 이미지', '이미지_유사도']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 행 높이 설정
        ws.row_dimensions[1].height = 30  # 헤더 행 높이
        for row in range(2, len(df) + 2):
            ws.row_dimensions[row].height = 100  # 데이터 행 높이
        
        # 열 너비 설정
        column_widths = {'A': 5, 'B': 30, 'C': 30, 'D': 15, 'E': 15, 'F': 15, 'G': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 데이터 및 이미지 추가
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            # 기본 데이터 추가
            ws.cell(row=row_idx, column=1, value=row['번호'])
            ws.cell(row=row_idx, column=2, value=row['상품명'])
            ws.cell(row=row_idx, column=3, value=row['파일명'])
            ws.cell(row=row_idx, column=7, value=row['이미지_유사도'])
            
            # 이미지 추가
            image_columns = {
                '본사 이미지': row['본사 이미지'],
                '고려기프트 이미지': row['고려기프트 이미지'],
                '네이버 이미지': row['네이버 이미지']
            }
            
            for col_idx, (col_name, img_path) in enumerate(image_columns.items(), 4):
                if img_path and pd.notna(img_path):
                    try:
                        # 이미지 파일 복사
                        img = Image(img_path)
                        # 이미지 크기 조정 (최대 100x100)
                        img.width = 100
                        img.height = 100
                        # 이미지 추가
                        ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")
                    except Exception as e:
                        logging.warning(f"이미지 추가 실패 ({img_path}): {e}")
                        ws.cell(row=row_idx, column=col_idx, value=str(img_path))
                else:
                    ws.cell(row=row_idx, column=col_idx, value="")
        
        # 엑셀 파일 저장
        wb.save(output_file)
        logging.info(f"이미지가 포함된 엑셀 파일이 저장되었습니다: {output_file}")
        
        # 임시 디렉토리 정리
        shutil.rmtree(temp_dir)
        
    except Exception as e:
        logging.error(f"엑셀 파일 생성 중 오류 발생: {e}", exc_info=True)

def test_image_integration():
    """이미지 통합 및 유사도 검사 테스트"""
    try:
        # 1. 설정 파일 로드
        config = configparser.ConfigParser()
        config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'config.ini')
        config.read(config_path, encoding='utf-8')
        
        # 2. 이미지 디렉토리 설정
        main_dir = Path("C:/RPA/Image/Main")
        haereum_dir = main_dir / "Haereum"
        kogift_dir = main_dir / "kogift"
        naver_dir = main_dir / "naver"
        
        # 3. 해오름 이미지 목록 가져오기
        haereum_images = sorted([f for f in haereum_dir.glob("*.jpg") if "_nobg" not in f.name])
        logging.info(f"해오름 이미지 {len(haereum_images)}개 발견")
        
        # 4. 초기 DataFrame 생성
        test_data = []
        for i, img_path in enumerate(haereum_images):
            product_name = img_path.stem
            if product_name.startswith('haereum_'):
                product_name = product_name[8:]  # 'haereum_' 접두사 제거
            product_name = product_name.replace('_', ' ')
            
            test_data.append({
                "번호": i+1,
                "상품명": product_name,
                "파일명": img_path.name,
                "본사 이미지": str(img_path),
                "고려기프트 이미지": None,
                "네이버 이미지": None,
                "이미지_유사도": None
            })
        
        df = pd.DataFrame(test_data)
        logging.info(f"초기 DataFrame 생성 완료: {len(df)}개 행")
        
        # 5. 이미지 유사도 계산을 위한 matcher 초기화
        matcher = EnhancedImageMatcher()
        
        # 6. 고려기프트 이미지와 유사도 계산
        kogift_images = sorted([f for f in kogift_dir.glob("*.jpg")])
        logging.info(f"고려기프트 이미지 {len(kogift_images)}개 발견")
        
        for i, row in df.iterrows():
            haereum_img = row['본사 이미지']
            best_similarity = 0
            best_kogift_img = None
            
            for kogift_img in kogift_images:
                similarity = matcher.calculate_similarity(haereum_img, str(kogift_img))
                if similarity > best_similarity:
                    best_similarity = similarity
                    best_kogift_img = str(kogift_img)
            
            if best_kogift_img:
                df.at[i, '고려기프트 이미지'] = best_kogift_img
                df.at[i, '이미지_유사도'] = best_similarity
                logging.info(f"상품 {row['상품명']} - 고려기프트 유사도: {best_similarity:.2f}")
        
        # 7. 네이버 이미지와 유사도 계산
        naver_images = sorted([f for f in naver_dir.glob("*.jpg")])
        logging.info(f"네이버 이미지 {len(naver_images)}개 발견")
        
        for i, row in df.iterrows():
            haereum_img = row['본사 이미지']
            best_similarity = 0
            best_naver_img = None
            
            for naver_img in naver_images:
                similarity = matcher.calculate_similarity(haereum_img, str(naver_img))
                if similarity > best_similarity:
                    best_similarity = similarity
                    best_naver_img = str(naver_img)
            
            if best_naver_img:
                df.at[i, '네이버 이미지'] = best_naver_img
                # 기존 유사도와 비교하여 더 높은 값으로 업데이트
                current_similarity = df.at[i, '이미지_유사도'] or 0
                if best_similarity > current_similarity:
                    df.at[i, '이미지_유사도'] = best_similarity
                logging.info(f"상품 {row['상품명']} - 네이버 유사도: {best_similarity:.2f}")
        
        # 8. 유사도 기반 필터링
        result_df = integrate_and_filter_images(df, config)
        
        # 9. 결과 출력
        logging.info("\n=== 최종 결과 ===")
        for _, row in result_df.iterrows():
            logging.info(f"\n상품명: {row['상품명']}")
            logging.info(f"본사 이미지: {row['본사 이미지']}")
            logging.info(f"고려기프트 이미지: {row['고려기프트 이미지']}")
            logging.info(f"네이버 이미지: {row['네이버 이미지']}")
            logging.info(f"유사도: {row['이미지_유사도']}")
        
        # 10. 결과를 엑셀 파일로 저장 (이미지 포함)
        output_dir = Path("C:/RPA/Output")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = output_dir / "test_image_integration_results.xlsx"
        create_excel_with_images(result_df, output_file)
        
    except Exception as e:
        logging.error(f"테스트 중 오류 발생: {e}", exc_info=True)

if __name__ == "__main__":
    test_image_integration() 