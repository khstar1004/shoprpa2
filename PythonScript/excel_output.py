import os
import logging
import pandas as pd
import openpyxl
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List, Union, Tuple
import json
from PIL import Image
import shutil
from openpyxl.utils import get_column_letter

# Import from other modules
from excel_constants import (
    FINAL_COLUMN_ORDER,
    UPLOAD_COLUMN_ORDER, IMAGE_COLUMNS,
    REQUIRED_INPUT_COLUMNS,
    UPLOAD_COLUMN_MAPPING
)
from excel_data_processing import (
    flatten_nested_image_dicts, prepare_naver_image_urls_for_upload,
    _prepare_data_for_excel, finalize_dataframe_for_excel
)
from excel_formatting import (
    _apply_basic_excel_formatting, _apply_upload_file_formatting,
    _add_hyperlinks_to_worksheet, _add_header_footer,
    _apply_conditional_formatting, ExcelFormatter
)
from excel_image_utils import (
    _process_image_columns, ImageProcessor,
    _adjust_image_cell_dimensions
)
from excel_utils import sanitize_dataframe_for_excel

# Initialize logger
logger = logging.getLogger(__name__)

def safe_excel_operation(func):
    """
    데코레이터: Excel 작업 중 발생할 수 있는 예외를 안전하게 처리합니다.
    """
    import functools
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logging.error(f"Excel operation failed in {func.__name__}: {str(e)}", exc_info=True)
            return False
    return wrapper

@safe_excel_operation
def create_split_excel_outputs(df_finalized: pd.DataFrame, output_path_base: str) -> Tuple[bool, bool, Optional[str], Optional[str]]:
    """
    작업메뉴얼에 따라 두 가지 Excel 파일을 생성합니다:
    1. Result file (A): 이미지 포함, 조회용 (원본 컬럼 이름 유지)
    2. Upload file (P): URL 링크만 포함, 업로드용 (컬럼 이름 변환)

    Args:
        df_finalized: 최종 처리된 DataFrame
        output_path_base: 출력 파일의 기본 경로

    Returns:
        tuple: (result_success, upload_success, result_path, upload_path)
    """
    # Ensure we have valid data
    if df_finalized is None or df_finalized.empty:
        logger.error("No data to write to Excel. DataFrame is empty or None.")
        return False, False, None, None

    # Flatten any nested image dictionaries to prevent Excel conversion errors
    df_finalized = flatten_nested_image_dicts(df_finalized)
    
    # Sanitize DataFrame to make sure all values are Excel-compatible
    df_finalized = sanitize_dataframe_for_excel(df_finalized)
    
    logger.info(f"Starting creation of split Excel outputs from finalized DataFrame (Shape: {df_finalized.shape})")
    
    # Default return values
    result_path = None
    result_success = False
    upload_path = None
    upload_success = False

    try:
        # -----------------------------------------
        # 1. Create Result File (A) - with images
        # -----------------------------------------
        result_path = f"{output_path_base}_result.xlsx"
        logger.info(f"Creating result file (A): {result_path} with {len(df_finalized)} rows.")
        
        # Create result DataFrame with images
        df_result = df_finalized.copy()
        
        # Create a new workbook for result file
        workbook_result = openpyxl.Workbook()
        worksheet_result = workbook_result.active
        worksheet_result.title = "제품 가격 비교"
        
        # Write data using the helper function
        if not _write_data_to_worksheet(worksheet_result, df_result):
            logger.error("Failed to write data to result worksheet")
            return False, False, None, None
            
        # Apply formatting based on manual requirements
        _apply_basic_excel_formatting(worksheet_result, df_result.columns.tolist())
        _add_hyperlinks_to_worksheet(worksheet_result, df_result, hyperlinks_as_formulas=False)
        _add_header_footer(worksheet_result)
        
        # Process and add images
        image_cols = [col for col in df_result.columns if col in IMAGE_COLUMNS]
        for col in image_cols:
            col_idx = df_result.columns.get_loc(col) + 1
            col_letter = get_column_letter(col_idx)
            worksheet_result.column_dimensions[col_letter].width = 22
            
            for row_idx, value in enumerate(df_result[col], 2):
                if isinstance(value, dict) and 'path' in value:
                    img_path = value['path']
                    if os.path.exists(img_path):
                        try:
                            img = openpyxl.drawing.image.Image(img_path)
                            img.width = 160
                            img.height = 160
                            img.anchor = f"{col_letter}{row_idx}"
                            worksheet_result.add_image(img)
                            worksheet_result.row_dimensions[row_idx].height = 120
                        except Exception as img_err:
                            logger.warning(f"Failed to add image at {img_path}: {img_err}")
        
        # Save result file
        workbook_result.save(result_path)
        result_success = True
        
        # -----------------------------------------
        # 2. Create Upload File (P) - without images, with column mapping
        # -----------------------------------------
        upload_path = f"{output_path_base}_upload.xlsx"
        logger.info(f"Creating upload file (P): {upload_path}")
        
        # Create upload version DataFrame with only URLs
        df_upload = df_finalized.copy()
        
        # Replace image data with web URLs only
        for col in IMAGE_COLUMNS:
            if col in df_upload.columns:
                df_upload[col] = df_upload[col].apply(
                    lambda x: x.get('url', '') if isinstance(x, dict) else 
                    (x if isinstance(x, str) and x.startswith(('http://', 'https://')) else '')
                )
        
        # Create upload workbook
        workbook_upload = openpyxl.Workbook()
        worksheet_upload = workbook_upload.active
        worksheet_upload.title = "제품 가격 비교"
        
        # Write data using the helper function
        if not _write_data_to_worksheet(worksheet_upload, df_upload):
            logger.error("Failed to write data to upload worksheet")
            return result_success, False, result_path, None
            
        # Apply upload-specific formatting
        _apply_basic_excel_formatting(worksheet_upload, df_upload.columns.tolist())
        _add_hyperlinks_to_worksheet(worksheet_upload, df_upload, hyperlinks_as_formulas=True)
        _add_header_footer(worksheet_upload)
        
        # Save upload file
        workbook_upload.save(upload_path)
        upload_success = True
        
        return result_success, upload_success, result_path, upload_path
        
    except Exception as e:
        logger.error(f"Error in create_split_excel_outputs: {e}")
        logger.debug(traceback.format_exc())
        return result_success, upload_success, result_path, upload_path

@safe_excel_operation
def create_final_output_excel(df: pd.DataFrame, output_path: str) -> bool:
    """
    Creates the final Excel output file with all formatting and data.
    
    Args:
        df: DataFrame containing the data to write
        output_path: Path where the Excel file should be saved
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Input validation
        if df is None or df.empty:
            logging.error("Cannot create Excel file: Input DataFrame is empty or None")
            return False
            
        # Check output directory permissions and space
        output_dir = os.path.dirname(output_path)
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except Exception as e:
                logging.error(f"Failed to create output directory {output_dir}: {e}")
                return False
                
        if not os.access(output_dir, os.W_OK):
            logging.error(f"No write permission for output directory: {output_dir}")
            return False
            
        # Check disk space (require at least 100MB free)
        try:
            import shutil
            free_space = shutil.disk_usage(output_dir).free
            if free_space < 100 * 1024 * 1024:  # 100MB
                logging.error(f"Insufficient disk space. Only {free_space/1024/1024:.1f}MB available")
                return False
        except Exception as e:
            logging.warning(f"Could not check disk space: {e}")
            
        # Create backup if file exists
        if os.path.exists(output_path):
            try:
                backup_path = f"{output_path}.bak"
                shutil.copy2(output_path, backup_path)
                logging.info(f"Created backup at: {backup_path}")
            except Exception as e:
                logging.warning(f"Failed to create backup: {e}")
        
        # Check if file is locked
        try:
            if os.path.exists(output_path):
                with open(output_path, 'a') as f:
                    pass
        except PermissionError:
            logging.error(f"File is locked or in use: {output_path}")
            return False
            
        # Estimate memory requirements
        estimated_size = len(df) * len(df.columns) * 100  # Rough estimate
        if estimated_size > 100 * 1024 * 1024:  # 100MB warning threshold
            logging.warning(f"Large file size estimated ({estimated_size/1024/1024:.1f}MB)")
            
        # Process data in chunks if necessary
        chunk_size = 1000
        processed_df = pd.DataFrame()
        
        for i in range(0, len(df), chunk_size):
            chunk = df.iloc[i:i+chunk_size].copy()
            # Process chunk
            chunk = finalize_dataframe_for_excel(chunk)
            processed_df = pd.concat([processed_df, chunk])
            
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write data
            processed_df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Get the worksheet
            worksheet = writer.sheets['Sheet1']
            
            # Apply formatting
            formatter = ExcelFormatter()
            formatter.format_result_file(worksheet, processed_df)
            
            # Process images if present
            image_processor = ImageProcessor()
            images_added = image_processor.process_image_columns(worksheet, processed_df)
            
            if images_added > 0:
                logging.info(f"Added {images_added} images to Excel file")
            
            # Adjust dimensions for images
            image_processor._adjust_dimensions_for_images(worksheet, processed_df)
            
        logging.info(f"Successfully created Excel file at: {output_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error creating Excel file: {e}")
        logging.debug(traceback.format_exc())
        
        # Try to restore from backup if available
        backup_path = f"{output_path}.bak"
        if os.path.exists(backup_path):
            try:
                shutil.copy2(backup_path, output_path)
                logging.info("Restored from backup after error")
            except Exception as restore_error:
                logging.error(f"Failed to restore from backup: {restore_error}")
        
        return False

def _write_data_to_worksheet(worksheet, df_for_excel):
    """Write data to worksheet with proper handling of complex data types"""
    try:
        # Helper function to safely extract URL from complex data structures
        def extract_url_from_complex_value(value):
            """Extract URL from complex dictionary objects or return string representation"""
            # Handle None/NaN values
            if pd.isna(value) or value is None:
                return ""

            # Handle strings
            if isinstance(value, str):
                return value

            # Handle numbers
            if isinstance(value, (int, float)):
                return value
                
            # Handle dictionary values
            if isinstance(value, dict):
                try:
                    # Case 1: Double-nested URL structure {'url': {'url': 'actual_url', ...}}
                    if 'url' in value and isinstance(value['url'], dict) and 'url' in value['url']:
                        return value['url']['url']
                    
                    # Case 2: Nested local_path
                    if 'url' in value and isinstance(value['url'], dict) and 'local_path' in value['url']:
                        return value['url']['local_path']
                    
                    # Case 3: Direct URL {'url': 'actual_url'}
                    elif 'url' in value and isinstance(value['url'], str):
                        return value['url']
                        
                    # Case 4: Local path
                    elif 'local_path' in value and value['local_path']:
                        return value['local_path']
                    
                    # Case 5: Product name
                    elif 'product_name' in value:
                        return f"Product: {value['product_name']}"
                    
                    # Case 6: Source property (haereum, kogift, naver)
                    elif 'source' in value and isinstance(value['source'], str):
                        if 'url' in value or 'local_path' in value:
                            return value.get('url', value.get('local_path', str(value)))
                        return f"Source: {value['source']}"
                    
                    # Default: Convert to string
                    return json.dumps(value, ensure_ascii=False)
                except Exception as dict_error:
                    logger.warning(f"Error extracting from dict: {dict_error}")
                    return str(value)
                    
            # Handle list/tuple
            if isinstance(value, (list, tuple)):
                try:
                    # Try to extract first meaningful item
                    if len(value) > 0:
                        first_item = value[0]
                        if isinstance(first_item, dict) and ('url' in first_item or 'local_path' in first_item):
                            return extract_url_from_complex_value(first_item)
                    return json.dumps(value, ensure_ascii=False)
                except Exception as list_error:
                    logger.warning(f"Error extracting from list: {list_error}")
                    return str(value)
                    
            # Default case
            return str(value)
        
        # Write header
        for col_idx, col_name in enumerate(df_for_excel.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, row in enumerate(df_for_excel.itertuples(), 2):
            for col_idx, value in enumerate(row[1:], 1):  # Skip the index
                try:
                    # Extract clean value for Excel
                    cell_value = extract_url_from_complex_value(value)
                    
                    # Set the cell value
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                except Exception as e:
                    # Log the error but continue with other cells
                    logger.error(f"Error writing cell at row {row_idx}, col {col_idx}: {str(e)}")
                    # Put a placeholder in the cell to avoid further errors
                    worksheet.cell(row=row_idx, column=col_idx, value="-")
        
        return True
    except Exception as e:
        logger.error(f"Error writing data to worksheet: {str(e)}")
        logger.error(f"Error details: {traceback.format_exc()}")
        return False

def flatten_nested_image_dicts(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten nested image dictionaries in DataFrame to simple URL strings or local file paths.
    This makes the data suitable for Excel output.
    """
    if df is None or df.empty:
        return df

    df = df.copy()
    
    # Define image-related columns - use both standard column names and alternative names
    image_cols = [col for col in df.columns if col in [
        # Standard image columns
        '본사 이미지', '고려기프트 이미지', '네이버 이미지',
        # Alternative column names
        '해오름(이미지링크)', '고려기프트(이미지링크)', '네이버쇼핑(이미지링크)'
    ]]
    
    # If no exact matches found, try partial matching
    if not image_cols:
        image_cols = [col for col in df.columns if any(img_type in col.lower() for img_type in ['이미지', 'image'])]
    
    logger.debug(f"Processing image columns for flattening: {image_cols}")
    
    for col in image_cols:
        for idx in df.index:
            value = df.at[idx, col]
            
            # Skip if value is already a string or None
            if isinstance(value, str) or pd.isna(value):
                continue
                
            try:
                # Handle dictionary values
                if isinstance(value, dict):
                    # Try to extract URL or local_path in order of preference
                    result = None
                    
                    # Case 1: Double-nested URL structure {'url': {'url': 'actual_url', ...}}
                    if 'url' in value and isinstance(value['url'], dict):
                        if 'url' in value['url'] and value['url']['url']:
                            result = value['url']['url']
                        elif 'local_path' in value['url'] and value['url']['local_path']:
                            result = value['url']['local_path']
                    
                    # Case 2: Direct URL {'url': 'actual_url'}
                    elif 'url' in value and isinstance(value['url'], str) and value['url']:
                        result = value['url']
                    
                    # Case 3: Local path
                    elif 'local_path' in value and value['local_path']:
                        result = value['local_path']
                        
                    # Case 4: Check for other possible URL fields
                    else:
                        for key in ['image_url', 'original_url', 'src', 'link', 'product_url']:
                            if key in value and isinstance(value[key], str) and value[key]:
                                result = value[key]
                                break
                    
                    # Set the flattened value (keep dictionary if we have path+url that might be useful for Excel)
                    if result and os.path.exists(result):
                        # If it's a local file that exists, keep a simplified dict for Excel embedding
                        df.at[idx, col] = {'url': result, 'local_path': result}
                    else:
                        # Otherwise just use the string value
                        df.at[idx, col] = result if result else '-'
                
                # Handle Series or other iterable types
                elif isinstance(value, (pd.Series, list, tuple)):
                    # Convert to list if it's a Series
                    items = value.tolist() if isinstance(value, pd.Series) else value
                    
                    # Try to find a valid URL in the items
                    result = None
                    for item in items:
                        if isinstance(item, dict):
                            # Apply the same URL extraction logic as above
                            if 'url' in item and isinstance(item['url'], dict) and 'url' in item['url']:
                                result = item['url']['url']
                                break
                            elif 'url' in item and isinstance(item['url'], str):
                                result = item['url']
                                break
                            elif 'local_path' in item and item['local_path']:
                                result = item['local_path']
                                break
                            else:
                                for key in ['image_url', 'original_url', 'src', 'link', 'product_url']:
                                    if key in item and isinstance(item[key], str) and item[key]:
                                        result = item[key]
                                        break
                                if result:
                                    break
                        elif isinstance(item, str) and item.startswith(('http://', 'https://')):
                            result = item
                            break
                    
                    # Set the flattened value
                    if result and os.path.exists(result):
                        # If it's a local file that exists, keep a simplified dict for Excel embedding
                        df.at[idx, col] = {'url': result, 'local_path': result}
                    else:
                        # Otherwise just use the string value
                        df.at[idx, col] = result if result else '-'
                
                else:
                    # For any other type, convert to string
                    df.at[idx, col] = str(value)
                    
            except Exception as e:
                logger.warning(f"Error flattening image data in column {col}, row {idx}: {e}")
                df.at[idx, col] = '-'
    
    return df

def _process_image_columns(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame) -> int:
    # Add image validation
    def validate_image(img_path):
        try:
            if not os.path.exists(img_path):
                return False
            if os.path.getsize(img_path) > 10 * 1024 * 1024:  # 10MB limit
                logging.warning(f"Image too large: {img_path}")
                return False
            with Image.open(img_path) as img:
                img.verify()
            return True
        except Exception as e:
            logging.error(f"Invalid image {img_path}: {e}")
            return False 

def finalize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    # Add column name validation
    def sanitize_column_name(name: str) -> str:
        # Remove invalid Excel characters
        invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
        for char in invalid_chars:
            name = name.replace(char, '_')
        # Limit length to Excel's maximum
        if len(name) > 255:
            name = name[:252] + '...'
        return name
        
    # Handle duplicate columns
    seen_columns = set()
    new_columns = []
    for col in df.columns:
        sanitized = sanitize_column_name(col)
        if sanitized in seen_columns:
            counter = 1
            while f"{sanitized}_{counter}" in seen_columns:
                counter += 1
            sanitized = f"{sanitized}_{counter}"
        seen_columns.add(sanitized)
        new_columns.append(sanitized)
    
    df.columns = new_columns 

def _apply_cell_styles_and_alignment(worksheet: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    # Add error handling for style application
    def safe_apply_style(cell, style_func):
        try:
            style_func(cell)
        except Exception as e:
            logging.warning(f"Failed to apply style to cell {cell.coordinate}: {e}")
            
    # Add value type validation
    def validate_cell_value(value):
        if isinstance(value, (str, int, float, bool, type(None))):
            return value
        return str(value)  # Convert other types to string 

def sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    # Add nested data handling
    def flatten_value(value):
        if isinstance(value, (dict, list)):
            return json.dumps(value)
        return value
        
    # Add numeric precision handling
    def validate_numeric(value):
        if isinstance(value, float):
            if abs(value) > 1e15:  # Excel's limit
                return str(value)
        return value 