import pandas as pd
import os
import sys
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from PythonScript.excel_utils import (
    create_final_output_excel, 
    _add_hyperlinks_to_worksheet as add_hyperlinks,
    _process_image_columns as process_image_cells,
    LINK_COLUMNS_FOR_HYPERLINK as LINK_COLUMN_MAP
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)

def test_excel_basic_writing():
    """Test basic Excel creation with sample data"""
    print("\n=== Testing Basic Excel Writing ===")
    
    # Define column names - these should match exactly with the data structure
    columns = [
        "Code", "중분류카테고리", "상품명", "기본수량(1)", "판매단가(V포함)", "본사상품링크", 
        "기본수량(2)", "판매가(V포함)(2)", "판매단가(V포함)(2)", "가격차이(2)", "가격차이(2)(%)", "고려기프트 상품링크", 
        "기본수량(3)", "판매단가(V포함)(3)", "가격차이(3)", "가격차이(3)(%)", "공급사명", "네이버 쇼핑 링크", "공급사 상품링크", 
        "본사 이미지", "고려기프트 이미지", "네이버 이미지"
    ]
    
    # Parse the provided sample data directly into a DataFrame
    data = {
        "Code": ["A", "A"],
        "중분류카테고리": ["가방(에코백/면)", "어린이우산"],
        "상품명": ["행운이 네잎클로버 투포켓 에코백", "캐치티니핑 53 스무디 입체리본 투명 아동우산"],
        "기본수량(1)": ["200", "50"],
        "판매단가(V포함)": ["2970", "17820"],
        "본사상품링크": ["http://www.jclgift.com/product/product_view.asp?p_idx=437570", "http://www.jclgift.com/product/product_view.asp?p_idx=437551"],
        "기본수량(2)": ["200", "50"],
        "판매가(V포함)(2)": ["570900", "842600"],
        "판매단가(V포함)(2)": ["2854.5", "16852"],
        "가격차이(2)": ["-115.5", "-968"],
        "가격차이(2)(%)": ["-3.9", "-5.4"],
        "고려기프트 상품링크": ["http://koreagift.com/ez/mall.php?cat=003011001&query=view&no=170297", "http://koreagift.com/ez/mall.php?cat=004002005&query=view&no=170277"],
        "기본수량(3)": ["", ""],
        "판매단가(V포함)(3)": ["", "14490"],
        "가격차이(3)": ["", "-3330"],
        "가격차이(3)(%)": ["", "-18.6"],
        "공급사명": ["", "네이버"],
        "네이버 쇼핑 링크": ["", "https://search.shopping.naver.com/catalog/53165134501"],
        "공급사 상품링크": ["", ""],
        "본사 이미지": ["http://i.jclgift.com/upload/product/bimg3/BBCH0009421b.png", "http://i.jclgift.com/upload/product/bimg3/LLAG0003250b.jpg"],
        "고려기프트 이미지": ["http://koreagift.com/ez/upload/mall/shop_1744178312138728_0.png", "http://koreagift.com/ez/upload/mall/shop_1744109588135407_0.jpg"],
        "네이버 이미지": ["", "https://shopping-phinf.pstatic.net/main_5316513/53165134501.20250222203926.jpg"]
    }
    
    # Create DataFrame from the dictionary
    df = pd.DataFrame(data)
    
    # Display DataFrame information
    print(f"DataFrame created with {len(df)} rows and {len(df.columns)} columns")
    
    # Create output directory if it doesn't exist
    output_dir = "test_excel_output"
    os.makedirs(output_dir, exist_ok=True)
    print(f"Output directory created: {output_dir}")
    
    # Call the Excel writing function
    try:
        output_file = os.path.join(output_dir, "excel_test_basic.xlsx")
        success = create_final_output_excel(
            df, 
            output_file
        )
    
        if success and os.path.exists(output_file):
            print(f"Excel file successfully created: {output_file}")
            print(f"File size: {os.path.getsize(output_file)} bytes")
            return output_file
        else:
            print("Failed to create Excel file or file not found")
            return None
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return None


def custom_excel_styling(file_path):
    """Custom function to apply Excel styling, avoiding the issues in the main function"""
    try:
        print("Applying custom Excel styles...")
        
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Define styles
        header_style = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Styling {max_row} rows and {max_col} columns")
        
        # Apply header styling
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_style
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply borders and alignment to all cells
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Set column widths
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            # Default width for most columns
            ws.column_dimensions[col_letter].width = 15
            
            # Custom widths for specific columns
            header = ws.cell(row=1, column=col).value
            if header:
                if '상품명' in str(header):
                    ws.column_dimensions[col_letter].width = 30
                elif '링크' in str(header):
                    ws.column_dimensions[col_letter].width = 20
                elif '이미지' in str(header):
                    ws.column_dimensions[col_letter].width = 25
        
        # Highlight price difference cells
        for row in range(2, max_row + 1):
            # Check 가격차이(2) column
            price_diff_col2 = None
            price_diff_col3 = None
            
            # Find the column indices for price difference columns
            for col in range(1, max_col + 1):
                header = ws.cell(row=1, column=col).value
                if header == '가격차이(2)':
                    price_diff_col2 = col
                elif header == '가격차이(3)':
                    price_diff_col3 = col
            
            # Highlight negative price differences (good) in yellow
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Apply highlighting to price difference columns
            if price_diff_col2:
                cell = ws.cell(row=row, column=price_diff_col2)
                cell_value = cell.value
                if cell_value and cell_value != '-':
                    # Try to convert to number and check if negative
                    try:
                        value = float(str(cell_value).replace(',', '').replace('-', '-'))
                        if value < 0:
                            cell.fill = yellow_fill
                    except ValueError:
                        pass
                        
            if price_diff_col3:
                cell = ws.cell(row=row, column=price_diff_col3)
                cell_value = cell.value
                if cell_value and cell_value != '-':
                    # Try to convert to number and check if negative
                    try:
                        value = float(str(cell_value).replace(',', '').replace('-', '-'))
                        if value < 0:
                            cell.fill = yellow_fill
                    except ValueError:
                        pass
        
        # Save the workbook
        wb.save(file_path)
        print(f"Custom styling applied successfully to {file_path}")
        return True
        
    except Exception as e:
        print(f"Error applying custom styles: {e}")
        return False


def test_excel_styling_and_links(excel_file):
    """Test applying styles and hyperlinks to an existing Excel file"""
    print("\n=== Testing Excel Styling and Links ===")
    
    if not excel_file or not os.path.exists(excel_file):
        print("No Excel file to style")
        return False
    
    # Test adding hyperlinks first (this seems to work)
    print("Adding hyperlinks...")
    links_result = add_hyperlinks(excel_file, LINK_COLUMN_MAP)
    if links_result:
        print(f"Successfully added hyperlinks to {excel_file}")
    else:
        print(f"Failed to add hyperlinks to {excel_file}")
    
    # Test applying custom styles
    styling_result = custom_excel_styling(excel_file)
    if styling_result:
        print(f"Successfully applied custom styles to {excel_file}")
    else:
        print(f"Failed to apply custom styles to {excel_file}")
    
    # Check file after modifications
    if os.path.exists(excel_file):
        print(f"Final file size after processing: {os.path.getsize(excel_file)} bytes")
        return True
    else:
        print("File not found after processing attempts")
        return False

def main():
    """Main test function"""
    print("Starting Excel writing tests...")
    
    # Test 1: Basic Excel creation
    excel_file = test_excel_basic_writing()
    
    # Test 2: Styling and hyperlinks
    if excel_file:
        test_excel_styling_and_links(excel_file)
    
    print("\nAll tests completed.")

if __name__ == "__main__":
    main() 